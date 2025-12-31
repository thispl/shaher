import frappe
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import requests
from frappe.utils import (getdate, cint, add_months, date_diff, add_days,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)

@frappe.whitelist()
def download_vehicle_register():
    filename = "Vehicle Register"
    xlsx_file = make_vehicle_xlsx()
    frappe.response["filename"] = filename + ".xlsx"
    frappe.response["filecontent"] = xlsx_file.getvalue()
    frappe.response["type"] = "binary"



def get_vehicle_data(args):
    filters = {}
    company = args.get("company")
    if company:
        filters["custom_company"] = company  
    
    vehicles = frappe.get_all(
        "Vehicle",
        filters=filters,
        fields=[
            "location", "custom_project", "name", "make",
            "custom_vehicle_category", "chassis_no",
            "custom_rop_due", "custom_ras_due", "custom_ivms_certificate_validity", "custom_voc_validity", "custom_lifting_tpi_validity",
            "custom_speed_limiter__roll_over_cage", "custom_speed_limiter__roll_over_cage", "custom_remark", "custom_model_year", "custom_engine_no", "custom_fire_extinguisher_next_due_date"
        ]
    )

    data = []
    for i, v in enumerate(vehicles, start=1):
        row = [
            i,
            v.location,
            v.custom_project,
            v.name,
            v.make,
            v.custom_model_year,
            v.custom_vehicle_category,
            v.custom_engine_no,
            v.chassis_no,
            frappe.utils.formatdate(v.custom_fire_extinguisher_next_due_date),
            frappe.utils.formatdate(v.custom_rop_due),
            frappe.utils.formatdate(v.custom_ras_due),
            frappe.utils.formatdate(v.custom_ivms_certificate_validity),
            frappe.utils.formatdate(v.custom_voc_validity),
            frappe.utils.formatdate(v.custom_lifting_tpi_validity),
            frappe.utils.formatdate(v.custom_speed_limiter__roll_over_cage),
            frappe.utils.formatdate(v.custom_speed_limiter__roll_over_cage),
            v.custom_remark
        ]
        data.append(row)
    return data

def make_vehicle_xlsx(sheet_name="Vehicle Register"):
    args = frappe.local.form_dict
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    data = get_vehicle_data(args)

    pink = PatternFill(start_color="F7D7C4", end_color="F7D7C4", fill_type="solid")
    blue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=ws.max_column if ws.max_column>1 else 18)
    ws["A1"] = "HSE VEHICLE REGISTER"
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].font = Font(bold=True, size=16)


    headers = [
        "S.No", "SITE LOCATION", "PROJECT/CONTRACT NO", "VEHICLE REG NO ", "MAKE",
        "MODEL YEAR", "TYPE OF VEHICLE", "ENGINE NO", "CHASSIS NO",
        "FIRE EXTINGUISHER NEXT DUE DATE", "ROP NEXT DUE DATE", "RAS NEXT DUE DATE ",
        "IVMS CERTIFICATE RENEWAL NEXT DUE DATE ", "VOC CERTIFICATE RENEWAL NEXT DUE DATE ",
        "LIFTING TPI NEXT DUE DATE", "ROLL OVER CAGE", "SPEED LIMITER", "REMARKS (IF ANY)"
    ]

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col)
        cell.value = h
        cell.font = Font(bold=True)
        cell.fill = blue
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, row in enumerate(data, start=4):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if c_idx == ws.max_column:
                cell.alignment = Alignment(wrap_text=True, vertical="top")


    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 15
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 20
    ws.column_dimensions["J"].width = 20
    ws.column_dimensions["K"].width = 20
    ws.column_dimensions["L"].width = 20
    ws.column_dimensions["M"].width = 20
    ws.column_dimensions["N"].width = 20
    ws.column_dimensions["O"].width = 20
    ws.column_dimensions["P"].width = 20
    ws.column_dimensions["Q"].width = 20
    ws.column_dimensions["R"].width = 50
    ws.row_dimensions[3].height = 25

    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for r in range(3, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file