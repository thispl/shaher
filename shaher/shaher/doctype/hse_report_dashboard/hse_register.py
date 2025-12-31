from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils import get_column_letter
from io import BytesIO
import frappe
from datetime import datetime, timedelta

@frappe.whitelist()
def download():
    filename = "HSE TRAINING MATRIX"
    file = make_hse_xlsx(filename)
    frappe.response["filename"] = filename + ".xlsx"
    frappe.response["filecontent"] = file.getvalue()
    frappe.response["type"] = "binary"


def make_hse_xlsx(sheet_name):

    detail_fields = [
        "S No","Full Name","Designation","Employee No",
        "Site Location","Date of Birth","Resident Card No",
        "Email ID","Contact Number","Resident Card No Expiry Date",
        "ROP License Expiry Date","branch / Contract No",
        "HSE Passport Issued Institute Name","HSE Passport Number"
    ]

    course_groups = frappe.get_all("Course Group", fields=["name"])
    group_courses = {}

    for grp in course_groups:
        courses = frappe.get_all(
            "Course",
            filters={"course_group": grp.name},
            fields=["course_name"],
            order_by="course_name asc"
        )
        group_courses[grp.name] = [c.course_name for c in courses]

    all_courses = []
    for group, courses in group_courses.items():
        all_courses.extend(courses)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    yellow   = PatternFill("solid", fgColor="FFC000")
    blue     = PatternFill("solid", fgColor="BDD7EE")
    border   = Border(left=Side(style='thin'), right=Side(style='thin'),
                      top=Side(style='thin'), bottom=Side(style='thin'))
    title_al = Alignment(horizontal='center', vertical='center')
    wrap_al  = Alignment(horizontal='center', vertical='center', wrap_text=True)

    total_cols = len(detail_fields) + len(all_courses)

    
    ws.append(["HSE TRAINING REGISTER"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws["A1"].fill = yellow
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = title_al

    
    row2 = ["DETAILS"] + [""] * (len(detail_fields) - 1)

    for group, courses in group_courses.items():
        if courses:
            row2.append(group)
            row2.extend([""] * (len(courses) - 1))

    ws.append(row2)

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=len(detail_fields))

    col_start = len(detail_fields) + 1
    for group, courses in group_courses.items():
        if courses:
            col_end = col_start + len(courses) - 1
            ws.merge_cells(start_row=2, start_column=col_start,
                           end_row=2, end_column=col_end)
            col_start = col_end + 1

    

    yellow = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    title_al = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            if row == 2:
                c.fill = yellow
                c.font = Font(bold=True)
                c.alignment = title_al
            c.border = thin_border
    

    
    row3 = detail_fields[:] + all_courses[:]
    ws.append(row3)

    for col in range(1, total_cols + 1):
        c = ws.cell(row=3, column=col)
        c.fill = blue
        c.font = Font(bold=True)
        c.alignment = wrap_al
        c.border = border
        ws.column_dimensions[get_column_letter(col)].width = 25

    
    training_map = get_training_matrix_latest()

   
    employees = get_filtered_employees()
    serial_no = 1

    for emp in employees:
        # frappe.log_error(emp.name)
        dob = emp.date_of_birth.strftime('%d-%m-%Y') if emp.date_of_birth else ''
        resident_expiry = emp.custom_resident_id_expiry_date.strftime('%d-%m-%Y') if emp.custom_resident_id_expiry_date else ''
        visa_expiry = emp.custom_visa_expiry_date.strftime('%d-%m-%Y') if emp.custom_visa_expiry_date else ''
        row = [
            serial_no,
            emp.employee_name,
            emp.designation,
            emp.name,
            emp.custom_site_location,
            dob,
            emp.custom_resident_id_numberqid_number,
            emp.user_id,
            emp.cell_number,
            resident_expiry,
            visa_expiry,
            emp.branch,
            emp.issued_institute_name,
            emp.hse_passport_number,
        ]

        for course in all_courses:
            key = (emp.name, course)
            expiry = training_map.get(key, "")
            row.append(expiry)

        ws.append(row)
        serial_no += 1
    red_font    = Font(color="FF0000", bold=True)      
    orange_font = Font(color="FFA500", bold=True)     

    today = datetime.today().date()
    warning_date = today + timedelta(days=30)

    start_row = 4
    start_col = 15  

    for row_idx in range(start_row, ws.max_row + 1):
        for col_idx in range(start_col, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            value = cell.value

            if not value:
                continue

            try:
                expiry = datetime.strptime(value, "%d-%m-%Y").date()
            except:
                continue

            if expiry < today:
                cell.font = red_font

            elif today <= expiry <= warning_date:
                cell.font = orange_font
    ws.column_dimensions['A'].width = 6
    cols=['D','E','F','G','I','K']
    for c in cols:
        ws.column_dimensions[c].width = 15
    ws.column_dimensions['H'].width = 30
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = ws.dimensions

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def get_training_matrix_latest():

    rows = frappe.db.sql("""
        SELECT
            parent AS employee,
            course_name,
            MAX(STR_TO_DATE(expiry_date, '%Y-%m-%d')) AS latest_expiry
        FROM `tabHSE Training Matrix`
        GROUP BY employee, course_name
    """, as_dict=True)

    out = {}
    for r in rows:
        dt = r.latest_expiry
        if dt:   
            formatted = dt.strftime('%d-%m-%Y')
        else:
            formatted = ""

        out[(r.employee, r.course_name)] = formatted

    return out


def get_filtered_employees():
    company    = frappe.local.form_dict.get("company")
    division   = frappe.local.form_dict.get("division")
    from_date  = frappe.local.form_dict.get("from_date")
    to_date    = frappe.local.form_dict.get("to_date")

    filters  = []
    params   = []

    if company:
        filters.append("e.company = %s")
        params.append(company)

    if division:
        filters.append("e.custom_division = %s")
        params.append(division)

    where_clause = " AND ".join(filters) if filters else "1=1"

    if from_date and to_date:
        # frappe.log_error('Filters',where_clause)
        # frappe.log_error('Filters2',params)
        return frappe.db.sql("""
            SELECT
                e.name,
                e.employee_name,
                e.designation,
                e.custom_site_location,
                e.date_of_birth,
                e.custom_resident_id_numberqid_number,
                e.user_id,
                e.cell_number,
                e.custom_resident_id_expiry_date,
                e.custom_visa_expiry_date,
                e.branch,
                ed.issued_institute_name,
                ed.hse_passport_number
            FROM `tabEmployee` e
            INNER JOIN `tabHSE` ed ON ed.employee = e.name
            INNER JOIN `tabHSE Training Matrix` tm ON tm.parent = ed.name
            WHERE
                {where_clause}
                AND STR_TO_DATE(tm.expiry_date, '%%Y-%%m-%%d') BETWEEN %s AND %s
            GROUP BY e.name
            ORDER BY e.employee_name
        """.format(where_clause=where_clause), params + [from_date, to_date], as_dict=True)


    return frappe.db.sql(f"""
        SELECT
            e.name,
            e.employee_name,
            e.designation,
            e.custom_site_location,
            e.date_of_birth,
            e.custom_resident_id_numberqid_number,
            e.user_id,
            e.cell_number,
            e.custom_resident_id_expiry_date,
            e.custom_visa_expiry_date,
            e.branch,
            ed.issued_institute_name,
            ed.hse_passport_number
        FROM `tabEmployee` e
        INNER JOIN `tabHSE` ed ON ed.employee = e.name
        WHERE
            e.status = 'Active'
            AND {where_clause}
        ORDER BY e.employee_name
    """, params, as_dict=True)
