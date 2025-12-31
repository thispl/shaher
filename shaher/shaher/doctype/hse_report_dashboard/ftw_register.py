import frappe
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

@frappe.whitelist()
def download_hse_excel():
    filename = "FTW Register.xlsx"
    filedata = make_hse_excel()

    frappe.response["filename"] = filename
    frappe.response["filecontent"] = filedata.getvalue()
    frappe.response["type"] = "binary"


def make_hse_excel():
    employees = get_filtered_employees()

    wb = Workbook()
    ws = wb.active
    ws.title = "HSE Register"
    ws.append(['HSE FITNESS TO WORK MEDICAL EXAMINATION REGISTER','','','','','','','','','','','',''])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=13)
    headers = [
        "S.No", "Employee Name", "Employee ID", "Designation","Site Location",
        "Project", "Date of Birth", "Age",
        "Medical Test Done On", "Next Due Date", "Result",
        "Chronic Disease", "Doctor Remarks"
    ]
    ws.append(headers)

    fill_yellow = PatternFill(start_color="9db7e0", end_color="9db7e0", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bold = Font(bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_num in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col_num)
        c.fill = fill_yellow
        c.font = bold
        c.alignment = center
        c.border = border
    for col_num in range(1, len(headers) + 1):
        c = ws.cell(row=2, column=col_num)
        c.fill = fill_yellow
        c.font = bold
        c.alignment = center
        c.border = border
    widths = [6, 25, 15, 18, 15, 25, 15, 7, 15, 15, 15, 20, 40]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = width

    # ws.freeze_panes = "A2"
    # ws.auto_filter.ref = ws.dimensions
    
    sn = 1
    today = datetime.today().date()

    for emp in employees:
        dob = emp.date_of_birth.strftime("%d-%m-%Y") if emp.date_of_birth else ""
        age = today.year - emp.date_of_birth.year if emp.date_of_birth else ""

        ftw = emp.get("ftw") or frappe._dict()

        row = [
            sn,
            emp.employee_name,
            emp.name,
            emp.designation,
            emp.custom_site_location,
            emp.project__contract_no or "",
            dob,
            age,
            format_date(emp.medical_test_done_on),
            format_date(emp.next_due_date),
            emp.results__fitunfit or "",
            emp.chronic_disease or "",
            emp.remarks_from_doctor or "",
        ]

        ws.append(row)
        sn += 1
    for row in ws.iter_rows(min_col=13, max_col=13):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.border = border
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def format_date(value):
    if not value:
        return ""
    try:
        return datetime.strptime(str(value), "%Y-%m-%d").strftime("%d-%m-%Y")
    except:
        return value


def get_filtered_employees():

    company    = frappe.local.form_dict.get("company")
    division   = frappe.local.form_dict.get("division")
    from_date  = frappe.local.form_dict.get("from_date")
    to_date    = frappe.local.form_dict.get("to_date")

    filters  = []
    params   = []

    if company:
        filters.append("e.company = %s")
        params.append(company)

    if division:
        filters.append("e.custom_division = %s")
        params.append(division)

    where_clause = " AND ".join(filters) if filters else "1=1"

    if from_date and to_date:
        # frappe.log_error('Filters',where_clause)
        # frappe.log_error('Filters2',params)
        return frappe.db.sql("""
            SELECT
                e.employee_name,
                e.name,
                e.designation,
                e.custom_site_location,
                ed.project__contract_no,
                e.date_of_birth,
                ftw.medical_test_done_on,
                ftw.next_due_date,
                ftw.results__fitunfit,
                ftw.chronic_disease,
                ftw.remarks_from_doctor
            FROM `tabEmployee` e
            INNER JOIN `tabHSE` ed ON ed.employee = e.name
            LEFT JOIN `tabFTW Register` ftw ON ftw.parent = ed.name  
            WHERE
                {where_clause}
                AND STR_TO_DATE(ftw.medical_test_done_on, '%%Y-%%m-%%d') BETWEEN %s AND %s
            GROUP BY e.name
            ORDER BY e.employee_name
        """.format(where_clause=where_clause), params + [from_date, to_date], as_dict=True)
    return frappe.db.sql("""
            SELECT
                e.employee_name,
                e.name,
                e.designation,
                e.custom_site_location,
                ed.project__contract_no,
                e.date_of_birth,
                ftw.medical_test_done_on,
                ftw.next_due_date,
                ftw.results__fitunfit,
                ftw.chronic_disease,
                ftw.remarks_from_doctor
            FROM `tabEmployee` e
            INNER JOIN `tabHSE` ed ON ed.employee = e.name
            LEFT JOIN `tabFTW Register` ftw ON ftw.parent = ed.name  
            WHERE
                {where_clause}
            GROUP BY e.name
            ORDER BY e.employee_name
        """.format(where_clause=where_clause), params, as_dict=True)
    