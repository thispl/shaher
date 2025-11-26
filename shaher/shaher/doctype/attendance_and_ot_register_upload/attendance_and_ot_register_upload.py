# Copyright (c) 2025, gifty.p@groupteampro.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from openpyxl import Workbook
import frappe
import io
from frappe import _
import openpyxl
from openpyxl import Workbook
import calendar
from datetime import datetime, date
from frappe.utils import getdate, date_diff, add_days

import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
from frappe.utils import get_site_path



class AttendanceandOTRegisterUpload(Document):
    pass


@frappe.whitelist()
def get_template():
    filename = 'Attendance and OT Register Upload'
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary' 

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
def make_xlsx(sheet_name=None):
    args = frappe.local.form_dict
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    headers = [
        "Employee Code", "Employee Name", "Division", "Month End Date", 
        "No of Days Worked", "Normal OT HRS","Night OT","WO/NH", "Food Allowance", "Other Additions",
        "Salary Advanced", "Other Deductions", "Remarks"
    ]

    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
        employees = frappe.db.sql("""
            SELECT name, employee_name, status, custom_site_location, relieving_date, date_of_joining
            FROM `tabEmployee`
            WHERE company = %(company)s
            AND date_of_joining <= %(to_date)s
            AND (
                    status = 'Active'
                OR (status = 'Left' AND relieving_date BETWEEN %(from_date)s AND %(to_date)s)
            )
        """, args, as_dict=True)
        if args.get('location'):
            employees = frappe.db.sql("""
                SELECT name, employee_name, status, custom_site_location, relieving_date, date_of_joining
                FROM `tabEmployee`
                WHERE company = %(company)s
                AND (%(location)s IS NULL OR custom_site_location = %(location)s)
                AND date_of_joining <= %(to_date)s
                AND (
                        status = 'Active'
                    OR (status = 'Left' AND relieving_date BETWEEN %(from_date)s AND %(to_date)s)
                )
            """, args, as_dict=True)

        today = datetime.today()
        year = today.year
        month = today.month

        
        if month == 1:
            previous_month = 12
            year -= 1
        else:
            previous_month = month - 1

        
        days_in_prev_month = calendar.monthrange(year, previous_month)[1]
        
        last_date = date(year, previous_month, days_in_prev_month)
        
        formatted_date = last_date.strftime("%d-%m-%Y")



    
        for row_num, emp in enumerate(employees, start=2):
            ws.cell(row=row_num, column=1).value = emp.name                    
            ws.cell(row=row_num, column=2).value = emp.employee_name          
            ws.cell(row=row_num, column=3).value = emp.custom_site_location          
            ws.cell(row=row_num, column=4).value = formatted_date          
            ws.cell(row=row_num, column=5).value = days_in_prev_month          

        


    xlsx_file = io.BytesIO()
    wb.save(xlsx_file)
    return xlsx_file
 



@frappe.whitelist()
def excel_preview(file_url):
    import frappe
    import openpyxl
    from frappe.utils.file_manager import get_file

    if not file_url:
        return ""

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = frappe.get_site_path("public", file_doc.file_url.replace("/files/", "files/"))

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    html = """
    <div style="overflow-x: auto; max-width: 100%;">
    <p style="font-size:15px;"><b>Upload Preview</b></p>
    <table style='border-collapse: collapse; width: max-content;'>
        <thead>
    """

    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        html += "<tr>"

       
        if row_count == 1:
            html += "<th style='background-color: #e8262e; color: white; text-align: center; padding: 5px; border: 1px solid white;'>S.No</th>"
        else:
            html += "<td style='text-align: center; padding: 5px; border: 1px solid  #e8262e;'>{}</td>".format(row_count - 1)

        for cell in row:
            is_number = isinstance(cell, (int, float))
            align_style = "text-align: right;" if is_number else "text-align: left;"

            if row_count == 1:
                html += f"<th style='background-color:#e8262e; color: white; border: 1px solid  white; {align_style} padding: 5px; white-space: nowrap;'>{cell if cell is not None else ''}</th>"
            else:
              
                html += f"<td style='white-space: nowrap; padding: 5px; {align_style} border: 1px solid  #e8262e;'>{cell if cell is not None else ''}</td>"

        html += "</tr>"

        if row_count == 1:
            html += "</thead><tbody>"

    html += "</tbody></table></div>"
    return html


@frappe.whitelist()
def create_doc(file_url, from_date, to_date, doc_name=None, submit=False):
    import openpyxl
    from frappe.utils import get_site_path

    header_map = {
        "Employee Code": "employee",
        "Employee Name": "employee_name",
        "Division":"division",
        "No of Days Worked":"no_of_days_worked",
        "Food Allowance":"food_allowance",
        "Other Additions":"other_earnings",
        "Other Deductions":"other_deduction",
        "Salary Advanced":"salary_advance_deduction",
        "Remarks":"remarks",
        "Normal OT HRS": "ot_hours",
        "Night OT":"night_ot",
        "WO/NH":"wo_nh"
       
    }

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = get_site_path("public", file_doc.file_url.replace("/files/", "files/"))

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = []
    created_docs = []
    # Collect all employee codes from Excel before inserting docs
    employee_codes = []
    not_exist_employees = []
    missing_doj = []
    joining_after_period = []
    left_before_period = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [header_map.get(str(cell).strip(), "") for cell in row]
            continue
        employee_code = row[headers.index("employee")] if "employee" in headers else None
        if employee_code:
            employee_codes.append(employee_code)
            
            # Fetch employee details
            emp_doc = frappe.db.get_value('Employee', {'name': employee_code}, ['name', 'date_of_joining', 'relieving_date'], as_dict=True)
            
            if not emp_doc or not emp_doc.name:
                not_exist_employees.append(employee_code)
                continue
            
            if not emp_doc.date_of_joining:
                missing_doj.append(employee_code)
            else:
                if date_diff(getdate(to_date), getdate(emp_doc.date_of_joining)) < 0:
                    joining_after_period.append(employee_code)
            
            if emp_doc.relieving_date and date_diff(getdate(emp_doc.relieving_date), getdate(from_date)) < 0:
                left_before_period.append(employee_code)
    
    employee_codes_tuple = tuple(employee_codes) if employee_codes else ("",)

    # Check for overlapping Attendance & OT Registers
    duplicates = frappe.db.sql("""
        SELECT employee, name
        FROM `tabAttendance and OT Register`
        WHERE docstatus != 2
        AND employee IN %s
        AND from_date <= %s
        AND to_date >= %s
    """, (employee_codes_tuple, to_date, from_date), as_dict=True)
    
    messages = []

    if not_exist_employees:
        messages.append(
            "Employees not found in system: " + ", ".join(map(str, not_exist_employees))
        )

    if missing_doj:
        messages.append("Employees missing Date of Joining: " + ", ".join(missing_doj))
    if joining_after_period:
        messages.append("Employees joining after Payroll Period: " + ", ".join(joining_after_period))
    if left_before_period:
        messages.append("Employees left before Payroll Period: " + ", ".join(left_before_period))
    if duplicates:
        dup_msgs = [
            f"Employee: {dup['employee']}, Attendance and OT Register: <a href='/app/attendance-and-ot-register/{dup['name']}' target='_blank'>{dup['name']}</a>"
            for dup in duplicates
        ]
        messages.append("Duplicate  Attendance and OT Register entries found:<br>" + "<br>".join(dup_msgs))
    
    if messages:
        frappe.throw("<br>".join(messages))
    
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
          
            headers = [header_map.get(str(cell).strip(), "") for cell in row]
            continue

        doc_data = {}
        for j, cell in enumerate(row):
            fieldname = headers[j]
            if fieldname:
                doc_data[fieldname] = cell

        employee_code = doc_data.get("employee")
        if not employee_code or str(employee_code).strip() == "":
            frappe.logger().info(f"Skipping row {i+1} — missing employee code.")
            continue

        doc = frappe.new_doc("Attendance and OT Register")
        doc.from_date = from_date
        doc.to_date = to_date
        doc.update(doc_data)
        doc.attendance_and_ot_register_upload = doc_name 

        if submit in ("1", "true", "True", True):
            doc.insert()
            doc.submit()
        else:
            doc.insert()

        created_docs.append(doc.name)

    return created_docs

@frappe.whitelist()
def create(file_url=None, from_date=None, to_date=None, doc_name=None, submit=False):
    import openpyxl
    from frappe.utils import get_site_path

    header_map = {
        "Employee Code": "employee",
        "Employee Name": "employee_name",
        "Division":"division",
        "No of Days Worked":"no_of_days_worked",
        "Food Allowance":"food_allowance",
        "Other Additions":"other_earnings",
        "Other Deductions":"other_deduction",
        "Salary Advanced":"salary_advance_deduction",
        "Remarks":"remarks",
        "Normal OT HRS": "ot_hours",
        "Night OT":"night_ot",
        "WO/NH":"wo_nh"
       
    }

    file_doc = frappe.get_doc("File", {"file_url": file_url})
    file_path = get_site_path("public", file_doc.file_url.replace("/files/", "files/"))

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    headers = []
    created_docs = []
    employee_codes = []
    not_exist_employees = []
    missing_doj = []
    joining_after_period = []
    left_before_period = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [header_map.get(str(cell).strip(), "") for cell in row]
            continue
        employee_code = row[headers.index("employee")] if "employee" in headers else None
        if employee_code:
            employee_codes.append(employee_code)
            
            # Fetch employee details
            emp_doc = frappe.db.get_value('Employee', {'name': employee_code}, ['name', 'date_of_joining', 'relieving_date'], as_dict=True)
            
            if not emp_doc or not emp_doc.name:
                not_exist_employees.append(employee_code)
                continue
            
            if not emp_doc.date_of_joining:
                missing_doj.append(employee_code)
            else:
                if date_diff(getdate(to_date), getdate(emp_doc.date_of_joining)) < 0:
                    joining_after_period.append(employee_code)
            
            if emp_doc.relieving_date and date_diff(getdate(emp_doc.relieving_date), getdate(from_date)) < 0:
                left_before_period.append(employee_code)
    
    employee_codes_tuple = tuple(employee_codes) if employee_codes else ("",)

    # Check for overlapping Attendance & OT Registers
    duplicates = frappe.db.sql("""
        SELECT employee, name
        FROM `tabAttendance and OT Register`
        WHERE docstatus != 2
        AND employee IN %s
        AND from_date <= %s
        AND to_date >= %s
    """, (employee_codes_tuple, to_date, from_date), as_dict=True)
    
    messages = []

    if not_exist_employees:
        messages.append(
            "Employees not found in system: " + ", ".join(map(str, not_exist_employees))
        )

    if missing_doj:
        messages.append("Employees missing Date of Joining: " + ", ".join(missing_doj))
    if joining_after_period:
        messages.append("Employees joining after Payroll Period: " + ", ".join(joining_after_period))
    if left_before_period:
        messages.append("Employees left before Payroll Period: " + ", ".join(left_before_period))
    if duplicates:
        dup_msgs = [
            f"Employee: {dup['employee']}, Attendance and OT Register: <a href='/app/attendance-and-ot-register/{dup['name']}' target='_blank'>{dup['name']}</a>"
            for dup in duplicates
        ]
        messages.append("Duplicate and OT Register entries found:<br>" + "<br>".join(dup_msgs))
    
    if messages:
        frappe.throw("<br>".join(messages))
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
         
            headers = [header_map.get(str(cell).strip(), "") for cell in row]
            continue

        doc_data = {}
        for j, cell in enumerate(row):
            fieldname = headers[j]
            if fieldname:
                doc_data[fieldname] = cell

        employee_code = doc_data.get("employee")
        if not employee_code or str(employee_code).strip() == "":
            frappe.logger().info(f"Skipping row {i+1} — missing employee code.")
            continue

        doc = frappe.new_doc("Attendance and OT Register")
        doc.from_date = from_date
        doc.to_date = to_date
        doc.attendance_and_ot_register_upload = doc_name 
        doc.update(doc_data)

        if submit in ("1", "true", "True", True):
            doc.insert()
           
        else:
            doc.insert()

        created_docs.append(doc.name)

    return created_docs