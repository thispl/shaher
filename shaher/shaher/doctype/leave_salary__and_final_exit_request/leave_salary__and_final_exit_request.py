# Copyright (c) 2025, gifty.p@groupteampro.com and contributors
# For license information, please see license.txt
from __future__ import unicode_literals
import frappe,erpnext
from frappe.utils import cint
from frappe.utils import flt, fmt_money
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today, time_diff_in_hours

from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta

import openpyxl
from collections import defaultdict
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from frappe.utils import date_diff,today,cstr
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from frappe.utils import formatdate


from erpnext.setup.utils import get_exchange_rate
from frappe import throw,_, db, get_doc, throw, whitelist


from frappe.utils import (
	add_days,
	cint,
	cstr,
	date_diff,
	flt,
	formatdate,
	get_first_day,
	get_link_to_form,
	getdate,
	money_in_words,
	rounded,
)
from frappe.model.document import Document


class LeaveSalaryandFinalExitRequest(Document):
	pass

#Update the leave count(Annual Leave and LOP) between the given Last Rejoin Date and Last Date of Service in leave salary
@frappe.whitelist()
def leave_count(start_date,end_date,employee):
	count = 0
	start_date = str(start_date)
	end_date = str(end_date)
	attendance = frappe.get_all(
	"Attendance",
		filters={
			"employee": employee,
			"attendance_date": ["between", (start_date, end_date)],
			"docstatus": ["!=", 2],
			"leave_type": ["in", ["Annual Leave"]]
		},
		fields=["status"]
	)

	for i in attendance:
		if i.status =="On Leave":
			count +=1
		if i.status =="Half Day":
			count+=0.5
	return count

#If the Given employee role is HOD , HOD Checkbox has been enabled
@frappe.whitelist()
def get_role(employee):
	user_id = frappe.get_value('Employee',{'employee':employee},['user_id'])
	hod = frappe.get_value('User',{'email':user_id},['name'])
	role = "HOD"
	hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
	if hod:
		return  "HI"
	else:
		return "HII"

#Update the Rejoining Date from Rejoining Form while open the leave salary
@frappe.whitelist()
def update_start_date(employee):
	value = frappe.db.get_value("Rejoining Form",{"employee_id":employee,"workflow_state":"Approved"},["re_join"])
	return value

