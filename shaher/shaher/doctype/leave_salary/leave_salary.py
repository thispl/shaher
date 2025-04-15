# Copyright (c) 2025, gifty.p@groupteampro.com and contributors
# For license information, please see license.txt

# Copyright (c) 2025, Teampro and contributors
# For license information, please see license.txt

from __future__ import unicode_literals
import frappe,erpnext
from frappe.utils import cint
from frappe.utils import flt, fmt_money
import json
from frappe.utils import date_diff, add_months,today,add_days,add_years,nowdate,flt
from frappe.utils import cstr, cint, getdate,get_first_day, get_last_day, today, time_diff_in_hours

from frappe.model.mapper import get_mapped_doc
from frappe.utils.file_manager import get_file
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
import datetime
from datetime import date,datetime,timedelta

import openpyxl
from collections import defaultdict
from openpyxl import Workbook
import openpyxl
import xlrd
import re
from frappe.utils import date_diff,today,cstr


from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from frappe.utils import formatdate
from datetime import datetime

from erpnext.setup.utils import get_exchange_rate
from frappe import throw,_, db, get_doc, throw, whitelist

from dateutil.relativedelta import relativedelta
from frappe.utils import (
	add_days,
	cint,
	cstr,
	date_diff,
	flt,
	formatdate,
	get_first_day,
	get_link_to_form,
	getdate,
	money_in_words,
	rounded,
)
from frappe.model.document import Document
import math

class LeaveSalary(Document):
	def on_submit(self):
		self.validate_asset()
	def validate_asset(self):
		for data in self.assets_allocated:
			if data.status == "Owned" and data.action=='Return':
				frappe.throw(_("All allocated assets should be returned before submission"))

	def on_update(self):
		self.set_total_asset_recovery_cost()
		self.total_deduction = self.loan + self.advance + self.others + self.immigration_cost + self.visa + self.air_ticket_expense +self.others + self.mess_expense +self.total_asset_recovery_cost

		self.net_pay = self.total_earnings - self.total_deduction
	# def validate(self):
	# 	# self.get_assets_statements()
	# 	# self.set_total_asset_recovery_cost()
	# 	user_roles = frappe.get_roles(frappe.session.user)
	# 	hr="HR" in user_roles
	# 	hr_manager="HR Manager" in user_roles
	# 	if not(hr or hr_manager):
			
	# 		if self.type != "Indemnity":
	# 			if self.joining_or_last_rejoining_date:
	# 				diff = date_diff(today(),self.joining_or_last_rejoining_date)
	# 				if diff<270:
	# 					frappe.throw("You are not eligible to apply for this document")
	# 			else:
	# 				diff = date_diff(today(),self.joining_date)
	# 				if diff<270:
	# 					frappe.throw("You are not eligible to apply for this document")
	# 		else:
				
	# 			if isinstance(self.joining_date, str):
	# 				start_date = datetime.strptime(self.joining_date, '%Y-%m-%d')
	# 			else:
	# 				start_date = self.joining_date 

	# 			if isinstance(today(), str):
	# 				end_date = datetime.strptime(today(), '%Y-%m-%d')
	# 			else:
	# 				end_date = today()
	# 			difference = relativedelta(end_date, start_date)
	# 			years = difference.years
	# 			if years<1:
	# 				frappe.throw("You are not eligible to apply , because your service is below 1 years")
	# 	else:
	# 		count =0
	# 		attendance = frappe.get_all(
	# 		"Attendance",
	# 			filters={
	# 				"employee": self.employee_number,
	# 				"attendance_date": ["between", (self.joining_date, today())],
	# 				"docstatus": ["!=", 2],
	# 			},
	# 			fields=["status"]
	# 		)

	# 		for i in attendance:
	# 			if i.status =="Present":
	# 				count +=1
	# 			if i.status =="Half Day":
	# 				count+=0.5
	# 		if count < 100 :
	# 			frappe.throw("You are not eligible to apply for this document, because you are not working for 100 days")

	
	# While save the document all values automatically fetch to the documents
	def before_save(self):
		if self.is_new():
			self.last_leave_salary_taken_day=frappe.db.get_value("Leave Salary",{"employee_number":self.employee_number},["posting_date"])
		count = 0
		absent_days = 0
		basic = frappe.db.get_value("Employee",{"name":self.employee_number},["custom_basic"])
		doj = frappe.db.get_value('Employee',{"name":self.employee_number},'date_of_joining')
		hra = frappe.db.get_value("Employee",{"name":self.employee_number},["custom_hra"])
		allowance = frappe.db.get_value("Employee",{"name":self.employee_number},["custom_other_allowance"])
		if self.type != "Indemnity":
			if not self.joining_or_last_rejoining_date:
				frappe.throw("Enter the Service Start Date")
			if not self.date_of_service:
				frappe.throw("Enter the Service End Date")
			start_date = str(self.joining_or_last_rejoining_date)
			end_date = str(self.date_of_service)
			

			attendance = frappe.get_all(
			"Attendance",
				filters={
					"employee": self.employee_number,
					"attendance_date": ["between", (start_date, end_date)],
					"docstatus": ["!=", 2],
					"leave_type": ["in", ["Annual Leave"]]
				},
				fields=["status"]
			)

			for i in attendance:
				if i.status =="On Leave":
					count +=1
				if i.status =="Half Day":
					count+=0.5
		else:
			start_date = str(self.joining_or_last_rejoining_date)
			end_date = str(self.date_of_service)
			
			absent_days += frappe.db.get_value("Employee",{"name":self.employee_number},['custom_absent_days'])
			
			attendance = frappe.get_all(
			"Attendance",
				filters={
					"employee": self.employee_number,
					"attendance_date": ["between", (start_date, end_date)],
					"docstatus": ["=", 1],
					"leave_application":'' 
				},
				fields=["status"]
			)

			for i in attendance:
				if i.status =="Absent":
					absent_days +=1
				if i.status =="Half Day":
					absent_days+=0.5
		self.total_leave = count
		self.total_absent_days = absent_days
		total_days = date_diff( self.date_of_service,self.joining_date)
		payment_days = int(total_days)  + 1 

		self.leave_salary_days = payment_days - absent_days
		date_2 = datetime.now()
		diff = relativedelta(date_2, doj)
		# current_yoe_days = date_diff(nowdate(),self.joining_date)
		# current_yoe = round((current_yoe_days / 365),3)

		if(1 <= diff.years <= 3):
			gratuity_per_year = (basic * .5)
			total_gratuity = math.ceil(gratuity_per_year/365*(payment_days - absent_days))
			self.gratuity = total_gratuity
		elif(diff.years > 3):
			gratuity_per_year = basic
			total_gratuity = math.ceil(gratuity_per_year/365*(payment_days - absent_days))
			self.gratuity = total_gratuity		
		date_object = datetime.strptime(self.date_of_service, "%Y-%m-%d")
		month_value = date_object.month
		year_value=date_object.year
		date = f"{year_value}-{month_value:02d}-01"
		diff = date_diff(self.date_of_service,date)

		self.salary_payable_days = diff + 1
		if self.type != "Indemnity":
			
			salary = (basic + allowance) /30 * (diff + 1)

		else:
			if isinstance(self.joining_date, str):
				start_date = datetime.strptime(self.joining_date, '%Y-%m-%d')
			else:
				start_date = self.joining_date  # Already a datetime.date object

			if isinstance(self.date_of_service, str):
				end_date = datetime.strptime(self.date_of_service, '%Y-%m-%d')
			else:
				end_date = self.date_of_service
			difference = relativedelta(end_date, start_date)
			years = difference.years
			ctc = frappe.db.get_value("Employee",{"name":self.employee_number},["ctc"])
			if years <=5:
				salary = ctc * 75/26
			else:
				salary = ctc * ((15 * 5) + (30 *(years - 5))) / 26
			if salary > (ctc*12):
				salary = ctc
			else:
				salary = salary
		self.salary = salary
		total = date_diff( end_date,start_date)
		salary_days = int(total) - int(count) + 1
		self.leave_salary_days_salary = salary_days


		leave_salary = basic * salary_days /26		
		self.leave_salary = leave_salary

		if self.type == "Indemnity":
			total_earning = ( 
				(self.salary or 0) + 
				(self.ticket_allowance or 0) + 
				(self.leave_salary or 0) + 
				(self.gratuity or 0)
			)

		if self.type != "Indemnity":
			total_earning = (
				(self.leave_salary or 0) + 
				(self.ticket_allowance or 0)
			)

		
		self.total_earnings = total_earning
		
		self.total_deduction = (
			(self.loan or 0) + 
			(self.advance or 0) + 
			(self.others or 0) + 
			(self.immigration_cost or 0) + 
			(self.visa or 0) + 
			(self.air_ticket_expense or 0) + 
			(self.mess_expense or 0) + 
			(self.total_asset_recovery_cost or 0)
		)

		
		self.net_pay = self.total_earnings or 0 -  self.total_deduction or 0
		
	@frappe.whitelist()
	def get_outstanding_statements(self):
		self.get_assets_statements()

	def get_assets_statements(self):
		if not len(self.get("assets_allocated", [])):
			for data in self.get_assets_movement():
				self.append("assets_allocated", data)
				
	def get_assets_movement(self):
		asset_movements = frappe.get_all(
			"Asset Movement Item",
			filters={"docstatus": 1},
			fields=["asset", "from_employee", "to_employee", "parent", "asset_name"],
			or_filters={"from_employee": self.employee_number, "to_employee": self.employee_number},
		)

		data = []
		inward_movements = []
		outward_movements = []
		for movement in asset_movements:
			if movement.to_employee and movement.to_employee == self.employee_number:
				inward_movements.append(movement)

			if movement.from_employee and movement.from_employee == self.employee_number:
				outward_movements.append(movement)

		for movement in inward_movements:
			outwards_count = [movement.asset for movement in outward_movements].count(movement.asset)
			inwards_counts = [movement.asset for movement in inward_movements].count(movement.asset)

			if inwards_counts > outwards_count:
				cost = frappe.db.get_value("Asset", movement.asset, "total_asset_cost")
				data.append(
					{
						"reference": movement.parent,
						"asset_name": movement.asset_name,
						"date": frappe.db.get_value("Asset Movement", movement.parent, "transaction_date"),
						"actual_cost": cost,
						"cost": cost,
						"action": "Return",
						"status": "Owned",
					}
				)
		return data
	def set_total_asset_recovery_cost(self):
		total_cost = 0
		for data in self.assets_allocated:
			if data.action == "Recover Cost":
				if not data.description:
					data.description = _("Asset Recovery Cost for {0}: {1}").format(
						data.reference, data.asset_name
					)
				total_cost += flt(data.cost)

		self.total_asset_recovery_cost = flt(total_cost, self.precision("total_asset_recovery_cost"))


#Get the Basic , HRA and Other Allowance from Employee MIS and set to the Leave salary document while enter the Employee
@frappe.whitelist()
def get_employee_details(employee):
	values = frappe.get_all("Employee",{"name":employee},["*"])
	for i in values:
		return i.custom_basic,i.custom_hra,i.custom_other_allowance


# Get the serive period between the days of Employee Joining Date and Indemnity Date in Leave Salary
@frappe.whitelist()
def update_service_period(start_date,end_date):

	start_date = datetime.strptime(start_date, '%Y-%m-%d')
	end_date = datetime.strptime(end_date, '%Y-%m-%d')
	difference = relativedelta(end_date, start_date)
	years = difference.years
	months = difference.months
	days = difference.days
	if days < 0:
		months -= 1
		prev_month = end_date - relativedelta(months=1)
		days_in_prev_month = (prev_month + relativedelta(months=1)).replace(day=1) - prev_month.replace(day=1)
		days += days_in_prev_month.days
	return f"{years} Years, {months} Months, {days} Days"


#If the Given employee role is HOD , HOD Checkbox has been enabled
@frappe.whitelist()
def get_role(employee):
	user_id = frappe.get_value('Employee',{'employee':employee},['user_id'])
	hod = frappe.get_value('User',{'email':user_id},['name'])
	role = "HOD"
	hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
	if hod:
		return  "HI"
	else:
		return "HII"


#Update the Rejoining Date from Rejoining Form while open the leave salary
@frappe.whitelist()
def update_start_date(employee,type):
	if type != "Indemnity":
		value = frappe.db.get_value("Rejoining Form",{"employee_id":employee,"workflow_state":"Approved"},["re_join"])
	else:
		value = frappe.db.get_value("Employee",{"name":employee},["date_of_joining"])
	return value



