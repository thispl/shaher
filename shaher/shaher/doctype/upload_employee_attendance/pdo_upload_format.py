# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from datetime import date, datetime, timedelta
import openpyxl
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from six import BytesIO
from frappe.utils import cint,today,flt,date_diff,add_days,add_months,date_diff,getdate,formatdate,cint,cstr
from frappe.utils.file_manager import get_file
import requests
from frappe.utils import getdate
from datetime import datetime, timedelta
from frappe import _
from frappe.utils.file_manager import get_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import re
# class UploadEmployeeAttendance(Document):
    
#     def on_cancel(self):
#         att_count = frappe.db.count("Attendance", {'custom_upload_attendance': self.name,'docstatus': 1})
#         if att_count > 40:
#             frappe.enqueue("shaher.shaher.doctype.upload_employee_attendance.upload_employee_attendance.cancel_attendance_records", 
#                 queue='long', 
#                 timeout=1000, 
#                 docname=self.name)
#         else:
#             cancel_attendance_records(self.name)


#     def on_submit(self):
#         file_url = frappe.db.get_value(
#             "File",
#             {"attached_to_doctype": "Upload Employee Attendance", "attached_to_name": self.name},
#             "file_url"
#         )
#         if not file_url:
#             frappe.throw(_("No file attached for this document."))

#         try:
#             in_memory_file = download_external_file_url(file_url)
#             wb = openpyxl.load_workbook(in_memory_file)
#             ws = wb.active
#         except Exception as e:
#             frappe.throw(f"Failed to read Excel file: {str(e)}")

#         emp_count = get_employee_count(ws)

#         if emp_count > 30:
#             frappe.enqueue(
#                 method="shaher.shaher.doctype.upload_employee_attendance.pdo_upload_format.process_attendance_from_excel_job",
#                 queue="default",
#                 timeout=1200,
#                 now=False,
#                 docname=self.name
#             )
#             frappe.msgprint("Attendance is updating in the background.Kindly check after few minutes.")
#         else:
#             self.create_attendance_from_excel(ws)

#     def create_attendance_from_excel(self, ws):
#         start_row = 6  
#         max_row = ws.max_row
#         max_col = ws.max_column

#         try:
#             dates = get_att_dates(self.from_date, self.to_date)
#         except Exception as e:
#             frappe.log_error("Failed to get dates", "Attendance Upload")
#             return

#         usable_col_end = max_col - 4
#         actual_cols = usable_col_end - 4 
#         expected_cols = len(dates)

#         if expected_cols != actual_cols:
#             frappe.log_error(
#                 f"Column mismatch: expected {expected_cols} date columns for period {self.from_date} to {self.to_date}, but found {actual_cols} in Excel.",
#                 "Attendance Upload"
#             )
#             cols_to_process = min(expected_cols, actual_cols)
#         else:
#             cols_to_process = expected_cols

#         row = start_row
#         while row <= max_row:
#             emp_code = ws.cell(row=row, column=2).value
#             emp_name = ws.cell(row=row, column=3).value

#             if not emp_code:
#                 row += 3  # skip 3 rows (Attendance + OT + PDO OT)
#                 continue

#             b_count = ws.cell(row=row, column=max_col - 3).value or 0
#             lunch = ws.cell(row=row, column=max_col - 2).value or 0
#             dinner = ws.cell(row=row, column=max_col - 1).value or 0
#             mobile_allowance = ws.cell(row=row, column=max_col).value or 0
#             emp_status = frappe.db.get_value("Employee", {'name': emp_code}, ['status'])

#             present_count = 0
#             rd_count = 0
#             absent_count = 0
#             holiday_count = 0
#             ww_count = 0
#             worked_days = 0
#             total_ww_days = 0
#             total_hh_days = 0
#             total_ot_hours = 0
#             total_pdo_ot_hours = 0  # NEW

#             for idx in range(cols_to_process):
#                 attendance_date = dates[idx]
#                 doj = frappe.db.get_value("Employee", {'name': emp_code}, ['date_of_joining'])
#                 rel = frappe.db.get_value("Employee", {'name': emp_code}, ['relieving_date'])

#                 if emp_status == 'Active':
#                     if getdate(attendance_date) < getdate(doj):
#                         continue
#                 else:
#                     if getdate(attendance_date) > getdate(rel):
#                         continue

#                 col = 5 + idx  # date columns start after 5 fixed columns

#                 # Attendance status
#                 status = ws.cell(row=row, column=col).value
#                 status = (str(status).strip().upper() if status else "A")

#                 # OT HR
#                 ot = ws.cell(row=row + 1, column=col).value
#                 ot = float(ot) if ot not in (None, '') else 0.0

#                 # PDO OT HR
#                 pdo_ot = ws.cell(row=row + 2, column=col).value
#                 pdo_ot = float(pdo_ot) if pdo_ot not in (None, '') else 0.0

#                 total_ot_hours += ot
#                 total_pdo_ot_hours += pdo_ot

#                 # Attendance counting
#                 if status == "P":
#                     att_status = "Present"
#                     present_count += 1
#                     rd_status = 0
#                 elif status == "HD":
#                     att_status = "Half Day"
#                     rd_status = 0
#                     present_count += 0.5
#                 elif status == "RD":
#                     att_status = "Absent"
#                     rd_status = 1
#                     rd_count += 1
#                 else:
#                     att_status = "Absent"
#                     rd_status = 0

#                 holiday_list = frappe.db.get_value('Employee', emp_code, 'holiday_list')
#                 if not holiday_list:
#                     holiday_list = frappe.db.get_value('Company', self.company, 'default_holiday_list')

#                 if holiday_list and att_status != 'Present' and rd_status == 0:
#                     holiday_type = check_holiday(attendance_date, holiday_list)
#                     if holiday_type == 'WO':
#                         ww_count += 0.5 if att_status == 'Half Day' else 1
#                         total_ww_days += 1
#                     elif holiday_type == 'HH':
#                         holiday_count += 0.5 if att_status == 'Half Day' else 1
#                         total_hh_days += 1
#                     else:
#                         absent_count += 0.5 if att_status == 'Half Day' else 1

#                 # Update or create Attendance
#                 if frappe.db.exists("Attendance", {
#                     "employee": emp_code,
#                     "attendance_date": attendance_date,
#                     "docstatus": ["!=", 2]
#                 }):
#                     try:
#                         existing_att = frappe.get_doc("Attendance", {
#                             "employee": emp_code,
#                             "attendance_date": attendance_date
#                         })
#                         if existing_att.docstatus == 0:
#                             existing_att.status = att_status
#                             existing_att.custom_rest_day = rd_status
#                             existing_att.custom_overtime_hours = ot
#                             existing_att.custom_pdo_overtime_hours = pdo_ot
#                             existing_att.save(ignore_permissions=True)
#                             existing_att.submit()
#                             frappe.db.commit()
#                     except Exception as e:
#                         frappe.log_error(
#                             f"Failed to update existing attendance for {emp_code} on {attendance_date}: {str(e)}",
#                             "Attendance Upload"
#                         )
#                 else:
#                     try:
#                         att = frappe.new_doc("Attendance")
#                         att.employee = emp_code
#                         att.employee_name = emp_name
#                         att.attendance_date = attendance_date
#                         att.custom_upload_attendance = self.name
#                         att.custom_overtime_hours = ot
#                         att.custom_pdo_overtime_hours = pdo_ot
#                         att.status = att_status
#                         att.custom_rest_day = rd_status
#                         att.insert(ignore_permissions=True)
#                         att.submit()
#                         frappe.db.commit()
#                     except Exception as e:
#                         frappe.log_error(
#                             f"Error creating attendance for {emp_code} on {attendance_date.strftime('%d-%m-%Y')}: {str(e)}",
#                             "Attendance Upload"
#                         )

#             # Move to next employee (3 rows per employee)
#             row += 3

#             # Update Attendance and OT Register
#             register_name = frappe.db.get_value(
#                 "Attendance and OT Register",
#                 {
#                     'from_date': ('<=', self.to_date),
#                     'to_date': ('>=', self.from_date),
#                     'employee': emp_code,
#                     'docstatus': ['!=', 2]
#                 },
#                 'name'
#             )

#             worked_days = present_count + rd_count + ww_count + holiday_count

#             if register_name:
#                 try:
#                     register_doc = frappe.get_doc("Attendance and OT Register", register_name)
#                     register_doc.breakfast = float(b_count)
#                     register_doc.lunch = float(lunch)
#                     register_doc.dinner = float(dinner)
#                     register_doc.mobile_allow = float(mobile_allowance)
#                     register_doc.attendance_upload = self.name
#                     register_doc.rest_days = rd_count
#                     register_doc.present_days = present_count
#                     register_doc.absent_days = absent_count
#                     register_doc.total_weekly_holidays = total_ww_days
#                     register_doc.holidays = total_hh_days
#                     register_doc.no_of_days_worked = worked_days
#                     register_doc.ot_hours = total_ot_hours
#                     register_doc.pdo_ot_hours = total_pdo_ot_hours
#                     register_doc.total_ot_hours = total_ot_hours + total_pdo_ot_hours
#                     register_doc.save(ignore_permissions=True)
#                     frappe.db.commit()
#                 except Exception as e:
#                     frappe.log_error(
#                         f"Failed to update 'Attendance and OT Register' for {emp_code}: {str(e)}",
#                         "Attendance Upload"
#                     )

        

@frappe.whitelist()
def process_attendance_from_excel_job(docname):
    doc = frappe.get_doc("Upload Employee Attendance", docname)

    file_url = frappe.db.get_value(
        "File",
        {"attached_to_doctype": "Upload Employee Attendance", "attached_to_name": doc.name},
        "file_url"
    )

    if not file_url:
        frappe.log_error("No file found", "Attendance Upload Background Job")
        return

    try:
        in_memory_file = download_external_file_url(file_url)
        wb = openpyxl.load_workbook(in_memory_file)
        ws = wb.active
        doc.create_attendance_from_excel(ws)
    except Exception as e:
        frappe.log_error(f"Failed to process attendance in background: {str(e)}", "Attendance Upload")


@frappe.whitelist()
def get_att_dates(from_date, to_date):
        no_of_days = date_diff(add_days(to_date, 1), from_date)
        dates = [add_days(from_date, i) for i in range(no_of_days)]
        return dates

def download_external_file_url(file_url):
    path = frappe.get_site_path("public", file_url.replace("/files/", ""))
    with open(path, "rb") as f:
        return BytesIO(f.read())


def get_employee_count(ws):
    # Count how many rows contain employee codes (assuming every 2nd row is OT)
    return sum(
        1 for r in range(6, ws.max_row + 1, 2)
        if ws.cell(row=r, column=2).value
    )

from frappe.utils import get_url
@frappe.whitelist()
def download_external_file_url(file_url):
    if not file_url.startswith("http"):
        file_url = get_url() + file_url
    response = requests.get(file_url, stream=True)
    if response.status_code != 200:
        frappe.throw(_("Failed to download file from external storage."))

    return BytesIO(response.content)

@frappe.whitelist()
def get_template(from_date, to_date, department=None, designation=None, site_location=None, name=None, division=None):
    filename = 'Attendance Template'
    test = build_xlsx_response(filename,division)

@frappe.whitelist()
def build_xlsx_response(filename,division):
    if "PDO" in division: 
        xlsx_file = make_xlsx_pdo(filename)
    else:
        xlsx_file = make_xlsx_oetc(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def make_xlsx_pdo(data=None, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    gray_fill = PatternFill(start_color="D9D9D9", fill_type="solid")
    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)

    dates = get_dates(args.from_date, args.to_date)

    

    comp = frappe.db.get_value("Upload Employee Attendance", {'name': args.name}, 'company')
    ws.append([comp])

    site_location = frappe.db.get_value("Upload Employee Attendance", {'name': args.name}, 'site_location')
    ws.append([site_location] if site_location else [])

    from_date_str = datetime.strptime(args.from_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    to_date_str = datetime.strptime(args.to_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    ws.append([f"TIME SHEET {from_date_str} to {to_date_str}"])

    header3 = [
        "SI No.",
        "Emp. Code",
        "Employee Name",
        "Designation",
        ""     
    ]

    header3.extend([d[0] for d in dates])

    header3.extend([
        
        "No of Working Days",
        "Normal OT ",
        "OT from PDO",
        "Total OT Hrs",
        "Total Extra Days",
        "Extra days rate",
        "Extra days Amount",
        "Total Rest Days",
        "Total Leave Days",
        "Total Absent Days",
        "Total Medical Leave Days",
        "Total Emergency Leave Days",
        "Total NHP Days",
        "NHP Rate",
        "NH Amount",
        "Breakfast",
        "Lunch",
        "Dinner",
        "Food Allowance",
        "Mobile Allowance",
        "Other Allowance",
        "Remarks - Other Allowance",
        "Other Deduction",
        "Remarks - Other Deduction",
        "Salary Adjustment",
        "Arrear Salary",
        "Remarks for paid and deducted now"
    ])

    header4 = ["", "", "", "", ""]
    header4.extend([d[1] for d in dates])
    header4.extend([""] * 13) 
    ws.append(header3)
    ws.append(header4)

    

    emp_data = get_employees(args)

    row_num = 1
    start_row = 6

    for emp in emp_data:
        doj = emp.get("date_of_joining")
        if isinstance(doj, str):
            doj = datetime.strptime(doj, "%Y-%m-%d").date()

        emp_row = [
            row_num,
            emp.get("name"),
            emp.get("employee_name"),
            emp.get("designation"),
            "Atted"
        ]   

        emp_row.extend([""] * len(dates))  
        def col_letter(n):
            result = ""
            while n:
                n, rem = divmod(n - 1, 26)
                result = chr(65 + rem) + result
            return result


        start_row_ot = (row_num -1)*3 + 5 + 2
        start_row_tot = (row_num -1)*3 + 5 + 1
        start_row_ot_pdo = (row_num -1)*3 + 5 + 3
        start_col_ot = "F"
        end_col_ot = col_letter(len(dates) + 5)
        start_col_tot = col_letter(len(dates) + 7)
        end_col_tot = col_letter(len(dates) + 8)
        emp_row.append(len(dates))
        emp_row.append(
            f"=SUM({start_col_ot}{start_row_ot}:{end_col_ot}{start_row_ot})"
        )
        emp_row.append(
            f"=SUM({start_col_ot}{start_row_ot_pdo}:{end_col_ot}{start_row_ot_pdo})"
        )
        emp_row.append(f"=SUM({start_col_tot}{start_row_tot}:{end_col_tot}{start_row_tot})")          

        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"E\")"
        )
        emp_row.append(1.25)
        start_col_tot_ex = col_letter(len(dates) + 10)
        end_col_tot_ex = col_letter(len(dates) + 11)
        emp_row.append(
            f"={start_col_tot_ex}{start_row_tot}*{end_col_tot_ex}{start_row_tot}"
        )
        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"RD\")"
        )
        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"L\")"
        )   
        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"A\")"
        )  
        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"M\")"
        )      
        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"EM\")"
        )     
        emp_row.append(
            f"=COUNTIF({start_col_ot}{start_row_ot-1}:{end_col_ot}{start_row_ot-1},\"NH-P\")"
        )  
        emp_row.append(1.25)
        start_col_tot_nh = col_letter(len(dates) + 18)
        end_col_tot_nh = col_letter(len(dates) + 19)
        emp_row.append(
            f"={start_col_tot_nh}{start_row_ot-1}*{end_col_tot_nh}{start_row_ot-1}"
        )
        emp_row.extend([""] * 3)  
        start_col_tot_break = col_letter(len(dates) + 21)
        start_col_tot_lun = col_letter(len(dates) + 22)
        end_col_tot_din = col_letter(len(dates) + 23)
        emp_row.append(
            f"={start_col_tot_break}{start_row_ot-1}+({start_col_tot_lun}{start_row_ot-1}*2)+({end_col_tot_din}{start_row_ot-1}*2)"
        )
        # emp_row.append()        

        ws.append(emp_row)
   
        ot_hr_row = ["", "", "", "", "OT HR"]
        ot_hr_row.extend([""] * (len(dates) + 13))
        ws.append(ot_hr_row)

        pdo_ot_hr_row = ["", "", "", "", "PDO OT HR"]
        pdo_ot_hr_row.extend([""] * (len(dates) + 13))
        ws.append(pdo_ot_hr_row)
        date_start_col = 6  # Column F
        for row_offset in range(3):  # main + OT + PDO OT
            for idx, (work_date, _) in enumerate(dates):
                col = date_start_col + idx
                cell_date = getdate(work_date)

                # Gray out BEFORE DOJ
                if doj and cell_date < getdate(doj):
                    ws.cell(row=start_row + row_offset, column=col).fill = gray_fill

                # Gray out AFTER Relieving Date
                if emp.get("relieving_date") and cell_date > getdate(emp.get("relieving_date")):
                    ws.cell(row=start_row + row_offset, column=col).fill = gray_fill

        for col in range(1, 5):
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=start_row + 2,
                end_column=col
            )

        row_num += 1
        
        start_row_merge=6+len(dates)
        for col in range(start_row_merge,start_row_merge+27):
            ws.merge_cells(
                start_row=start_row,
                start_column=col,
                end_row=start_row + 2,
                end_column=col
            )
        start_row += 3
    
    merge_length = 5 + 1 + len(dates) + 26 

    if not site_location:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=merge_length)
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_length)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_length)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=merge_length)
    from openpyxl.utils import get_column_letter

# Columns for width 20
    wide_columns = [
        "Total OT",
        "OT from PDO",
        "Total OT Hrs",
        "No of Working Days",
        "Breakfast",
        "Lunch",
        "Dinner",
        "Rate Per Extra Day",
        "Rate Per Day NH",
        "Mobile Allowance",
        "Other Allowance",
        "Remarks - Other Allowance",
        "Other Deduction",
        "Remarks - Other Deduction",
        "Salary Adjustment",
        "Arrear Salary",
        "Remarks for paid and deducted now"
    ]

    header_row = 1  # change if your header starts later

    # Loop through all headers in this row and adjust where matches
    for col in range(1, ws.max_column + 1):
        header_text = ws.cell(row=header_row, column=col).value
        if header_text in wide_columns:
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 20

    # Merge fixed 5 columns in header3/header4
    fixed_columns_after_dates = 27

    # Merge columns 1–5 (SI No → Attent.)
    for col in range(1, 6):
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)

    # Merge *each fixed column* after the dates
    start_fixed_col = 6 + len(dates)  # col 6 is first date column so start at col = 6 + len(dates)

    for offset in range(fixed_columns_after_dates):
        col = start_fixed_col + offset
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
    
    

    for i in range(fixed_columns_after_dates):
        col_letter = get_column_letter(start_fixed_col + i)
        ws.column_dimensions[col_letter].width = 20 

    header_fill = PatternFill(start_color="E6B8B7", fill_type="solid")
    friday_fill = PatternFill(start_color="92D050", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)

    for row in [4, 5]:
        for col in range(1, merge_length + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_alignment

    for col_idx, cell in enumerate(ws[5], start=1):
        if cell.value == "Fr":
            ws.cell(row=5, column=col_idx).fill = friday_fill

    for col in range(1, merge_length + 1):
        ws.column_dimensions[get_column_letter(col)].width = 8
    
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 43
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16

    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=merge_length):
        for cell in rows:
            cell.border = border

    for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=merge_length):
        for cell in rows:
            cell.alignment = center_alignment
            cell.font = bold_font
    from openpyxl.utils import get_column_letter

    # Set width for the last 27 columns to 18
    for col in range(merge_length - 26, merge_length + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


@frappe.whitelist()
def make_xlsx_oetc(data=None, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    gray_fill = PatternFill(start_color="D9D9D9", fill_type="solid")
    from openpyxl.utils import get_column_letter

    if wb is None:
        wb = openpyxl.Workbook()

    ws = wb.create_sheet(sheet_name, 0)

    dates = get_dates(args.from_date, args.to_date)
    from_date = getdate(args.from_date)
    to_date = getdate(args.to_date)
    gross_days = (to_date - from_date).days + 1
    comp = frappe.db.get_value("Upload Employee Attendance", {'name': args.name}, 'company')
    ws.append([comp])

    site_location = frappe.db.get_value("Upload Employee Attendance", {'name': args.name}, 'site_location')
    ws.append([site_location] if site_location else [])

    from_date_str = datetime.strptime(args.from_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    to_date_str = datetime.strptime(args.to_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    ws.append([f"TIME SHEET {from_date_str} to {to_date_str}"])

    # ---------------- HEADER ---------------- #
    header3 = ["SI No.", "Emp. Code", "Employee Name", "Designation", ""]
    header3.extend([d[0] for d in dates])
    header3.extend([
        "No of Working Days",
        "Normal OT(1.25)",
        "Night OT(1.5)",
        "Holiday OT(2.0)",
        "Mobile Allowance",
        "Other Allowance",
        "Remarks - Other Allowance",
        "Additional Allowance",
        "Other Deduction",
        "Remarks - Other Deduction",
        "Medical Leave",
        "Salary Adjustment",
        "Arrear Salary",
        "Remarks for paid and deducted now",
        "Net Days"
    ])

    header4 = ["", "", "", "", ""]
    header4.extend([d[1] for d in dates])
    header4.extend([""] * 12)

    ws.append(header3)
    ws.append(header4)

    # ---------------- EMPLOYEES ---------------- #
    emp_data = get_employees(args)

    row_num = 1
    start_row = 6

    date_cell_count = len(dates)
    fixed_cols = 12  # updated

    for emp in emp_data:
        date_blank = [""] * date_cell_count
        tail_blank = [""] * fixed_cols

        # Row 1 – Attendance
        r1 = [row_num, emp.get("name"), emp.get("employee_name"), emp.get("designation"), "Attendance"]
        r1.extend(date_blank)
        def col_letter(n):
            result = ""
            while n:
                n, rem = divmod(n - 1, 26)
                result = chr(65 + rem) + result
            return result


        start_row_ot = (row_num -1)*4 + 2 +5
        start_row_tot = (row_num -1)*4 + 4 +5
        start_row_ot_pdo = (row_num -1)*4 + 3 +5
        start_col_ot = "F"
        end_col_ot = col_letter(len(dates) + 5)
        start_col_tot = col_letter(len(dates) + 6)
        end_col_tot = col_letter(len(dates) + 7)
        r1.extend([gross_days])
        r1.append(
            f"=SUM({start_col_ot}{start_row_ot}:{end_col_ot}{start_row_ot})"
        )
        r1.append(
            f"=SUM({start_col_ot}{start_row_ot_pdo}:{end_col_ot}{start_row_ot_pdo})"
        )
        r1.append(f"=SUM({start_col_ot}{start_row_tot}:{end_col_ot}{start_row_tot})")
                # <-- Inserted correctly
        r1.extend(tail_blank[:-1])     # <-- Keep equal columns
        ws.append(r1)

        # Row 2 – Normal OT
        r2 = ["", "", "", "", "Normal OT"]
        r2.extend(date_blank)
        r2.extend(tail_blank)
        ws.append(r2)

        # Row 3 – Night OT
        r3 = ["", "", "", "", "Night OT"]
        r3.extend(date_blank)
        r3.extend(tail_blank)
        ws.append(r3)

        # Row 4 – Holiday OT
        r4 = ["", "", "", "", "Holiday OT"]
        r4.extend(date_blank)
        r4.extend(tail_blank)
        ws.append(r4)
        date_start_col = 6  # Column F
        for row_offset in range(4): 
            for idx, (work_date, _) in enumerate(dates):
                col = date_start_col + idx
                cell_date = getdate(work_date)

                # Gray out BEFORE DOJ
                if emp.date_of_joining and cell_date < getdate(emp.date_of_joining):
                    ws.cell(row=start_row + row_offset, column=col).fill = gray_fill

                # Gray out AFTER Relieving Date
                if emp.get("relieving_date") and cell_date > getdate(emp.get("relieving_date")):
                    ws.cell(row=start_row + row_offset, column=col).fill = gray_fill
        # Merge ID columns
        for col in range(1, 5):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + 3, end_column=col)

        # Merge tail columns (including Gross Days)
        start_fixed_col = 6 + date_cell_count
        for col in range(start_fixed_col, start_fixed_col + fixed_cols+3):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + 3, end_column=col)

        row_num += 1
        start_row += 4


    merge_length = 5 + len(dates) + fixed_cols+3

    # Merge top headers
    if not site_location:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=merge_length)
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_length)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_length)

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=merge_length)

    # Merge row 4 & 5 headers
    for col in range(1, 6):
        ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)

    start_fixed_col = 6 + date_cell_count
    for i in range(fixed_cols+3):
        ws.merge_cells(start_row=4, start_column=start_fixed_col + i,
                       end_row=5, end_column=start_fixed_col + i)

    # ------------------------------------- STYLES ------------------------------------- #
    header_fill = PatternFill(start_color="E6B8B7", fill_type="solid")
    friday_fill = PatternFill(start_color="92D050", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)

    for row in [4, 5]:
        for col in range(1, merge_length + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_alignment

    for col_idx, cell in enumerate(ws[5], start=1):
        if cell.value == "Fr":
            ws.cell(row=5, column=col_idx).fill = friday_fill
    merge_column=merge_length+1
    for col in range(1, merge_column + 3):
        ws.column_dimensions[get_column_letter(col)].width = 8

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 43
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 16
    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=merge_length):
        for cell in rows:
            cell.border = border

    # Top header center + bold
    for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=merge_length):
        for cell in rows:
            cell.alignment = center_alignment
            cell.font = bold_font
    from openpyxl.utils import get_column_letter

    # Set width for the last 14 columns to 20
    for col in range(merge_length - 11, merge_length + 1):  # last 14 columns
        ws.column_dimensions[get_column_letter(col)].width = 20
    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def get_dates(from_date, to_date):
    start_date = datetime.strptime(from_date, "%Y-%m-%d")
    end_date = datetime.strptime(to_date, "%Y-%m-%d")
    delta = end_date - start_date
    days = []
    for i in range(delta.days + 1):
        date = start_date + timedelta(days=i)
        day_num = date.strftime("%d")  
        weekday = date.strftime("%a")[:2]
        days.append((day_num, weekday))
    return days

def get_employees(args):
    from_date = args.from_date
    to_date = args.to_date

    filters = []
    or_filters = []

    # Joined on or before payroll period end
    filters.append(['date_of_joining', '<=', to_date])

    # Relieving condition (OR)
    or_filters.append(['relieving_date', 'is', 'not set'])
    or_filters.append(['relieving_date', '>=', from_date])

    dept = frappe.db.get_value(
        "Upload Employee Attendance", {'name': args.name}, 'department'
    )
    desi = frappe.db.get_value(
        "Upload Employee Attendance", {'name': args.name}, 'designation'
    )
    comp = frappe.db.get_value(
        "Upload Employee Attendance", {'name': args.name}, 'company'
    )
    location = frappe.db.get_value(
        "Upload Employee Attendance", {'name': args.name}, 'site_location'
    )

    if args.department and args.department != 'None' and dept:
        filters.append(['department', '=', dept])

    if args.designation and args.designation != 'None' and desi:
        filters.append(['designation', '=', desi])

    if args.company and args.company != 'None' and comp:
        filters.append(['company', '=', comp])

    if args.site_location and args.site_location != 'None' and location:
        filters.append(['custom_site_location', '=', location])

    employees = frappe.db.get_all(
        'Employee',
        filters=filters,
        or_filters=or_filters,
        fields=[
            'name',
            'employee_name',
            'designation',
            'date_of_joining',
            'relieving_date'
        ]
    )

    def sort_key(emp):
        code = emp['name']
        match = re.match(r"([A-Za-z]+)(\d*)", code)
        prefix = match.group(1)
        number = int(match.group(2)) if match.group(2).isdigit() else 0
        return (prefix, number)

    return sorted(employees, key=sort_key)

# def get_employees(args):
#     filters = {'status': 'Active'}
#     dept=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['department'])
#     desi=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['designation'])
#     comp=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['company'])
#     location = frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['site_location'])
#     division = frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['division'])
#     if args.department and args.department!='None':
#         if dept:
#             filters['department'] = dept
#     if args.designation and args.designation!='None':
#         if desi:
#             filters['designation'] = desi
#     if args.company and args.company!='None':
#         if comp:
#             filters['company'] = comp
#     if args.site_location and args.site_location!='None':
#         if location:
#             filters['custom_site_location'] = location
#     # if args.division and args.division!='None':
#     #     if division:
#     #         filters['custom_division'] = division
#     employees = frappe.db.get_all('Employee', filters=filters, fields=['name', 'employee_name','designation','date_of_joining','relieving_date'])
#     def sort_key(emp):
#         code = emp['name']
#         match = re.match(r"([A-Za-z]+)(\d*)", code)

#         prefix = match.group(1)
#         number = match.group(2)

#         number = int(number) if number.isdigit() else 0

#         return (prefix, number)

#     employees = sorted(employees, key=sort_key)
#     return employees

import frappe

@frappe.whitelist()
def cancel_attendance_records(docname):
    att_list = frappe.db.get_all("Attendance", {
        'custom_upload_attendance': docname,
        'docstatus': 1
    }, ['name'])

    for a in att_list:
        try:
            att_doc = frappe.get_doc('Attendance', a.name)
            att_doc.cancel()
            frappe.db.commit()
        except Exception as e:
            frappe.log_error(f"{a.name}: {e}", "Upload Employee Attendance Cancel")

    reg_list = frappe.db.get_all("Attendance and OT Register", {
        'attendance_upload': docname
    }, ['name'])

    for r in reg_list:
        try:
            reg_doc = frappe.get_doc('Attendance and OT Register', r.name)
            reg_doc.breakfast = 0
            reg_doc.lunch = 0
            reg_doc.dinner = 0
            reg_doc.mobile_allow = 0
            reg_doc.attendance_upload = docname
            reg_doc.save(ignore_permissions=True)
            frappe.db.commit()
        except Exception as e:
            frappe.log_error(f"{r.name}: {e}", "Upload Employee Attendance Cancel")

@frappe.whitelist()
def check_holiday(date, holiday_list):
    query = """
        SELECT `tabHoliday`.holiday_date, `tabHoliday`.weekly_off
        FROM `tabHoliday List`
        LEFT JOIN `tabHoliday` ON `tabHoliday`.parent = `tabHoliday List`.name
        WHERE `tabHoliday List`.name = %s AND holiday_date = %s
    """
    holiday = frappe.db.sql(query, (holiday_list, date), as_dict=True)
    if holiday:
        if holiday[0].weekly_off == 1:
            return "WO"
        else:
            return "HO"
    return None
