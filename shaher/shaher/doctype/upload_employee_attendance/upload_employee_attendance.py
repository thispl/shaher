# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from datetime import date, datetime, timedelta
import openpyxl
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from six import BytesIO
from frappe.utils import cint,today,flt,date_diff,add_days,add_months,date_diff,getdate,formatdate,cint,cstr
from frappe.utils.file_manager import get_file
import requests
from frappe.utils import getdate
from datetime import datetime, timedelta
from frappe import _
from frappe.utils.file_manager import get_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
import re
class UploadEmployeeAttendance(Document):
    
    def on_cancel(self):
        att_count = frappe.db.count("Attendance", {'custom_upload_attendance': self.name,'docstatus': 1})
        if att_count > 40:
            frappe.enqueue("shaher.shaher.doctype.upload_employee_attendance.upload_employee_attendance.cancel_attendance_records", 
                queue='long', 
                timeout=1000, 
                docname=self.name)
        else:
            cancel_attendance_records(self.name)


    def on_submit(self):
        file_url = frappe.db.get_value(
            "File",
            {"attached_to_doctype": "Upload Employee Attendance", "attached_to_name": self.name},
            "file_url"
        )
        if not file_url:
            frappe.throw(_("No file attached for this document."))

        try:
            in_memory_file = download_external_file_url(file_url)
            wb = openpyxl.load_workbook(in_memory_file)
            ws = wb.active
        except Exception as e:
            frappe.throw(f"Failed to read Excel file: {str(e)}")

        emp_count = get_employee_count(ws)
        frappe.errprint(emp_count)
        if "PDO" in self.division: 
            if emp_count > 30:
                frappe.enqueue(
                    method="shaher.shaher.doctype.upload_employee_attendance.upload_employee_attendance.process_attendance_from_excel_job_pdo",
                    queue="default",
                    timeout=1200,
                    now=False,
                    docname=self.name,
                    docstatus=1
                )
                frappe.msgprint("Attendance is updating in the background.Kindly check after few minutes.")
            else:
                self.create_attendance_from_excel_pdo(ws)
        else:
            if emp_count > 30:
                frappe.enqueue(
                    method="shaher.shaher.doctype.upload_employee_attendance.upload_employee_attendance.process_attendance_from_excel_job_oetc",
                    queue="default",
                    timeout=1200,
                    now=False,
                    docname=self.name,
                    docstatus=1
                )
                frappe.msgprint("Attendance is updating in the background.Kindly check after few minutes.")
            else:
                self.create_attendance_from_excel_oetc_sub(ws)
    def validate(self):
        if not self.division:
            frappe.throw("Kindly enter the division.")
        import openpyxl
        from io import BytesIO

        self.upload_att_employee = []

        if not self.upload_attendance:
            return 

        file_doc = frappe.get_doc("File", {"file_url": self.upload_attendance})
        file_content = file_doc.get_content()
        wb = openpyxl.load_workbook(BytesIO(file_content))
        ws = wb.active

        start_row = 6
        max_row = ws.max_row

        row = start_row
        while row <= max_row:
            emp_code = ws.cell(row=row, column=2).value
            emp_name = ws.cell(row=row, column=3).value

            if emp_code:
                self.append("upload_att_employee", {
                    "employee": emp_code,
                })
            if "PDO" in self.division:
                row += 3  
            else:
                row += 4
        emp_count = get_employee_count(ws)
        
        if "PDO" in self.division:
            if emp_count > 30:
                frappe.enqueue(
                    method="shaher.shaher.doctype.upload_employee_attendance.upload_employee_attendance.process_attendance_from_excel_job_pdo",
                    queue="default",
                    timeout=1200,
                    now=False,
                    docname=self.name,
                    docstatus=0
                )
                frappe.msgprint("Attendance is updating in the background.Kindly check after few minutes.")
            else:
                self.create_att_ot_in_draft_pdo(ws)
        else:
            if emp_count > 30:
                frappe.enqueue(
                    method="shaher.shaher.doctype.upload_employee_attendance.upload_employee_attendance.process_attendance_from_excel_job_oetc",
                    queue="default",
                    timeout=1200,
                    now=False,
                    docname=self.name,
                    docstatus=0
                )
                frappe.msgprint("Attendance is updating in the background.Kindly check after few minutes.")
            else:
                self.create_attendance_from_excel_oetc(ws)
    
    def create_attendance_from_excel_oetc(self, ws):
        start_row = 6
        max_row = ws.max_row

        processed_employees = []

        # Fetch Time Period
        try:
            dates = get_att_dates(self.from_date, self.to_date)
        except Exception:
            frappe.log_error("Failed to get dates", "Attendance Upload")
            return

        date_start_col = 6
        date_cols = len(dates)

        # Gross days from date diff (inclusive)
        gross_days_actual = (getdate(self.to_date) - getdate(self.from_date)).days + 1

        row = start_row

        while row <= max_row:

            emp_code = ws.cell(row=row, column=2).value
            if not emp_code:
                row += 4
                continue

            processed_employees.append(emp_code)

            emp_status = frappe.db.get_value("Employee", emp_code, "status")
            doj, rel = frappe.db.get_value("Employee", emp_code, ["date_of_joining", "relieving_date"])

            present_count = absent_count = holiday_count = 0
            ww_count = rd_count = 0
            total_ot_hours = total_pdo_ot_hours = total_night_ot_hours = 0
            extra_days = nh_days = m_days = 0
            total_ww_days = total_hh_days = 0

            # Loop through Dates
            for idx in range(date_cols):
                attendance_date = getdate(dates[idx])
                col = date_start_col + idx

                # Skip dates before DOJ and after relieving
                if emp_status == "Active" and attendance_date < doj:
                    continue
                if emp_status != "Active" and rel and attendance_date > rel:
                    continue

                status = ws.cell(row=row, column=col).value
                status = str(status).strip().upper() if status else "A"

                ot = ws.cell(row=row + 1, column=col).value
                pdo_ot = ws.cell(row=row + 2, column=col).value
                night_ot = ws.cell(row=row + 3, column=col).value

                ot = float(ot) if ot else 0
                pdo_ot = float(pdo_ot) if pdo_ot else 0
                night_ot = float(night_ot) if night_ot else 0

                total_ot_hours += ot
                total_pdo_ot_hours += pdo_ot
                total_night_ot_hours += night_ot

                # Attendance count
                if status == "P":
                    present_count += 1
                elif status == "HD":
                    present_count += 0.5
                elif status == "RD":
                    rd_count += 1
                else:
                    if status == "E":
                        extra_days += 1
                    if status == "NH-P":
                        nh_days += 1
                    if status == "M":
                        m_days += 1
                    absent_count += 1

                # Holiday check
                holiday_list = frappe.db.get_value("Employee", emp_code, "holiday_list") \
                                or frappe.db.get_value("Company", self.company, "default_holiday_list")

                if holiday_list:
                    holiday_type = check_holiday(attendance_date, holiday_list)
                    if holiday_type == "WO":
                        ww_count += 1
                        total_ww_days += 1
                    elif holiday_type == "HH":
                        holiday_count += 1
                        total_hh_days += 1

            # Read only other earnings & deduction from Excel
            extra_start_col = date_start_col + date_cols +3

            mobile_allow = ws.cell(row=row, column=extra_start_col + 1).value or 0
            other_allowance = ws.cell(row=row, column=extra_start_col + 2).value or 0
            remarks_other_allowance = ws.cell(row=row, column=extra_start_col + 3).value or ""
            additional_allowance = ws.cell(row=row, column=extra_start_col + 4).value or 0
            other_deduction = ws.cell(row=row, column=extra_start_col + 5).value or 0
            remarks_other_deduction = ws.cell(row=row, column=extra_start_col + 6).value or ""
            medical_leave = ws.cell(row=row, column=extra_start_col + 7).value or 0
            salary_adjustment = ws.cell(row=row, column=extra_start_col + 8).value or 0
            paid_earlier = ws.cell(row=row, column=extra_start_col + 9).value or ""
            remarks_salary = ws.cell(row=row, column=extra_start_col + 10).value or ""

            # Move to next employee block
            row += 4

            worked_days = present_count + rd_count + ww_count + holiday_count

            register_name = frappe.db.get_value(
                "Attendance and OT Register",
                {
                    "employee": emp_code,
                    "from_date": self.from_date,
                    "to_date": self.to_date,
                    "docstatus": ["!=", 2]
                },
                "name"
            )

            try:
                if register_name:
                    reg = frappe.get_doc("Attendance and OT Register", register_name)
                else:
                    reg = frappe.new_doc("Attendance and OT Register")
                    reg.employee = emp_code
                    reg.from_date = self.from_date
                    reg.to_date = self.to_date
                    reg.attendance_upload = self.name

                reg.present_days = present_count
                reg.absent_days = absent_count
                reg.rest_days = rd_count
                reg.holidays = total_hh_days
                reg.total_weekly_holidays = total_ww_days
                reg.no_of_days_worked = worked_days

                reg.ot_hours = total_ot_hours
                reg.night_ot = total_pdo_ot_hours
                reg.holiday_ot = total_night_ot_hours
                reg.ot_amount=total_ot_hours*1.25
                reg.holiday_ot_amount=total_night_ot_hours*2
                reg.night_ot_amount=total_pdo_ot_hours*1.5
                reg.total_ot_hours = total_ot_hours + total_pdo_ot_hours + total_night_ot_hours

                reg.medical_leave_days = medical_leave
                reg.arrear_salary = paid_earlier

                reg.gross_days = gross_days_actual

                reg.mobile_allow = mobile_allow
                reg.other_earnings = int(other_allowance) + int(additional_allowance)
                reg.other_earnings_remark = remarks_other_allowance
                reg.other_deduction = other_deduction
                reg.other_deduction_remark = remarks_other_deduction
                reg.salary_adjustment = salary_adjustment
                reg.remarks_for_adjustment = remarks_salary

                if not register_name:
                    reg.insert(ignore_permissions=True)
                reg.save(ignore_permissions=True)
                frappe.db.commit()

            except Exception as e:
                frappe.log_error(f"{emp_code} Import Error", str(e))

        return processed_employees


    def create_attendance_from_excel_oetc_sub(self, ws):
        start_row = 6
        max_row = ws.max_row
        processed_employees = []

        # Fetch Time Period
        try:
            dates = get_att_dates(self.from_date, self.to_date)
        except Exception:
            frappe.log_error("Failed to get dates", "Attendance Upload")
            return

        date_start_col = 6
        date_cols = len(dates)

        # Gross days from date diff (inclusive)
        gross_days_actual = (getdate(self.to_date) - getdate(self.from_date)).days + 1

        row = start_row

        while row <= max_row:

            emp_code = ws.cell(row=row, column=2).value
            if not emp_code:
                row += 4
                continue

            processed_employees.append(emp_code)

            emp_status = frappe.db.get_value("Employee", emp_code, "status")
            doj, rel = frappe.db.get_value("Employee", emp_code, ["date_of_joining", "relieving_date"])

            present_count = absent_count = holiday_count = 0
            ww_count = rd_count = 0
            total_ot_hours = total_pdo_ot_hours = total_night_ot_hours = 0
            extra_days = nh_days = m_days = 0
            total_ww_days = total_hh_days = 0

            # Loop through Dates
            for idx in range(date_cols):
                attendance_date = getdate(dates[idx])
                col = date_start_col + idx

                # Skip dates before DOJ and after relieving
                if emp_status == "Active" and attendance_date < doj:
                    continue
                if emp_status != "Active" and rel and attendance_date > rel:
                    continue

                status = ws.cell(row=row, column=col).value
                status = str(status).strip().upper() if status else "A"

                ot = ws.cell(row=row + 1, column=col).value
                pdo_ot = ws.cell(row=row + 2, column=col).value
                night_ot = ws.cell(row=row + 3, column=col).value

                ot = float(ot) if ot else 0
                pdo_ot = float(pdo_ot) if pdo_ot else 0
                night_ot = float(night_ot) if night_ot else 0

                total_ot_hours += ot
                total_pdo_ot_hours += pdo_ot
                total_night_ot_hours += night_ot

                # Attendance count
                if status == "P":
                    att_status = "Present"
                    present_count += 1
                elif status == "HD":
                    att_status = "Half Day"
                    present_count += 0.5
                elif status == "RD":
                    att_status = "Absent"
                    rd_count += 1
                else:
                    att_status = "Absent"

                    if status == "E":
                        extra_days += 1
                    if status == "NH-P":
                        nh_days += 1
                    if status == "M":
                        m_days += 1

                    absent_count += 1

                rest_day = 1 if status == "RD" else 0

                # Holiday check
                holiday_list = frappe.db.get_value("Employee", emp_code, "holiday_list") \
                                or frappe.db.get_value("Company", self.company, "default_holiday_list")

                if holiday_list:
                    holiday_type = check_holiday(attendance_date, holiday_list)
                    if holiday_type == "WO":
                        ww_count += 1
                        total_ww_days += 1
                    elif holiday_type == "HH":
                        holiday_count += 1
                        total_hh_days += 1

                # ►► CREATE / UPDATE ATTENDANCE HERE
                att_name = frappe.db.get_value(
                    "Attendance",
                    {
                        "employee": emp_code,
                        "attendance_date": attendance_date,
                        "docstatus": ["!=", 2]
                    },
                    "name"
                )

                try:
                    if att_name:
                        att = frappe.get_doc("Attendance", att_name)
                        if att.docstatus == 0:
                            att.status = att_status
                            att.ot_hours = ot
                            att.custom_holiday_ot_hours = pdo_ot
                            att.custom_rest_day = rest_day
                            att.custom_upload_attendance = self.name
                            att.save(ignore_permissions=True)
                    else:
                        att = frappe.new_doc("Attendance")
                        att.employee = emp_code
                        att.attendance_date = attendance_date
                        att.company = self.company
                        att.status = att_status
                        att.ot_hours = ot
                        att.custom_holiday_ot_hours = pdo_ot
                        att.custom_rest_day = rest_day
                        att.custom_upload_attendance = self.name
                        att.insert(ignore_permissions=True)

                except Exception as e:
                    frappe.log_error(
                        f"Attendance failed: {emp_code} - {attendance_date}: {str(e)}",
                        "OETC Attendance Upload"
                    )

            # Read other earnings & deduction
            extra_start_col = date_start_col + date_cols +3

            mobile_allow = ws.cell(row=row, column=extra_start_col + 1).value or 0
            other_allowance = ws.cell(row=row, column=extra_start_col + 2).value or 0
            remarks_other_allowance = ws.cell(row=row, column=extra_start_col + 3).value or ""
            additional_allowance = ws.cell(row=row, column=extra_start_col + 4).value or 0
            other_deduction = ws.cell(row=row, column=extra_start_col + 5).value or 0
            remarks_other_deduction = ws.cell(row=row, column=extra_start_col + 6).value or ""
            medical_leave = ws.cell(row=row, column=extra_start_col + 7).value or 0
            salary_adjustment = ws.cell(row=row, column=extra_start_col + 8).value or 0
            paid_earlier = ws.cell(row=row, column=extra_start_col + 9).value or ""
            remarks_salary = ws.cell(row=row, column=extra_start_col + 10).value or ""

            # Move to next employee block
            row += 4

            worked_days = present_count + rd_count + ww_count + holiday_count

            register_name = frappe.db.get_value(
                "Attendance and OT Register",
                {
                    "employee": emp_code,
                    "from_date": self.from_date,
                    "to_date": self.to_date,
                    "docstatus": ["!=", 2]
                },
                "name"
            )

            try:
                if register_name:
                    reg = frappe.get_doc("Attendance and OT Register", register_name)
                else:
                    reg = frappe.new_doc("Attendance and OT Register")
                    reg.employee = emp_code
                    reg.from_date = self.from_date
                    reg.to_date = self.to_date
                    reg.attendance_upload = self.name

                reg.present_days = present_count
                reg.absent_days = absent_count
                reg.rest_days = rd_count
                reg.holidays = total_hh_days
                reg.total_weekly_holidays = total_ww_days
                reg.no_of_days_worked = worked_days

                reg.ot_hours = total_ot_hours
                reg.holiday_ot = total_pdo_ot_hours
                reg.night_ot = total_night_ot_hours
                reg.ot_amount = total_ot_hours*1.25
                reg.holiday_ot_amount = total_pdo_ot_hours*2
                reg.night_ot_amount = total_night_ot_hours*1.5
                reg.total_ot_hours = total_ot_hours + total_pdo_ot_hours + total_night_ot_hours

                reg.medical_leave_days = medical_leave
                reg.arrear_salary = paid_earlier

                reg.gross_days = gross_days_actual

                reg.mobile_allow = mobile_allow
                reg.other_earnings = int(other_allowance) + int(additional_allowance)
                reg.other_earnings_remark = remarks_other_allowance
                reg.other_deduction = other_deduction
                reg.other_deduction_remark = remarks_other_deduction
                reg.salary_adjustment = salary_adjustment
                reg.remarks_for_adjustment = remarks_salary

                if not register_name:
                    reg.insert(ignore_permissions=True)
                reg.save(ignore_permissions=True)

            except Exception as e:
                frappe.log_error(f"{emp_code} Import Error", str(e))

        frappe.db.commit()
        return processed_employees

        # def create_attendance_from_excel(self, ws):
        #     start_row = 6  
        #     max_row = ws.max_row
        #     max_col = ws.max_column

        #     try:
        #         dates = get_att_dates(self.from_date, self.to_date)
        #     except Exception as e:
        #         frappe.log_error("failed to get dates", "Attendance Upload")
        #         return

        #     usable_col_end = max_col - 4
        #     actual_cols = usable_col_end - 4 
        #     expected_cols = len(dates)

        #     if expected_cols != actual_cols:
        #         frappe.log_error(
        #             f"Column mismatch: expected {expected_cols} date columns for period {self.from_date} to {self.to_date}, but found {actual_cols} in Excel.",
        #             "Attendance Upload"
        #         )
        #         cols_to_process = min(expected_cols, actual_cols)
        #     else:
        #         cols_to_process = expected_cols

        #     row = start_row
        #     while row <= max_row:
        #         emp_code = ws.cell(row=row, column=2).value
        #         emp_name = ws.cell(row=row, column=3).value

        #         if not emp_code:
        #             row += 2
        #             continue
                
        #         b_count = ws.cell(row=row, column=max_col - 3).value or 0
        #         lunch = ws.cell(row=row, column=max_col - 2).value or 0
        #         dinner = ws.cell(row=row, column=max_col - 1).value or 0
        #         mobile_allowance = ws.cell(row=row, column=max_col).value or 0
        #         emp_status=frappe.db.get_value("Employee",{'name':emp_code},['status'])
                
        #         present_count = 0
        #         rd_count =0
        #         absent_count =0
        #         holiday_count = 0
        #         ww_count =0
        #         worked_days = 0
        #         total_ww_days =0
        #         total_hh_days =0
        #         for idx in range(cols_to_process):
        #             attendance_date = dates[idx]
        #             doj=frappe.db.get_value("Employee",{'name':emp_code},['date_of_joining'])
        #             rel=frappe.db.get_value("Employee",{'name':emp_code},['relieving_date'])
                
        #             if emp_status=='Active':
        #                 if getdate(attendance_date) < getdate(doj):
        #                     continue
        #             else:
        #                 if getdate(attendance_date) > getdate(rel):
        #                     continue
        #             col = 5 + idx 
        #             status = ws.cell(row=row, column=col).value
        #             ot = ws.cell(row=row + 1, column=col).value
        #             status = (str(status).strip().upper() if status else "A")
        #             ot = float(ot) if ot not in (None, '') else 0.0
        #             if status == "P":
        #                 att_status = "Present"
        #                 present_count +=1
        #                 rd_status = 0
        #             elif status == "HD":
        #                 att_status = "Half Day"
        #                 rd_status = 0
        #                 present_count +=0.5
        #             elif status == "RD":
        #                 att_status = "Absent"
        #                 rd_status = 1
        #                 rd_count +=1
        #             else:
        #                 att_status = "Absent"
        #                 rd_status = 0
        #             holiday_list = frappe.db.get_value('Employee', emp_code, 'holiday_list')
        #             if not holiday_list:
        #                 holiday_list = frappe.db.get_value('Company', self.company, 'default_holiday_list')
        #             if holiday_list and att_status !='Present' and rd_status ==0:
        #                 if check_holiday(attendance_date,holiday_list) =='WO':
        #                     if att_status == 'Half Day':
        #                         ww_count +=1
        #                     else:
        #                         ww_count +=0.5
        #                     total_ww_days +=1
        #                 elif check_holiday(attendance_date,holiday_list) =='HH':
        #                     if att_status == 'Half Day':
        #                         holiday_count +=1
        #                     else:
        #                         holiday_count +=0.5
        #                     total_hh_days +=1
        #                 else:
        #                     if att_status == 'Half Day':
        #                         absent_count +=1
        #                     else:
        #                         absent_count +=0.5


        #             if frappe.db.exists("Attendance", {
        #                 "employee": emp_code,
        #                 "attendance_date": attendance_date,
        #                 "docstatus": ["!=", 2]
        #             }):
        #                 try:
        #                     existing_att = frappe.get_doc("Attendance", {
        #                         "employee": emp_code,
        #                         "attendance_date": attendance_date
        #                     })
        #                     if existing_att.docstatus == 0:
        #                         existing_att.status = att_status
        #                         existing_att.custom_rest_day = rd_status
        #                         existing_att.custom_overtime_hours = ot
        #                         existing_att.save(ignore_permissions=True)
        #                         existing_att.submit()
        #                         frappe.db.commit()
        #                     else:
        #                         continue
        #                 except Exception as e:
        #                     frappe.log_error(
        #                         f"Failed to update existing attendance for {emp_code} on {attendance_date}: {str(e)}",
        #                         "Attendance Upload"
        #                     )
        #             else:
        #                 try:
        #                     att = frappe.new_doc("Attendance")
        #                     att.employee = emp_code
        #                     att.employee_name = emp_name
        #                     att.attendance_date = attendance_date
        #                     att.custom_upload_attendance = self.name
        #                     att.custom_overtime_hours = ot
        #                     att.status = att_status
        #                     att.custom_rest_day = rd_status
        #                     att.insert(ignore_permissions=True)
        #                     att.submit()
        #                     frappe.db.commit()
        #                 except Exception as e:
        #                     frappe.log_error(
        #                         f"Error creating attendance for {emp_code} on {attendance_date.strftime('%d-%m-%Y')}: {str(e)}",
        #                         "Attendance Upload"
        #                     )
        #         row += 2


        #         register_name = frappe.db.get_value(
        #             "Attendance and OT Register",
        #             {
        #                 'from_date': ('<=', self.to_date),
        #                 'to_date': ('>=', self.from_date),
        #                 'employee': emp_code,
        #                 'docstatus':['!=',2]
        #             },
        #             'name'
        #         )
        #         worked_days = present_count + rd_count + ww_count + holiday_count
        #         if register_name:
        #             try:
        #                 register_doc = frappe.get_doc("Attendance and OT Register", register_name)
        #                 register_doc.breakfast = float(b_count)
        #                 register_doc.lunch = float(lunch)
        #                 register_doc.dinner = float(dinner)
        #                 register_doc.mobile_allow = float(mobile_allowance)
        #                 register_doc.attendance_upload=self.name
        #                 register_doc.rest_days = rd_count
        #                 register_doc.present_days = present_count
        #                 register_doc.absent_days = absent_count
        #                 register_doc.total_weekly_holidays = total_ww_days
        #                 register_doc.holidays = total_hh_days
        #                 register_doc.no_of_days_worked = worked_days
        #                 register_doc.save(ignore_permissions=True)
        #                 frappe.db.commit()
        #             except Exception as e:
        #                 frappe.log_error(
        #                     f"Failed to update 'Attendance and OT Register' for {emp_code}: {str(e)}",
        #                     "Attendance Upload"
        #                 )

        def create_attendance_from_excel(self, ws):
            start_row = 6
            max_row = ws.max_row
            max_col = ws.max_column

            try:
                dates = get_att_dates(self.from_date, self.to_date)
            except Exception as e:
                frappe.log_error("Failed to get dates", "Attendance Upload")
                return

            usable_col_end = max_col - 15  # 15 extra columns at the end
            actual_cols = usable_col_end - 5  # 5 fixed columns before date columns
            expected_cols = len(dates)

            if expected_cols != actual_cols:
                frappe.log_error(
                    f"Column mismatch: expected {expected_cols} date columns for period {self.from_date} to {self.to_date}, but found {actual_cols} in Excel.",
                    "Attendance Upload"
                )
                cols_to_process = min(expected_cols, actual_cols)
            else:
                cols_to_process = expected_cols

            row = start_row
            while row <= max_row:
                emp_code = ws.cell(row=row, column=2).value
                emp_name = ws.cell(row=row, column=3).value

                if not emp_code:
                    row += 3  
                    continue

                emp_status = frappe.db.get_value("Employee", {'name': emp_code}, ['status'])

                present_count = 0
                rd_count = 0
                absent_count = 0
                holiday_count = 0
                ww_count = 0
                total_ww_days = 0
                total_hh_days = 0
                total_ot_hours = 0
                total_pdo_ot_hours = 0
                extra_days=0
                nh_days=0
                m_days=0
                # Loop through dates
                for idx in range(cols_to_process):
                    attendance_date = dates[idx]
                    doj = frappe.db.get_value("Employee", {'name': emp_code}, ['date_of_joining'])
                    rel = frappe.db.get_value("Employee", {'name': emp_code}, ['relieving_date'])

                    if emp_status == 'Active' and getdate(attendance_date) < getdate(doj):
                        continue
                    if emp_status != 'Active' and rel and getdate(attendance_date) > getdate(rel):
                        continue
                    e_type = frappe.db.get_value("Employee", {'name': emp_code}, ['custom_nationality'])
                    col = 6 + idx  # Date columns start at column 6 (after fixed 5)

                    # Attendance status
                    status = ws.cell(row=row, column=col).value
                    status = (str(status).strip().upper() if status else "A")

                    # OT HR
                    ot = ws.cell(row=row + 1, column=col).value
                    ot = float(ot) if ot not in (None, '') else 0.0

                    # PDO OT HR
                    pdo_ot = ws.cell(row=row + 2, column=col).value
                    pdo_ot = float(pdo_ot) if pdo_ot not in (None, '') else 0.0

                    total_ot_hours += ot
                    total_pdo_ot_hours += pdo_ot

                    # Count attendance
                    if status == "P":
                        att_status = "Present"
                        present_count += 1
                        rd_status = 0
                    elif status == "HD":
                        att_status = "Half Day"
                        present_count += 0.5
                        rd_status = 0
                    elif status == "RD":
                        att_status = "Absent"
                        rd_count += 1
                        rd_status = 1
                    else:
                        if status=='E':
                            extra_days+=1
                        if status=='NH-P':
                            nh_days+=1
                        if status=='M':
                            m_days+=1
                        att_status = "Absent"
                        rd_status = 0

                    # Check holidays
                    holiday_list = frappe.db.get_value('Employee', emp_code, 'holiday_list')
                    if not holiday_list:
                        holiday_list = frappe.db.get_value('Company', self.company, 'default_holiday_list')

                    if holiday_list and att_status != 'Present' and rd_status == 0:
                        holiday_type = check_holiday(attendance_date, holiday_list)
                        if holiday_type == 'WO':
                            ww_count += 0.5 if att_status == 'Half Day' else 1
                            total_ww_days += 1
                        elif holiday_type == 'HH':
                            holiday_count += 0.5 if att_status == 'Half Day' else 1
                            total_hh_days += 1
                        else:
                            absent_count += 0.5 if att_status == 'Half Day' else 1
                    if status=='M':
                        if e_type:
                            if e_type=="Nationality":
                                absent_count+=2
                            else:
                                absent_count+=1
                        

                    # Update or create Attendance
                    if frappe.db.exists("Attendance", {
                        "employee": emp_code,
                        "attendance_date": attendance_date,
                        "docstatus": ["!=", 2]
                    }):
                        try:
                            existing_att = frappe.get_doc("Attendance", {
                                "employee": emp_code,
                                "attendance_date": attendance_date
                            })
                            if existing_att.docstatus == 0:
                                existing_att.status = att_status
                                existing_att.custom_rest_day = rd_status
                                existing_att.custom_overtime_hours = ot
                                existing_att.custom_pdo_overtime_hours = pdo_ot
                                existing_att.save(ignore_permissions=True)
                                existing_att.submit()
                                frappe.db.commit()
                        except Exception as e:
                            frappe.log_error(
                                
                                "Attendance Upload",f"Failed to update existing attendance for {emp_code} on {attendance_date}"
                            )
                    else:
                        try:
                            att = frappe.new_doc("Attendance")
                            att.employee = emp_code
                            att.employee_name = emp_name
                            att.attendance_date = attendance_date
                            att.custom_upload_attendance = self.name
                            att.custom_overtime_hours = ot
                            att.custom_pdo_overtime_hours = pdo_ot
                            att.status = att_status
                            att.custom_rest_day = rd_status
                            att.insert(ignore_permissions=True)
                            att.submit()
                            frappe.db.commit()
                        except Exception as e:
                            frappe.log_error(
                                
                                "Attendance Upload",f"Error creating attendance for {emp_code} on {attendance_date.strftime('%d-%m-%Y')}"
                            )

                # --- Read extra columns after date columns ---
                extra_start_col = 5 + len(dates)+3
                gross_row = row  # Correct row for Gross/Allowance values
                frappe.log_error(
                        
                        "Attendance row",gross_row
                    )
                gross_days = ws.cell(row=gross_row, column=extra_start_col + 1).value or 0
                breakfast = ws.cell(row=gross_row, column=extra_start_col + 2).value or 0
                lunch = ws.cell(row=gross_row, column=extra_start_col + 3).value or 0
                dinner = ws.cell(row=gross_row, column=extra_start_col + 4).value or 0
                rate_per_extra_day = ws.cell(row=gross_row, column=extra_start_col + 5).value or 0
                rate_per_day_nh = ws.cell(row=gross_row, column=extra_start_col + 6).value or 0
                mobile_allow = ws.cell(row=gross_row, column=extra_start_col + 7).value or 0
                other_allowance = ws.cell(row=gross_row, column=extra_start_col + 8).value or 0
                remarks_other_allowance = ws.cell(row=gross_row, column=extra_start_col + 9).value or ""
                other_deduction = ws.cell(row=gross_row, column=extra_start_col + 10).value or 0
                remarks_other_deduction = ws.cell(row=gross_row, column=extra_start_col + 11).value or ""
                salary_adjustment = ws.cell(row=gross_row, column=extra_start_col + 12).value or 0
                arrear_salary = ws.cell(row=gross_row, column=extra_start_col + 13).value or 0
                remarks_paid_deducted = ws.cell(row=gross_row, column=extra_start_col + 14).value or ""

                row += 3

                # --- Update or create Attendance and OT Register ---
                register_name = frappe.db.get_value(
                    "Attendance and OT Register",
                    {
                        'from_date': self.to_date,
                        'to_date': self.from_date,
                        'employee': emp_code,
                        'docstatus': ['!=', 2]
                    },
                    'name'
                )

                worked_days = present_count + rd_count + ww_count + holiday_count

                try:
                    if register_name:
                        register_doc = frappe.get_doc("Attendance and OT Register", register_name)
                    else:
                        register_doc = frappe.new_doc("Attendance and OT Register")
                        register_doc.employee = emp_code
                        register_doc.from_date = self.from_date
                        register_doc.to_date = self.to_date
                        register_doc.attendance_upload = self.name
                    
                    register_doc.extra_days = extra_days*rate_per_extra_day
                    register_doc.nh_days = nh_days*rate_per_day_nh
                    register_doc.gross_days = gross_days
                    register_doc.breakfast = breakfast
                    register_doc.lunch = lunch
                    register_doc.dinner = dinner
                    register_doc.extra_rate_per_day = rate_per_extra_day
                    register_doc.nh_rate_per_day = rate_per_day_nh
                    register_doc.mobile_allow = mobile_allow
                    register_doc.other_earnings = other_allowance
                    register_doc.other_earnings_remark = remarks_other_allowance
                    register_doc.other_deduction = other_deduction
                    register_doc.other_deduction_remark = remarks_other_deduction
                    register_doc.medical_leave_days = m_days
                    register_doc.salary_adjustment = salary_adjustment
                    register_doc.arrear_salary = arrear_salary
                    register_doc.remarks_for_adjustment = remarks_paid_deducted
                    register_doc.food_allowance=breakfast+(lunch*2)+(dinner*2)
                    register_doc.rest_days = rd_count
                    register_doc.present_days = present_count
                    register_doc.absent_days = absent_count
                    register_doc.total_weekly_holidays = total_ww_days
                    register_doc.holidays = total_hh_days
                    register_doc.no_of_days_worked = worked_days
                    register_doc.ot_hours = total_ot_hours
                    register_doc.pdo_ot = total_pdo_ot_hours
                    register_doc.total_ot_hours = total_ot_hours + total_pdo_ot_hours

                    if not register_name:
                        register_doc.insert(ignore_permissions=True)
                    register_doc.save(ignore_permissions=True)
                    frappe.db.commit()

                except Exception as e:
                    frappe.log_error(
                        
                        "Attendance Upload",f"Failed to create/update 'Attendance and OT Register' for {emp_code}: {str(e)}"
                    )
        # def create_attendance_from_excel_pdo(self, ws):
        #     start_row = 6
        #     max_row = ws.max_row
        #     max_col = ws.max_column
        #     processed_employees = []

        #     try:
        #         dates = get_att_dates(self.from_date, self.to_date)
        #     except Exception as e:
        #         frappe.log_error("Failed to get dates", "Attendance Upload")
        #         return

        #     usable_col_end = max_col - 15
        #     actual_cols = usable_col_end - 5  
        #     expected_cols = len(dates)

        #     cols_to_process = expected_cols if expected_cols == actual_cols else min(expected_cols, actual_cols)

        #     row = start_row
        #     while row <= max_row:
        #         emp_code = ws.cell(row=row, column=2).value
        #         if not emp_code:
        #             row += 3
        #             continue

        #         processed_employees.append(emp_code)

        #         # Fetch Employee general info once
        #         doj, rel, emp_status = frappe.db.get_value(
        #             "Employee", emp_code, ["date_of_joining", "relieving_date", "status"]
        #         )

        #         for idx in range(cols_to_process):
        #             attendance_date = dates[idx]
        #             col = 6 + idx

        #             # DOJ/Relieving filter
        #             if emp_status == 'Active' and getdate(attendance_date) < getdate(doj):
        #                 continue
        #             if emp_status != 'Active' and rel and getdate(attendance_date) > getdate(rel):
        #                 continue

        #             # Main Status
        #             status = ws.cell(row=row, column=col).value
        #             status = (str(status).strip().upper() if status else "A")

        #             # OT Hours
        #             ot = ws.cell(row=row + 1, column=col).value
        #             ot = float(ot) if ot not in (None, "") else 0.0

                
        #             status_map = {
        #                 "P": "Present",
        #                 "HD": "Half Day",
        #                 "RD": "On Leave",  # Change to Absent if required
        #                 "E": "Present",
        #                 "NH-P": "Present",
        #                 "M": "On Leave",
        #             }
        #             att_status = status_map.get(status, "Absent")

        #             # Check existing attendance
        #             att_name = frappe.db.get_value(
        #                 "Attendance",
        #                 {
        #                     "employee": emp_code,
        #                     "attendance_date": attendance_date,
        #                     "docstatus": ["!=", 2]
        #                 },
        #                 "name"
        #             )

        #             try:
        #                 if att_name:
        #                     att = frappe.get_doc("Attendance", att_name)
        #                     if att.docstatus == 0:
        #                         att.status = att_status
        #                         att.custom_overtime_hours = ot
        #                         att.custom_upload_attendance = self.name
        #                         att.save(ignore_permissions=True)
        #                 else:
        #                     att = frappe.new_doc("Attendance")
        #                     att.employee = emp_code
        #                     att.attendance_date = attendance_date
        #                     att.company = self.company
        #                     att.status = att_status
        #                     att.custom_overtime_hours = ot
        #                     att.custom_upload_attendance = self.name
        #                     att.insert(ignore_permissions=True)

        #             except Exception as e:
        #                 frappe.log_error(
        #                     f"Attendance creation/update failed: {emp_code} - {attendance_date}: {str(e)}",
        #                     "Attendance Upload"
        #                 )

        #         row += 3

        #     frappe.db.commit()
        #     return processed_employees



    def create_att_ot_in_draft_pdo(self, ws):
        from openpyxl import load_workbook

        start_row = 6
        max_row = ws.max_row
        max_col = ws.max_column
        processed_employees = []
        try:
            dates = get_att_dates(self.from_date, self.to_date)
        except Exception as e:
            frappe.log_error("Failed to get dates", "Attendance Upload")
            return

        usable_col_end = max_col - 27  # 15 extra columns at the end
        actual_cols = usable_col_end - 5  # 5 fixed columns before date columns
        expected_cols = len(dates)

        if expected_cols != actual_cols:
            frappe.log_error(
                f"Column mismatch: expected {expected_cols} date columns for period {self.from_date} to {self.to_date}, but found {actual_cols} in Excel.",
                "Attendance Upload"
            )
            cols_to_process = min(expected_cols, actual_cols)
        else:
            cols_to_process = expected_cols

        row = start_row
        while row <= max_row:
            emp_code = ws.cell(row=row, column=2).value
            emp_name = ws.cell(row=row, column=3).value

            if not emp_code:
                row += 3  # Skip Attendance + OT + PDO OT rows
                continue
            processed_employees.append(emp_code)
            emp_status = frappe.db.get_value("Employee", {'name': emp_code}, ['status'])

            present_count = 0
            rd_count = 0
            absent_count = 0
            holiday_count = 0
            ww_count = 0
            total_ww_days = 0
            total_hh_days = 0
            total_ot_hours = 0
            total_pdo_ot_hours = 0
            extra_days=0
            nh_days=0
            m_days=0
            l_days=0
            em_days=0
            # Loop through dates
            for idx in range(cols_to_process):
                attendance_date = dates[idx]
                doj = frappe.db.get_value("Employee", {'name': emp_code}, ['date_of_joining'])
                rel = frappe.db.get_value("Employee", {'name': emp_code}, ['relieving_date'])
                e_type=frappe.db.get_value("Employee", {'name': emp_code}, ['custom_nationality_type'])
                if emp_status == 'Active' and getdate(attendance_date) < getdate(doj):
                    continue
                if emp_status != 'Active' and rel and getdate(attendance_date) > getdate(rel):
                    continue

                col = 6 + idx  # Date columns start at column 6 (after fixed 5)

                # Attendance status
                status = ws.cell(row=row, column=col).value
                status = (str(status).strip().upper() if status else "A")

                # OT HR
                ot = ws.cell(row=row + 1, column=col).value
                ot = float(ot) if ot not in (None, '') else 0.0

                # PDO OT HR
                pdo_ot = ws.cell(row=row + 2, column=col).value
                pdo_ot = float(pdo_ot) if pdo_ot not in (None, '') else 0.0

                total_ot_hours += ot
                total_pdo_ot_hours += pdo_ot

                # Count attendance
                if status == "P":
                    att_status = "Present"
                    present_count += 1
                    rd_status = 0
                elif status == "HD":
                    att_status = "Half Day"
                    present_count += 0.5
                    rd_status = 0
                elif status == "RD":
                    att_status = "Absent"
                    rd_count += 1
                    rd_status = 1
                else:
                    if status=='E':
                        extra_days+=1
                    if status=='NH-P':
                        nh_days+=1
                    if status=='M':
                        m_days+=1
                    if status=='L':
                        l_days+=1
                    if status=='EM':
                        em_days+=1
                    
                    att_status = "Absent"
                    rd_status = 0

                # Check holidays
                holiday_list = frappe.db.get_value('Employee', emp_code, 'holiday_list')
                if not holiday_list:
                    holiday_list = frappe.db.get_value('Company', self.company, 'default_holiday_list')

                if holiday_list and att_status != 'Present' and rd_status == 0:
                    holiday_type = check_holiday(attendance_date, holiday_list)
                    if holiday_type == 'WO':
                        ww_count += 0.5 if att_status == 'Half Day' else 1
                        total_ww_days += 1
                    elif holiday_type == 'HH':
                        holiday_count += 0.5 if att_status == 'Half Day' else 1
                        total_hh_days += 1
                    else:
                        absent_count += 0.5 if att_status == 'Half Day' else 1
                if status=='M':
                    if e_type:
                        if e_type=="Nationality":
                            absent_count+=2
                        else:
                            absent_count+=1
                
            extra_start_col = 5 + len(dates)+1
            gross_row = row  
            # frappe.log_error(
                    
            #         "Attendance row",gross_row
            #     )
           
            
            extra_day_rate = ws.cell(row=gross_row, column=extra_start_col + 5).value or 0
            rate_per_day_nh = ws.cell(row=gross_row, column=extra_start_col + 13).value or 0
            breakfast = ws.cell(row=gross_row, column=extra_start_col + 15).value or 0
            lunch = ws.cell(row=gross_row, column=extra_start_col + 16).value or 0
            dinner = ws.cell(row=gross_row, column=extra_start_col + 17).value or 0
            mobile_allow = ws.cell(row=gross_row, column=extra_start_col + 19).value or 0
            other_allowance = ws.cell(row=gross_row, column=extra_start_col + 20).value or 0
            remarks_other_allowance = ws.cell(row=gross_row, column=extra_start_col + 21).value or ""
            other_deduction = ws.cell(row=gross_row, column=extra_start_col + 22).value or 0
            remarks_other_deduction = ws.cell(row=gross_row, column=extra_start_col + 23).value or ""
            salary_adjustment = ws.cell(row=gross_row, column=extra_start_col + 24).value or 0
            arrear_salary = ws.cell(row=gross_row, column=extra_start_col + 25).value or 0
            remarks_paid_deducted = ws.cell(row=gross_row, column=extra_start_col + 26).value or ""

            row += 3

            register_name = frappe.db.get_value(
                "Attendance and OT Register",
                {
                    "employee": emp_code,
                    "from_date": self.from_date,
                    "to_date": self.to_date,
                    "docstatus": ["!=", 2]
                },
                "name"
            )

            try:
                if register_name:
                    register_doc = frappe.get_doc("Attendance and OT Register", register_name)
                else:
                    register_doc = frappe.new_doc("Attendance and OT Register")
                    register_doc.employee = emp_code
                    register_doc.from_date = self.from_date
                    register_doc.to_date = self.to_date
                    register_doc.attendance_upload = self.name
                
                register_doc.ot_hours = total_ot_hours
                register_doc.pdo_ot=total_pdo_ot_hours
                register_doc.total_ot_hours=total_ot_hours+total_pdo_ot_hours

                register_doc.extra_days = extra_days
                register_doc.extra_rate_per_day = extra_day_rate or 0
                register_doc.extra_days_amount = extra_days*extra_day_rate
                register_doc.rest_days = rd_count
                register_doc.absent_days = absent_count
                register_doc.medical_leave_days = m_days
                register_doc.leave_days = l_days
                register_doc.el_days = em_days
                register_doc.nh_days = nh_days
                register_doc.nh_rate_per_day = rate_per_day_nh or 0
                register_doc.nh_days_amount = rate_per_day_nh*nh_days
                register_doc.breakfast = breakfast
                register_doc.lunch = lunch
                register_doc.dinner = dinner
                register_doc.food_allowance = (dinner*2)+(lunch*2)+breakfast
                register_doc.mobile_allow = mobile_allow
                register_doc.other_earnings = other_allowance
                register_doc.other_earnings_remark = remarks_other_allowance
                register_doc.other_deduction = other_deduction
                register_doc.other_deduction_remark = remarks_other_deduction
                register_doc.salary_adjustment=salary_adjustment
                register_doc.arrear_salary=arrear_salary
                register_doc.remarks_for_adjustment=remarks_paid_deducted

                if not register_name:
                    register_doc.insert(ignore_permissions=True)
                register_doc.save(ignore_permissions=True)
                frappe.db.commit()

            except Exception as e:
                frappe.log_error(
                    
                    "Attendance Upload",f"Failed to create/update 'Attendance and OT Register' for {emp_code}: {str(e)}"
                )
        return processed_employees
        
        
    def create_attendance_from_excel_pdo(self, ws):
        start_row = 6
        date_col_start = 6
        max_row = ws.max_row
        max_col = ws.max_column
        processed_employees = []

        try:
            dates = get_att_dates(self.from_date, self.to_date)
        except Exception:
            frappe.log_error("Failed to get dates", "Attendance Upload")
            return []

        row = start_row
        while row <= max_row:
            emp_code = ws.cell(row=row, column=2).value
            if not emp_code:
                row += 3
                continue

            processed_employees.append(emp_code)

            emp_status, doj, rel, e_type, holiday_list = frappe.db.get_value(
                "Employee",
                emp_code,
                ["status", "date_of_joining", "relieving_date", "custom_nationality_type", "holiday_list"]
            )

            if not holiday_list:
                holiday_list = frappe.db.get_value("Company", self.company, "default_holiday_list")

            present_count = 0
            rd_count = 0
            absent_count = 0
            holiday_count = 0
            ww_count = 0
            total_ww_days = 0
            total_hh_days = 0
            total_ot_hours = 0
            total_pdo_ot_hours = 0
            extra_days = 0
            nh_days = 0
            m_days = 0
            l_days=0
            em_days=0
            for idx, att_date in enumerate(dates):
                col = date_col_start + idx
                if col > max_col:
                    continue

                status_cell = ws.cell(row=row, column=col).value
                status = (str(status_cell).strip().upper() if status_cell else "A")

                ot_cell = ws.cell(row=row + 1, column=col).value
                pdo_cell = ws.cell(row=row + 2, column=col).value
                ot = float(ot_cell) if ot_cell not in (None, "") else 0.0
                pdo_ot = float(pdo_cell) if pdo_cell not in (None, "") else 0.0

                total_ot_hours += ot
                total_pdo_ot_hours += pdo_ot

                if emp_status == "Active" and doj and getdate(att_date) < getdate(doj):
                    continue
                if emp_status != "Active" and rel and getdate(att_date) > getdate(rel):
                    continue

                rest_day = 0
                if status == "P":
                    att_status = "Present"
                    present_count += 1
                elif status == "HD":
                    att_status = "Half Day"
                    present_count += 0.5
                elif status == "RD":
                    att_status = "Absent"
                    rd_count += 1
                    rest_day = 1
                else:
                    if status=='E':
                        extra_days+=1
                    if status=='NH-P':
                        nh_days+=1
                    if status=='M':
                        m_days+=1
                    if status=='L':
                        l_days+=1
                    if status=='EM':
                        em_days+=1
                    att_status = "Absent"
                    

                # Holiday logic
                if holiday_list and att_status != "Present" and rest_day == 0:
                    h_type = check_holiday(att_date, holiday_list)
                    if h_type == 'WO':
                        ww_count += 1
                        total_ww_days += 1
                    elif h_type == 'HH':
                        holiday_count += 1
                        total_hh_days += 1
                    else:
                        absent_count += 1

                if status == "M":
                    absent_count += 2 if e_type == "Nationality" else 1

                # ⚡ ATTENDANCE CREATE/UPDATE HERE ⚡
                att_name = frappe.db.get_value(
                    "Attendance",
                    {"employee": emp_code, "attendance_date": att_date, "docstatus": ["!=", 2]},
                    "name"
                )

                try:
                    if att_name:
                        att = frappe.get_doc("Attendance", att_name)
                        if att.docstatus == 0:
                            att.status = att_status
                            att.ot_hours = ot
                            att.custom_pdo_ot = pdo_ot
                            att.custom_rest_day = rest_day
                            att.custom_upload_attendance = self.name
                            att.save(ignore_permissions=True)
                    else:
                        att = frappe.new_doc("Attendance")
                        att.employee = emp_code
                        att.attendance_date = att_date
                        att.company = self.company
                        att.status = att_status
                        att.ot_hours = ot
                        att.custom_pdo_ot = pdo_ot
                        att.custom_rest_day = rest_day
                        att.custom_upload_attendance = self.name
                        att.insert(ignore_permissions=True)

                except Exception as e:
                    frappe.log_error(
                        f"Attendance creation/update failed: {emp_code} - {att_date}: {str(e)}",
                        "Attendance Upload"
                    )

            # Process Register (same as your code)
            extra_start_col = 5 + len(dates)+3
            gross_row = row

            extra_day_rate = ws.cell(row=gross_row, column=extra_start_col + 5).value or 0
            rate_per_day_nh = ws.cell(row=gross_row, column=extra_start_col + 13).value or 0
            breakfast = ws.cell(row=gross_row, column=extra_start_col + 15).value or 0
            lunch = ws.cell(row=gross_row, column=extra_start_col + 16).value or 0
            dinner = ws.cell(row=gross_row, column=extra_start_col + 17).value or 0
            mobile_allow = ws.cell(row=gross_row, column=extra_start_col + 19).value or 0
            other_allowance = ws.cell(row=gross_row, column=extra_start_col + 20).value or 0
            remarks_other_allowance = ws.cell(row=gross_row, column=extra_start_col + 21).value or ""
            other_deduction = ws.cell(row=gross_row, column=extra_start_col + 22).value or 0
            remarks_other_deduction = ws.cell(row=gross_row, column=extra_start_col + 23).value or ""
            salary_adjustment = ws.cell(row=gross_row, column=extra_start_col + 24).value or 0
            arrear_salary = ws.cell(row=gross_row, column=extra_start_col + 25).value or 0
            remarks_paid_deducted = ws.cell(row=gross_row, column=extra_start_col + 26).value or ""

            register_name = frappe.db.get_value(
                "Attendance and OT Register",
                {"employee": emp_code, "from_date": self.from_date, "to_date": self.to_date, "docstatus": ["!=", 2]},
                "name"
            )

            worked_days = present_count + rd_count + ww_count + holiday_count

            try:
                if register_name:
                    reg = frappe.get_doc("Attendance and OT Register", register_name)
                    
                else:
                    reg = frappe.new_doc("Attendance and OT Register")
                    reg.employee = emp_code
                    reg.from_date = self.from_date
                    reg.to_date = self.to_date
                    reg.attendance_upload = self.name

                reg.ot_hours = total_ot_hours
                reg.pdo_ot=total_pdo_ot_hours
                reg.total_ot_hours=total_ot_hours+total_pdo_ot_hours
                reg.no_of_days_worked=worked_days
                reg.extra_days = extra_days
                reg.extra_rate_per_day = extra_day_rate or 0
                reg.extra_days_amount = extra_days*extra_day_rate
                reg.rest_days = rd_count
                reg.absent_days = absent_count
                reg.medical_leave_days = m_days
                reg.leave_days = l_days
                reg.el_days = em_days
                reg.nh_days = nh_days
                reg.nh_rate_per_day = rate_per_day_nh or 0
                reg.nh_days_amount = rate_per_day_nh*nh_days
                reg.breakfast = breakfast
                reg.lunch = lunch
                reg.dinner = dinner
                reg.food_allowance = (dinner*2)+(lunch*2)+breakfast
                reg.mobile_allow = mobile_allow
                reg.other_earnings = other_allowance
                reg.other_earnings_remark = remarks_other_allowance
                reg.other_deduction = other_deduction
                reg.other_deduction_remark = remarks_other_deduction
                reg.salary_adjustment=salary_adjustment
                reg.arrear_salary=arrear_salary
                reg.remarks_for_adjustment=remarks_paid_deducted

                if not register_name:
                    reg.insert(ignore_permissions=True)
                if register_name and reg.docstatus==1:
                    pass
                else:
                    reg.save(ignore_permissions=True)

            except Exception as e:
                frappe.log_error(
                    f"Register create/update failed for {emp_code}: {str(e)}",
                    "Attendance Upload"
                )

            row += 3

        frappe.db.commit()
        return processed_employees
    
@frappe.whitelist()
def process_attendance_from_excel_job(docname):
    doc = frappe.get_doc("Upload Employee Attendance", docname)

    file_url = frappe.db.get_value(
        "File",
        {"attached_to_doctype": "Upload Employee Attendance", "attached_to_name": doc.name},
        "file_url"
    )

    if not file_url:
        frappe.log_error("No file found", "Attendance Upload Background Job")
        return

    try:
        in_memory_file = download_external_file_url(file_url)
        wb = openpyxl.load_workbook(in_memory_file)
        ws = wb.active
        if "PDO" in doc.division:
            doc.create_attendance_from_excel_pdo(ws)
        else:
            doc.create_attendance_from_excel_oetc(ws)
    except Exception as e:
        frappe.log_error(f"Failed to process attendance in background: {str(e)}", "Attendance Upload")

@frappe.whitelist()
def process_attendance_from_excel_job_pdo(docname,docstatus):
    doc = frappe.get_doc("Upload Employee Attendance", docname)

    file_url = frappe.db.get_value(
        "File",
        {"attached_to_doctype": "Upload Employee Attendance", "attached_to_name": doc.name},
        "file_url"
    )

    if not file_url:
        frappe.log_error("No file found", "Attendance Upload Background Job")
        return

    try:
        in_memory_file = download_external_file_url(file_url)
        wb = openpyxl.load_workbook(in_memory_file)
        ws = wb.active
        if docstatus==0:
            doc.create_att_ot_in_draft_pdo(ws)
        else:
            doc.create_attendance_from_excel_pdo(ws)
    except Exception as e:
        frappe.log_error(f"Failed to process attendance in background: {str(e)}", "Attendance Upload")
@frappe.whitelist()
def process_attendance_from_excel_job_oetc(docname,docstatus):
    doc = frappe.get_doc("Upload Employee Attendance", docname)

    file_url = frappe.db.get_value(
        "File",
        {"attached_to_doctype": "Upload Employee Attendance", "attached_to_name": doc.name},
        "file_url"
    )

    if not file_url:
        frappe.log_error("No file found", "Attendance Upload Background Job")
        return

    try:
        in_memory_file = download_external_file_url(file_url)
        wb = openpyxl.load_workbook(in_memory_file)
        ws = wb.active
        if docstatus==0:
            doc.create_attendance_from_excel_oetc(ws)
        else:
            doc.create_attendance_from_excel_oetc_sub(ws)
    except Exception as e:
        frappe.log_error(f"Failed to process attendance in background: {str(e)}", "Attendance Upload")

@frappe.whitelist()
def get_att_dates(from_date, to_date):
        no_of_days = date_diff(add_days(to_date, 1), from_date)
        dates = [add_days(from_date, i) for i in range(no_of_days)]
        return dates

def download_external_file_url(file_url):
    path = frappe.get_site_path("public", file_url.replace("/files/", ""))
    with open(path, "rb") as f:
        return BytesIO(f.read())


def get_employee_count(ws):
    # Count how many rows contain employee codes (assuming every 2nd row is OT)
    return sum(
        1 for r in range(6, ws.max_row + 1, 2)
        if ws.cell(row=r, column=2).value
    )

from frappe.utils import get_url
@frappe.whitelist()
def download_external_file_url(file_url):
    if not file_url.startswith("http"):
        file_url = get_url() + file_url
    response = requests.get(file_url, stream=True)
    if response.status_code != 200:
        frappe.throw(_("Failed to download file from external storage."))

    return BytesIO(response.content)

@frappe.whitelist()
def get_template(from_date, to_date, department=None, designation=None, site_location=None, name=None):
    filename = 'Attendance Template'
    test = build_xlsx_response(filename)

@frappe.whitelist()
def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    dates = get_dates(args.from_date, args.to_date)
    comp=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['company'])
    header1=[comp]
    ws.append(header1)
    site_location=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['site_location'])
    if not site_location:
        header12=[]
    else:
        header12=[site_location]
    ws.append(header12)
    from_date_str = datetime.strptime(args.from_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    to_date_str = datetime.strptime(args.to_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    date_str=f"TIME SHEET {from_date_str} to {to_date_str}"
    header2=[date_str]
    ws.append(header2)
    header3 = ['SI No.','Emp. Code','Employee Name','Designation']
    header3.extend([d[0] for d in dates])  
    header3.extend(['Food','','','Mobile Allowance'])
    header4 = ['', '', '', '']  
    header4.extend([d[1] for d in dates]) 
    header4.extend(['B','L','D'])
    ws.append(header3)
    ws.append(header4)

    emp_data = get_employees(args)
    row_num = 1
    start_row = 6
    for emp in emp_data:
        emp_row = [row_num,emp.get("name"),emp.get("employee_name"),emp.get("designation")]
        empty_line=[]
        empty_line.extend([''] * len(dates)) 

        ws.append(emp_row)
        ws.append(empty_line)
        for col in range(1, 5):
            ws.merge_cells(start_row=start_row, start_column=col,end_row=start_row + 1, end_column=col)
        row_num += 1
        start_row+=2
    
    merge_length=date_diff(args.to_date,args.from_date)
    merge_length+=5
    
    if not site_location:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=merge_length+4)
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_length+4)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_length+4)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=merge_length+4)
    ws.merge_cells(start_row=4, start_column=merge_length+1, end_row=4, end_column=merge_length+3)
    for i in range(1,5):
        ws.merge_cells(start_row=4,start_column=i,end_row=5,end_column=i)
    header_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    for row in [4, 5]:
        for col in range(1, len(header3) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_alignment
    friday_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type="solid")  
    for col_idx, cell in enumerate(ws[5], start=1): 
        if cell.value == 'Fr':
            ws.cell(row=5, column=col_idx).fill = friday_fill
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 43
    ws.column_dimensions['D'].width = 18
    last_col_index = ws.max_column
    from openpyxl.utils import get_column_letter

    last_col_letter = get_column_letter(last_col_index)
    ws.column_dimensions[last_col_letter].width = 18
    for i in range(1, 4): 
        col_letter = get_column_letter(last_col_index - i)
        ws.column_dimensions[col_letter].width = 6
    for row in range(4, ws.max_row + 1, 2):
        ws.merge_cells(
            start_row=row,
            start_column=last_col_index,
            end_row=row + 1,
            end_column=last_col_index
        )
    for col_offset in range(1, 4):  # 1 to 3
        col_idx = last_col_index - col_offset
        for row in range(6, ws.max_row + 1, 2):
            ws.merge_cells(
                start_row=row,
                start_column=col_idx,
                end_row=row + 1,
                end_column=col_idx
            )
    from openpyxl.utils import get_column_letter

    # Columns to check
    target_headers = [
        "Gross Days",
        "Mobile Allowance",
        "Other Allowance",
        "Remarks - Other Allowance",
        "Additional Allowance",
        "Other Deduction",
        "Remarks - Other Deduction",
        "Medical Leave",
        "Salary Adjustment",
        "Paid Earlier",
        "Remarks for Salary",
        "Net Days"
    ]

    # Column number where header may appear (3rd column)
    col_num = 3

    # Read value from the 3rd column header row
    header_value = ws.cell(row=1, column=col_num).value

    # If matches any of those headers, increase width
    if header_value in target_headers:
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = 20

    border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1,max_col=ws.max_column):
        for cell in rows:
            cell.border = border
    for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1,max_col=ws.max_column):
        for cell in rows:
            cell.alignment = center_alignment
            cell.font = bold_font
    for col in range(5, merge_length + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 6

    xlsx_file=BytesIO()
    wb.save(xlsx_file)
    return xlsx_file





def get_dates(from_date, to_date):
    start_date = datetime.strptime(from_date, "%Y-%m-%d")
    end_date = datetime.strptime(to_date, "%Y-%m-%d")
    delta = end_date - start_date
    days = []
    for i in range(delta.days + 1):
        date = start_date + timedelta(days=i)
        day_num = date.strftime("%d")  
        weekday = date.strftime("%a")[:2]
        days.append((day_num, weekday))
    return days


def get_employees(args):
    filters = {'status': 'Active'}
    dept=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['department'])
    desi=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['designation'])
    comp=frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['company'])
    location = frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['site_location'])
    division = frappe.db.get_value("Upload Employee Attendance",{'name':args.name},['division'])
    if args.department and args.department!='None':
        if dept:
            filters['department'] = dept
    if args.designation and args.designation!='None':
        if desi:
            filters['designation'] = desi
    if args.company and args.company!='None':
        if comp:
            filters['company'] = comp
    if args.site_location and args.site_location!='None':
        if location:
            filters['custom_site_location'] = location
    # if args.division and args.division!='None':
    #     if division:
    #         filters['custom_division'] = division
    import re

    employees = frappe.db.get_all(
        'Employee',
        filters=filters,
        fields=['name', 'employee_name', 'designation']
    )

    def sort_key(emp):
        code = emp['name']
        match = re.match(r"([A-Za-z]+)(\d*)", code)

        prefix = match.group(1)
        number = match.group(2)

        number = int(number) if number.isdigit() else 0

        return (prefix, number)

    employees = sorted(employees, key=sort_key)
    
    
    return employees

import frappe

@frappe.whitelist()
def cancel_attendance_records(docname):
    att_list = frappe.db.get_all("Attendance", {
        'custom_upload_attendance': docname,
        'docstatus': 1
    }, ['name'])

    for a in att_list:
        try:
            att_doc = frappe.get_doc('Attendance', a.name)
            att_doc.cancel()
            frappe.db.commit()
        except Exception as e:
            frappe.log_error(f"{a.name}: {e}", "Upload Employee Attendance Cancel")

    reg_list = frappe.db.get_all("Attendance and OT Register", {
        'attendance_upload': docname
    }, ['name'])

    for r in reg_list:
        try:
            reg_doc = frappe.get_doc('Attendance and OT Register', r.name)
            reg_doc.breakfast = 0
            reg_doc.lunch = 0
            reg_doc.dinner = 0
            reg_doc.mobile_allow = 0
            reg_doc.attendance_upload = docname
            reg_doc.save(ignore_permissions=True)
            frappe.db.commit()
        except Exception as e:
            frappe.log_error(f"{r.name}: {e}", "Upload Employee Attendance Cancel")

@frappe.whitelist()
def check_holiday(date, holiday_list):
    query = """
        SELECT `tabHoliday`.holiday_date, `tabHoliday`.weekly_off
        FROM `tabHoliday List`
        LEFT JOIN `tabHoliday` ON `tabHoliday`.parent = `tabHoliday List`.name
        WHERE `tabHoliday List`.name = %s AND holiday_date = %s
    """
    holiday = frappe.db.sql(query, (holiday_list, date), as_dict=True)
    if holiday:
        if holiday[0].weekly_off == 1:
            return "WO"
        else:
            return "HO"
    return None

@frappe.whitelist()
def preview_html_new(name):
    upload_doc = frappe.get_doc("Upload Employee Attendance", name)

    fields_order = [
        "ot_hours", "pdo_ot", "breakfast", "lunch", "dinner", "food_allowance",
        "extra_days", "extra_rate_per_day", "nh_days", "nh_rate_per_day",
        "mobile_allow", "other_earnings", "other_earnings_remark",
        "other_deduction", "other_deduction_remark", "medical_leave_days",
        "salary_adjustment", "arrear_salary", "remarks_for_adjustment"
    ]

    registers = frappe.get_all(
        "Attendance and OT Register",
        filters={"attendance_upload": name},
        fields=["name", "employee", "employee_name", "docstatus"] + fields_order,
        order_by="employee asc"
    )

    if not registers:
        return "<p style='color:red;'>No A&O Register rows found!</p>"

    html = """
    <style>
    .preview-container {
        max-height: 450px;
        max-width: 100%;
        overflow: auto;
        border: 1px solid #dcdcdc;
        position: relative;
    }
    .preview-table {
        border-collapse: collapse;
        font-size: 12px;
        width: max-content;
    }
    .preview-table th, .preview-table td {
        border: 1px solid #dcdcdc;
        padding: 5px 8px;
        white-space: nowrap;
        background: #fff;
    }
    .preview-table thead th {
        position: sticky;
        top: 0;
        z-index: 50;
        background: #f4f4f4;
    }
    .preview-table th:nth-child(1), td:nth-child(1) { width: 50px; }
    .preview-table th:nth-child(2), td:nth-child(2) { width: 120px; }
    .preview-table th:nth-child(3), td:nth-child(3) { width: 200px; }

    .preview-table th:nth-child(1), td:nth-child(1) {
        position: sticky;
        left: 0;
        background: #f4f4f4;
        z-index: 52;
    }
    .preview-table th:nth-child(2), td:nth-child(2) {
        position: sticky;
        left: 50px;
        background: #fafafa;
        z-index: 51;
        font-weight: bold;
    }
    .preview-table th:nth-child(3), td:nth-child(3) {
        position: sticky;
        left: 170px;
        background: #fafafa;
        z-index: 51;
    }

    .edit-field {
        width: 100%;
        border: none;
        outline: none;
        background: transparent;
        padding: 4px;
        text-align: center;
    }
    .edit-field:focus {
        background: #ffffe6;
    }
    .readonly {
        background: #f3f3f3;
        pointer-events: none;
        font-weight: bold;
    }
    </style>

    <div class="preview-container">
    <table class="preview-table">
        <thead>
            <tr>
                <th>Select</th>
                <th>Employee</th>
                <th>Employee Name</th>
                <th>Document ID</th>
                <th>Status</th>
    """

    # Headers
    for f in fields_order:
        header = f.replace('_', ' ').title()
        header = header.replace('Ot', 'OT').replace('Pdo', 'PDO')
        html += f"<th>{header}</th>"

    html += "</tr></thead><tbody>"

    for reg in registers:
        status = "Draft" if reg.docstatus == 0 else "Submitted"
        checkbox_attr = "disabled" if reg.docstatus == 1 else ""

        html += f'<tr data-regname="{reg.name}">'
        html += f'<td><input class="emp_select" type="checkbox" value="{reg.employee}" {checkbox_attr}></td>'
        html += f'<td>{reg.employee}</td>'
        html += f'<td>{reg.employee_name or ""}</td>'
        html += f'<td><a href="/app/attendance-and-ot-register/{reg.name}" target="_blank">{reg.name}</a></td>'
        html += f'<td>{status}</td>'

        for f in fields_order:
            val = reg.get(f) or ""
            classes = "edit-field"
            readonly = ""

            if f in ("breakfast", "lunch", "dinner"):
                classes += " food-input"
                # html += (
                #     f'<td><input type="number" step="1" class="{classes}" '
                #     f'data-field="{'food_allowance'}" value="{frappe.utils.escape_html(str(val))}" {readonly}></td>'
                # )
                    
            if f == "food_allowance":
                classes += " readonly"
                readonly = "readonly"

            html += (
                f'<td><input type="number" step="1" class="{classes}" '
                f'data-field="{f}" value="{frappe.utils.escape_html(str(val))}" {readonly}></td>'
            )

        html += "</tr>"

    html += """
    </tbody></table></div>
    
    """

    return html

@frappe.whitelist()
def preview_html(name):
    upload_doc = frappe.get_doc("Upload Employee Attendance", name)

    fields_order = [
        "ot_hours", "pdo_ot", "breakfast", "lunch", "dinner", "food_allowance",
        "extra_days", "extra_rate_per_day", "nh_days", "nh_rate_per_day",
        "mobile_allow", "other_earnings", "other_earnings_remark",
        "other_deduction", "other_deduction_remark", "medical_leave_days",
        "salary_adjustment", "arrear_salary", "remarks_for_adjustment"
    ]

    registers = frappe.get_all(
        "Attendance and OT Register",
        filters={"attendance_upload": name},
        fields=["name", "employee", "employee_name", "docstatus"] + fields_order,
        order_by="employee asc"
    )
    
    if not registers:
        return "<p style='color:red;'>No A&O Register rows found!</p>"
    def sort_key(emp):
        code = emp['employee']
        match = re.match(r"([A-Za-z]+)(\d*)", code)

        prefix = match.group(1)
        number = match.group(2)

        number = int(number) if number.isdigit() else 0

        return (prefix, number)

    registers = sorted(registers, key=sort_key)
    html = """
    <style>
    .preview-container {
        max-height: 600px;
        max-width: 100%;
        overflow: auto;
        border: 1px solid #dcdcdc;
        position: relative;
    }
    .preview-table {
        border-collapse: collapse;
        font-size: 12px;
        width: max-content;
    }
    .preview-table th,
    .preview-table td {
        border: 1px solid #dcdcdc;
        padding: 5px 8px;
        
    }

    .preview-table thead th {
        position: sticky;
        top: 0;
        background: #e8262e;
        color: white;
        z-index: 100;
        text-transform: uppercase;
    }

        .preview-table th:nth-child(1),
        .preview-table td:nth-child(1) { width: 50px; }

        .preview-table th:nth-child(2),
        .preview-table td:nth-child(2) { width: 120px; }

        .preview-table th:nth-child(3),
        .preview-table td:nth-child(3) { width: 200px; }
        .preview-table thead th:nth-child(1){
            position: sticky;
            left: 0;
            z-index: 199;
        }
        .preview-table thead th:nth-child(2){
            position: sticky;
            left: 50px;
            z-index: 198;
        }
        .preview-table thead th:nth-child(3){
            position: sticky;
            left: 170px;
            z-index: 197;
        }
        .preview-table td:nth-child(1) {
            position: sticky;
            left: 0;
            background: #f4f4f4;
            z-index: 105;
        }
        .preview-table td:nth-child(2) {
            position: sticky;
            left: 50px;
            background: #fafafa;
            z-index: 105;
            font-weight: bold;
        }
        .preview-table td:nth-child(3) {
            position: sticky;
            left: 170px;
            background: #fafafa;
            z-index: 105;
        }
   
   
    .edit-field {
        width: 100%;
        border: none;
        outline: none;
        background: transparent;
        padding: 4px;
        text-align: center;
    }
    .edit-field:focus {
        background: #ffffe6;
    }
    
    .preview-table.table-hover tbody tr:hover td {
        background-color: #f5f5f5 !important;
    }

    .preview-table.table-hover tbody tr:hover td:nth-child(1),
    .preview-table.table-hover tbody tr:hover td:nth-child(2),
    .preview-table.table-hover tbody tr:hover td:nth-child(3) {
        background-color: #f5f5f5 !important;
    }

    </style>
    <div class="preview-container">
    <table class="preview-table table-hover mb-0" id="members-table">
        <thead>
            <tr>
                <th class="text-center">
                    <input type="checkbox" id="select_all">
                </th>

                <th>Employee</th>
                <th>Employee Name</th>
                <th>Document ID</th>
                <th>Status</th>
    """

    # Add dynamic column headers with OT/PDO capitalization fix
    for f in fields_order:
        header = f.replace('_', ' ').title()
        header = header.replace('Ot', 'OT').replace('Pdo', 'PDO')
        if not header.endswith("Remark"):
            html += f"""<th style="width:100px;">{header.upper()}</th>"""
        else:
            html += f"""<th>{header.upper()}</th>"""

    html += "</tr></thead><tbody>"

    for reg in registers:
        status = "Draft" if reg.docstatus == 0 else "Submitted"
        is_submitted = reg.docstatus == 1
        # Disable checkbox for submitted rows
        checkbox_attr = "disabled" if reg.docstatus == 1 else ""

        html += f'<tr data-regname="{reg.name}">'
        # html += f'<td><input class="emp_select" type="checkbox" {{ disabled }}></td>'

        html += f'<td style="text-align:center;"><input class="emp_select" type="checkbox" value="{reg.employee}" {checkbox_attr}></td>'
        html += f'<td>{reg.employee}</td>'
        html += f'<td>{reg.employee_name or ""}</td>'
        html += f'<td><a href="/app/attendance-and-ot-register/{reg.name}" target="_blank">{reg.name}</a></td>'
        html += f'<td>{status}</td>'

        for f in fields_order:
            val = reg.get(f) or ""
            if f == "food_allowance":
                html += (
                    f'<td><input type="text" class="edit-field" '
                    f'data-field="{f}" value="{val}" readonly></td>'
                )
            elif is_submitted:
                html += (
                    f'<td>'
                    f'<input type="text" class="edit-field" '
                    f'data-field="{f}" value="{val}" disabled>'
                    f'</td>'
                )

            # Draft → editable
            else:
                html += (
                    f'<td>'
                    f'<input type="text" class="edit-field" '
                    f'data-field="{f}" value="{val}">'
                    f'</td>'
                )

        html += "</tr>"

    html += "</tbody></table></div>"
    return html

@frappe.whitelist()
def preview_html_oetc(name):
    upload_doc = frappe.get_doc("Upload Employee Attendance", name)

    fields_order = [
        "ot_hours",
        "ot_amount",
        "night_ot",
        "night_ot_amount",
        "holiday_ot",
        "holiday_ot_amount",
        "medical_leave_days",
        "salary_adjustment",
        "arrear_salary",
        "remarks_for_adjustment",
        "other_earnings",
        "other_earnings_remark",
        "other_deduction",
        "other_deduction_remark"
    ]


    registers = frappe.get_all(
        "Attendance and OT Register",
        filters={"attendance_upload": name},
        fields=["name", "employee", "employee_name", "docstatus"] + fields_order,
        order_by="employee asc"
    )

    if not registers:
        return "<p style='color:red;'>No A&O Register rows found!</p>"
    def sort_key(emp):
        code = emp['employee']
        match = re.match(r"([A-Za-z]+)(\d*)", code)

        prefix = match.group(1)
        number = match.group(2)

        number = int(number) if number.isdigit() else 0

        return (prefix, number)

    registers = sorted(registers, key=sort_key)
    html = """
    <style>
    .preview-container {
        max-height: 450px;
        max-width: 100%;
        overflow: auto;
        border: 1px solid #dcdcdc;
        position: relative;
    }
    .preview-table {
        border-collapse: collapse;
        font-size: 12px;
        width: max-content;
    }
        .preview-table td {
        border: 1px solid #dcdcdc;
        padding: 5px 8px;
        white-space: nowrap;
        background: #fff;
    }

    /* ================= Sticky Header (vertical) ================= */
    .preview-table thead th {
        position: sticky;
        top: 0;
        background: #e8262e;
        color: white;
        z-index: 100;
        text-transform: uppercase;
    }

        /* ================= Column Widths ================= */
        .preview-table th:nth-child(1),
        .preview-table td:nth-child(1) { width: 50px; }

        .preview-table th:nth-child(2),
        .preview-table td:nth-child(2) { width: 120px; }

        .preview-table th:nth-child(3),
        .preview-table td:nth-child(3) { width: 200px; }
        .preview-table thead th:nth-child(1){
            position: sticky;
            left: 0;
            z-index: 199;
        }
        .preview-table thead th:nth-child(2){
            position: sticky;
            left: 50px;
            z-index: 198;
        }
        .preview-table thead th:nth-child(3){
            position: sticky;
            left: 170px;
            z-index: 197;
        }
        .preview-table td:nth-child(1) {
            position: sticky;
            left: 0;
            background: #f4f4f4;
            z-index: 105;
        }
        .preview-table td:nth-child(2) {
            position: sticky;
            left: 50px;
            background: #fafafa;
            z-index: 105;
            font-weight: bold;
        }
        .preview-table td:nth-child(3) {
            position: sticky;
            left: 170px;
            background: #fafafa;
            z-index: 105;
        }
    .edit-field {
        width: 100%;
        border: none;
        outline: none;
        background: transparent;
        padding: 4px;
        text-align: center;
    }
    .edit-field:focus {
        background: #ffffe6;
    }.preview-table.table-hover tbody tr:hover td {
        background-color: #f5f5f5 !important;
    }

    .preview-table.table-hover tbody tr:hover td:nth-child(1),
    .preview-table.table-hover tbody tr:hover td:nth-child(2),
    .preview-table.table-hover tbody tr:hover td:nth-child(3) {
        background-color: #f5f5f5 !important;
    }
    </style>
    <div class="preview-container">
    <table class="preview-table">
        <thead>
            <tr>
                <th class="text-center">
                    <input type="checkbox" id="select_all">
                </th>
                <th>Employee</th>
                <th>Employee Name</th>
                <th>Document ID</th>
                <th>Status</th>
    """

    # Add dynamic column headers with OT/PDO capitalization fix
    for f in fields_order:
        header = f.replace('_', ' ').title()
        header = header.replace('Ot', 'OT').replace('Pdo', 'PDO')
        # html += f"<th >{header}</th>"
        if not header.endswith("Remark"):
            html += f"""<th style="width:100px;">{header.upper()}</th>"""
        else:
            html += f"""<th>{header.upper()}</th>"""

    html += "</tr></thead><tbody>"

    for reg in registers:
        status = "Draft" if reg.docstatus == 0 else "Submitted"

        # Disable checkbox for submitted rows
        checkbox_attr = "disabled" if reg.docstatus == 1 else ""
        is_submitted = reg.docstatus == 1
        html += f'<tr data-regname="{reg.name}">'
        # html += f'<td><input class="emp_select" type="checkbox" value="{reg.employee}" {checkbox_attr}></td>'
        html += f'<td style="text-align:center;"><input class="emp_select" type="checkbox" value="{reg.employee}" {checkbox_attr}></td>'

        html += f'<td>{reg.employee}</td>'
        html += f'<td>{reg.employee_name or ""}</td>'
        html += f'<td><a href="/app/attendance-and-ot-register/{reg.name}" target="_blank">{reg.name}</a></td>'
        html += f'<td>{status}</td>'

        for f in fields_order:
            val = reg.get(f) or ""
            if is_submitted:
                html += (
                    f'<td><input type="text" class="edit-field" '
                    f'data-field="{f}" value="{frappe.utils.escape_html(str(val))}"></td>'
                )
            else:
                html += (
                    f'<td><input type="text" class="edit-field" '
                    f'data-field="{f}" value="{frappe.utils.escape_html(str(val))}"></td>'
                )

        html += "</tr>"

    html += "</tbody></table></div>"
    return html


# @frappe.whitelist()
# def preview_html_oetc(name):
#     upload_doc = frappe.get_doc("Upload Employee Attendance", name)

#     # Only required fields
#     fields_order = [
#         "ot_hours",
#         "night_ot",
#         "holiday_ot",
#         "medical_leave_days",
#         "salary_adjustment",
#         "arrear_salary",
#         "remarks_for_adjustment",
#         "other_earnings",
#         "other_earnings_remark",
#         "other_deduction",
#         "other_deduction_remark"
#     ]

#     registers = frappe.get_all(
#         "Attendance and OT Register",
#         filters={"attendance_upload": name},
#         fields=["name", "employee", "employee_name", "docstatus"] + fields_order,
#         order_by="employee asc"
#     )

#     if not registers:
#         return "<p style='color:red;'>No A&O Register rows found!</p>"

#     html = """
#     <style>
#     .preview-container {
#         max-height: 450px;
#         max-width: 100%;
#         overflow: auto;
#         border: 1px solid #dcdcdc;
#         position: relative;
#     }
#     .preview-table {
#         border-collapse: separate;
#         width: max-content;
#         font-size: 12px;
#     }
#     .preview-table th,
#     .preview-table td {
#         border: 1px solid #dcdcdc;
#         padding: 5px 8px;
#         white-space: nowrap;
#         background: #fff;
#     }

#     /* Freeze Header */
#     .preview-table thead th {
#         position: sticky;
#         top: 0;
#         z-index: 50;
#         background: #f4f4f4;
#     }

#     /* Freeze Employee (2nd col) */
#     .preview-table th:nth-child(2),
#     .preview-table td:nth-child(2) {
#         position: sticky;
#         left: 40px; /* width of Select column */
#         background: #fafafa;
#         z-index: 51;
#         font-weight: bold;
#     }

#     /* Freeze Employee Name (3rd col) */
#     .preview-table th:nth-child(3),
#     .preview-table td:nth-child(3) {
#         position: sticky;
#         left: 160px; /* adjust when columns resized */
#         background: #fafafa;
#         z-index: 51;
#     }

#     /* Freeze Select (1st col) */
#     .preview-table th:nth-child(1),
#     .preview-table td:nth-child(1) {
#         position: sticky;
#         left: 0;
#         background: #f4f4f4;
#         z-index: 52;
#     }

#     .edit-field {
#         width: 100%;
#         border: none;
#         outline: none;
#         background: transparent;
#         padding: 4px;
#         text-align: center;
#     }
#     .edit-field:focus {
#         background: #ffffe6;
#     }
#     </style>

#     <div class="preview-container">
#     <table class="preview-table">
#         <thead>
#             <tr>
#                 <th>Select</th>
#                 <th>Employee</th>
#                 <th>Employee Name</th>
#                 <th>Document ID</th>
#                 <th>Status</th>
#     """

#     for f in fields_order:
#         html += f"<th>{f.replace('_',' ').title()}</th>"

#     html += "</tr></thead><tbody>"

#     for reg in registers:
#         status = "Draft" if reg.docstatus == 0 else "Submitted"
#         html += f'<tr data-regname="{reg.name}">'
#         html += f'<td><input class="emp_select" type="checkbox" value="{reg.employee}"></td>'
#         html += f'<td>{reg.employee}</td>'
#         html += f'<td>{reg.employee_name or ""}</td>'
#         html += f'<td><a href="/app/attendance-and-ot-register/{reg.name}" target="_blank">{reg.name}</a></td>'
#         html += f'<td>{status}</td>'

#         for f in fields_order:
#             val = reg.get(f) or ""
#             html += (
#                 f'<td><input type="text" class="edit-field" '
#                 f'data-field="{f}" '
#                 f'value="{frappe.utils.escape_html(str(val))}"></td>'
#             )

#         html += "</tr>"

#     html += "</tbody></table></div>"
#     return html

import json

@frappe.whitelist()
def submit_employees(name, employees):
    if isinstance(employees, str):
        employees = json.loads(employees)

    for emp in employees:
        ao_doc = frappe.db.get_value(
            "Attendance and OT Register",
            {"attendance_upload": name, "employee": emp},
            "name"
        )
        if ao_doc:
            doc = frappe.get_doc("Attendance and OT Register", ao_doc)
            if doc.docstatus == 0:
                doc.submit()
    return "OK"

@frappe.whitelist()
def update_register_data(updates):
    if isinstance(updates, str):
        updates = json.loads(updates)

    for up in updates:
        reg = frappe.get_doc("Attendance and OT Register", up["register"])
        for field, value in up["fields"].items():
            
            if value == "":
                value = None
            reg.set(field, value)
        breakfast = reg.breakfast or 0
        lunch = reg.lunch or 0
        dinner = reg.dinner or 0

        food_allowance = float(breakfast) + (float(lunch) * 2) + (float(dinner) * 2)
        reg.set("food_allowance", food_allowance)
        reg.save(ignore_permissions=True)

    frappe.db.commit()
    return "OK"

@frappe.whitelist()
def submit_selected_registers(registers):
    registers = frappe.parse_json(registers)

    if not registers:
        frappe.throw("No registers selected")

    for reg in registers:
        doc = frappe.get_doc("Attendance and OT Register", reg)
        if doc.docstatus == 0:
            doc.submit()

    return "OK"

