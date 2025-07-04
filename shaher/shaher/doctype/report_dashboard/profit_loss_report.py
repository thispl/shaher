import frappe
from frappe.utils import nowdate
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from six import BytesIO
from erpnext.setup.utils import get_exchange_rate
from frappe.utils import flt
# from forex_python.converter import CurrencyRates
from frappe.utils import flt, getdate, fmt_money
import requests
from openpyxl.styles import Border, Side, Font

exchange_rates = {}

@frappe.whitelist()
def download():
	filename = 'Exchange Currency - Gain / Loss'
	build_xlsx_response(filename)

def make_xlsx(data=None, sheet_name="Sheet1", wb=None, column_widths=None):
	if wb is None:
		wb = Workbook()

	ws = wb.active
	ws.title = sheet_name

	thin_border = Border(
		left=Side(style='thin'),
		right=Side(style='thin'),
		top=Side(style='thin'),
		bottom=Side(style='thin')
	)
	green_font = Font(color="00B050") 
	red_font = Font(color="FF0000")


	ws["A3"] = 'Exchange Currency - Gain / Loss'
	ws["A3"].font = Font(bold=True, size=14)
	ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
	ws.merge_cells("A3:J3")
 
	ws.append([""])
 
	headers = ["Party", "Party Name", "Invoice no", "Billing Currency Name", 
			   "Billing Amount", "Payment Currency Name", "Payment Amount", "Payment Type", "Gain / Loss", "Status"]

	ws.append(headers)
	header_row = ws.max_row

	for col, header in enumerate(headers, start=1):
		cell = ws.cell(row=header_row, column=col, value=header)
		cell.font = Font(bold=True)
		cell.alignment = Alignment(horizontal="center")
		cell.border = thin_border

	column_widths = [10, 50, 20, 25, 15, 23, 18, 13, 15]
	for i, width in enumerate(column_widths, start=1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
  
	sales_invoice = frappe.get_all("Sales Invoice", {"status": "Paid"}, ["name", "currency", "grand_total", "customer"])
	for si in sales_invoice:
		pe = frappe.db.sql("""
			select c.reference_doctype, c.reference_name, p.paid_to_account_currency, sum(c.allocated_amount) as allocated_amount, p.name, p.payment_type
			from `tabPayment Entry` p
			inner join `tabPayment Entry Reference` c on c.parent = p.name
			where c.reference_name = %s and p.paid_to_account_currency is not NULL and p.docstatus = 1
			group by c.reference_name
		""", (si.name,), as_dict=True)

		if pe:
			# frappe.errprint(pe)
			pe = pe[0]
			gain_loss = get_gain_loss(pe.paid_to_account_currency, pe.allocated_amount, si.currency, si.grand_total, pe.payment_type)
			status = get_status(gain_loss)
			row_data = ["Customer", si.customer, si.name, si.currency, fmt_money(round(si.grand_total, 2)), 
						pe.paid_to_account_currency, fmt_money(round(pe.allocated_amount, 2)), pe.payment_type, fmt_money(round(gain_loss, 2)), status]
		

			ws.append(row_data)
  
		last_row = ws.max_row
		for col in range(1, len(row_data) + 1): 
			cell = ws.cell(row=last_row, column=col)
			cell.border = thin_border
			if col in [5, 7, 9] and last_row != 5:
				cell.alignment = Alignment(horizontal="right")
				if col == 9:  
					if gain_loss > 0:
						cell.font = green_font
					elif gain_loss < 0:
						cell.font = red_font
			if col == 10:  
				if status == "Gain":
					cell.font = green_font
				elif status == "Loss":
					cell.font = red_font
   
	purchase_invoice = frappe.get_all("Purchase Invoice", {"status": "Paid"}, ["name", "currency", "grand_total", "supplier"])
	for pi in purchase_invoice:
		pe = frappe.db.sql("""
			select c.reference_doctype, c.reference_name, p.paid_to_account_currency, sum(c.allocated_amount) as allocated_amount, p.name, p.payment_type
			from `tabPayment Entry` p
			inner join `tabPayment Entry Reference` c on c.parent = p.name
			where c.reference_name = %s and p.paid_to_account_currency is not NULL and p.docstatus = 1
			group by c.reference_name
		""", (pi.name,), as_dict=True)

		if pe:
			pe = pe[0]
			gain_loss = get_gain_loss(pe.paid_to_account_currency, pe.allocated_amount, pi.currency, pi.grand_total, pe.payment_type)
			status = get_status(gain_loss)
			row_data = ["Supplier", pi.supplier, pi.name, pi.currency, fmt_money(round(pi.grand_total, 2)), 
						pe.paid_to_account_currency, fmt_money(round(pe.allocated_amount, 2)), pe.payment_type, fmt_money(round(gain_loss, 2)), status]
		

			ws.append(row_data)
		last_row = ws.max_row
		for col in range(1, len(row_data) + 1): 
			cell = ws.cell(row=last_row, column=col)
			cell.border = thin_border
			if col in [5, 7, 9] and last_row != 5:
				cell.alignment = Alignment(horizontal="right")
				if col == 9:  
					if gain_loss > 0:
						cell.font = green_font
					elif gain_loss < 0:
						cell.font = red_font
			if col == 10:  
				if status == "Gain":
					cell.font = green_font
				elif status == "Loss":
					cell.font = red_font

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx()
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def fetch_exchange_rates():
	"""Fetch exchange rates once and store them globally."""
	global exchange_rates
	if not exchange_rates:
		url = "https://api.freecurrencyapi.com/v1/latest?apikey=fca_live_87Iqr6cHBRxY0xoi0NOqcMyR6Xlhlsaq2VZOFgKs"
		response = requests.get(url)
		exchange_rates = response.json().get("data", {})

@frappe.whitelist(allow_guest=True)
def get_gain_loss(payment_currency, amount, billed_currency, billed_amount, payment_type):
	exchange_rate = get_exchange_rate(payment_currency, billed_currency)
	converted_amount = round((amount * exchange_rate), 2)
	return (billed_amount - converted_amount) if payment_type == "Pay" else (converted_amount - billed_amount)


@frappe.whitelist()
def get_status(gain_loss):
	if gain_loss > 0:
		result = "Gain"
		
	if gain_loss < 0:
		result = "Loss"
	if gain_loss == 0:
		result = "-"
	return result	

@frappe.whitelist()
def test_check():
	import requests
	payment_currency = "USD"
	billed_currency = "INR"
	amount = 500
	payment_type = "Pay"
	billed_amount = 500
 
	url = f"https://api.freecurrencyapi.com/v1/latest?apikey=fca_live_87Iqr6cHBRxY0xoi0NOqcMyR6Xlhlsaq2VZOFgKs"

	response = requests.get(url)
	data = response.json()
	print(data)
	source_rate = data["data"].get(payment_currency)
	target_rate = data["data"].get(billed_currency)
	converted_amount = round((amount / source_rate) * target_rate, 2)
	print(converted_amount)
	if payment_type == "Pay":
		gain_loss = billed_amount - converted_amount
	else:
		gain_loss = converted_amount - billed_amount
	return gain_loss

@frappe.whitelist()
def check_er():
	payment_type = "Pay"
	billed_amount = 53200.000
	amount = 53200.000
	exchange_rate = get_exchange_rate("INR", "KWD")
	print(exchange_rate)
	converted_amount = round((amount * exchange_rate), 2)
	print(converted_amount)
	return (billed_amount - converted_amount) if payment_type == "Pay" else (converted_amount - billed_amount)