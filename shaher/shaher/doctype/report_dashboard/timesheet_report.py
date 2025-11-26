# Copyright (c) 2025, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from datetime import date, datetime, timedelta
import openpyxl
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from six import BytesIO
from frappe.utils import cint,today,flt,date_diff,add_days,add_months,date_diff,getdate,formatdate,cint,cstr
from frappe.utils.file_manager import get_file
import requests
from frappe.utils import getdate
from datetime import datetime, timedelta
from frappe import _
from frappe.utils.file_manager import get_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill

@frappe.whitelist()
def download():
    filename = 'Monthly Attendance Sheet.xlsx'
    build_xlsx_response(filename)

@frappe.whitelist()
def get_att_dates(from_date, to_date):
        no_of_days = date_diff(add_days(to_date, 1), from_date)
        dates = [add_days(from_date, i) for i in range(no_of_days)]
        return dates


@frappe.whitelist()
def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'

@frappe.whitelist()
def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)

    dates = get_dates(args.from_date, args.to_date)
    comp=frappe.db.get_single_value("Report Dashboard",'company')
    header1=[comp]
    ws.append(header1)
    site_location=frappe.db.get_single_value("Report Dashboard",'site_location')
    if not site_location:
        header12=[]
    else:
        header12=[site_location]
    ws.append(header12)
    from_date_str = datetime.strptime(args.from_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    to_date_str = datetime.strptime(args.to_date, '%Y-%m-%d').strftime('%d/%m/%Y')
    date_str=f"TIME SHEET {from_date_str} to {to_date_str}"
    header2=[date_str]
    ws.append(header2)
    header3 = ['SI No.','Emp. Code','Employee Name','Designation']
    header3.extend([d[0] for d in dates])  
    header3.extend(['OT from PDO','Total Present Days','Total Extra Days','Total Rest Days','Total Leave Days','Total Absent Days','Total M Leave Days','Total EM Leave Days','Total OT Hrs','Food','','','Mobile Allowance','Remarks'])
    header4 = ['', '', '', '']  
    header4.extend([d[1] for d in dates]) 
    header4.extend(['','','','','','','','','','B','L','D'])
    ws.append(header3)
    ws.append(header4)

    emp_data = get_employees(args)
    row_num = 1
    start_row = 6 

    for emp in emp_data:
        emp_row = [
            row_num,
            emp.get("name"),
            emp.get("employee_name"),
            emp.get("designation")
        ]

        emp_row.extend(emp.get("daily_status"))
        emp_row.append('')
        emp_row.extend(emp["totals"][1:8])
        emp_row.extend([
            emp.get("breakfast", 0),
            emp.get("lunch", 0),
            emp.get("dinner", 0),
            emp.get("mobile_allow", 0)
        ])
        ws.append(emp_row)
        emp_ot_row = ['', '', '', ''] 
        emp_ot_row.extend(emp.get("daily_ot"))
        # PDO OT currently given as 0 and tot OT should come in 2nd row of employee name
        for i in range(9):
            if i == 0:
                emp_ot_row.append(emp["totals"][0])
            elif i == 8:
                emp_ot_row.append(emp["totals"][8])
            else:
                emp_ot_row.append('')
        emp_ot_row.extend(['', '', '', ''])
        ws.append(emp_ot_row)
        for col in range(1, 5):
            ws.merge_cells(start_row=start_row, start_column=col, end_row=start_row + 1, end_column=col)
        row_num += 1
        start_row += 2
    merge_length=date_diff(args.to_date,args.from_date)+5
    location_col = merge_length + 1
    col_letter = get_column_letter(location_col)
    ws.cell(row=1, column=location_col).value = f"Location: {args.site_location}"
    ws.cell(row=2, column=location_col).value = f"{from_date_str} to {to_date_str}"
    to_date_obj = datetime.strptime(args.to_date, '%Y-%m-%d')
    month_year = to_date_obj.strftime('%B %Y')
    ws.cell(row=3, column=location_col).value = f"Month: {month_year}"
    # if site not present merge the company row with the 2nd row
    if not site_location:
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=merge_length)
    else:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=merge_length)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=merge_length)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=merge_length)
    ws.merge_cells(start_row=4, start_column=merge_length+10, end_row=4, end_column=merge_length+12)
    ws.merge_cells(start_row=1, start_column=merge_length+1, end_row=1, end_column=merge_length+14)
    ws.merge_cells(start_row=2, start_column=merge_length+1, end_row=2, end_column=merge_length+14)
    ws.merge_cells(start_row=3, start_column=merge_length+1, end_row=3, end_column=merge_length+14)
    for i in range(1,5):
        ws.merge_cells(start_row=4,start_column=i,end_row=5,end_column=i)
    header_fill = PatternFill(start_color="E6B8B7", end_color="E6B8B7", fill_type="solid")
    bold_font = Font(bold=True)
    center_alignment = Alignment(horizontal="center", vertical="center")
    # header alignment (bold,center)
    for row in [4, 5]:
        for col in range(1, len(header3) + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = header_fill
            cell.font = bold_font
            cell.alignment = center_alignment
    # for friday fill green
    friday_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type="solid")  
    for col_idx, cell in enumerate(ws[5], start=1): 
        if cell.value == 'Fr':
            ws.cell(row=5, column=col_idx).fill = friday_fill
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 43
    ws.column_dimensions['D'].width = 18
    last_col_index = ws.max_column-1
    last_col_letter_index = get_column_letter(last_col_index +1)
    last_col_letter = get_column_letter(last_col_index)
    ws.column_dimensions[last_col_letter].width = 18
    ws.column_dimensions[last_col_letter_index].width = 30
    for i in range(1, 4): 
        col_letter = get_column_letter(last_col_index - i)
        ws.column_dimensions[col_letter].width = 6
    for row in range(4, ws.max_row + 1, 2):
        ws.merge_cells(
            start_row=row,
            start_column=last_col_index,
            end_row=row + 1,
            end_column=last_col_index
        )
    for lcol in range(1, 4): 
        col_idx = last_col_index - lcol
        for row in range(6, ws.max_row + 1, 2):
            ws.merge_cells(
                start_row=row,
                start_column=col_idx,
                end_row=row + 1,
                end_column=col_idx
            )
    last_col = ws.max_column - 1
    start_col = last_col - 4 - 8
    end_col = last_col - 4

    for col in range(start_col, end_col + 1):
        ws.merge_cells(
            start_row=4,
            start_column=col,
            end_row=5,
            end_column=col
        )

        cell = ws.cell(row=4, column=col)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.font = Font(bold=True)
    for row in range(4, ws.max_row, 2):
        end_row = row + 1
        if end_row > ws.max_row:
            end_row = ws.max_row  

        ws.merge_cells(
            start_row=row,
            start_column=ws.max_column,
            end_row=end_row,
            end_column=ws.max_column
        )
    border = Border(left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000'))

    for rows in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1,max_col=ws.max_column):
        for cell in rows:
            cell.border = border
    for rows in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1,max_col=ws.max_column):
        for cell in rows:
            cell.border = border
            if cell.column != 3:  
                cell.alignment = center_alignment
            else:
               cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for rows in ws.iter_rows(min_row=1, max_row=3, min_col=1,max_col=ws.max_column):
        for cell in rows:
            cell.alignment = center_alignment
            cell.font = bold_font
    for col in range(5, merge_length + 1):
        col_letter = get_column_letter(col)
        ws.column_dimensions[col_letter].width = 6
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    start_row = 6 
    end_row = ws.max_row
    start_col = 5  
    end_col = merge_length + 1  

    for row in range(start_row, end_row + 1, 2):  
        current_val = None
        merge_start_col = None

        for col in range(start_col, end_col + 1):
            cell_val = ws.cell(row=row, column=col).value

            if cell_val in ['New Joining', 'Left']:
                if current_val == cell_val:
                    continue 
                else:
                    if merge_start_col is not None and col - merge_start_col > 1:
                        ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row + 1, end_column=col - 1)
                        # ws.merge_cells(start_row=row + 1, start_column=merge_start_col, end_row=row + 1, end_column=col - 1)
                        for c in range(merge_start_col, col):
                            ws.cell(row=row, column=c).fill = yellow_fill
                            ws.cell(row + 1, column=c).fill = yellow_fill

                    current_val = cell_val
                    merge_start_col = col
            else:
                if merge_start_col is not None and col - merge_start_col > 1:
                    ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row + 1, end_column=col - 1)
                    # ws.merge_cells(start_row=row + 1, start_column=merge_start_col, end_row=row + 1, end_column=col - 1)
                    for c in range(merge_start_col, col):
                        ws.cell(row=row, column=c).fill = yellow_fill
                        ws.cell(row + 1, column=c).fill = yellow_fill
                current_val = None
                merge_start_col = None
        if merge_start_col is not None and end_col - merge_start_col >= 1:
            ws.merge_cells(start_row=row, start_column=merge_start_col, end_row=row+1, end_column=end_col)
            # ws.merge_cells(start_row=row + 1, start_column=merge_start_col, end_row=row + 1, end_column=end_col)
            for c in range(merge_start_col, end_col + 1):
                ws.cell(row=row, column=c).fill = yellow_fill
                ws.cell(row + 1, column=c).fill = yellow_fill

    ws.freeze_panes = 'E1'
    xlsx_file=BytesIO()
    wb.save(xlsx_file)
    return xlsx_file

def get_dates(from_date, to_date):
    start_date = datetime.strptime(from_date, "%Y-%m-%d")
    end_date = datetime.strptime(to_date, "%Y-%m-%d")
    delta = end_date - start_date
    days = []
    for i in range(delta.days + 1):
        date = start_date + timedelta(days=i)
        day_num = date.strftime("%d") 
        weekday = date.strftime("%a")[:2] 
        days.append((day_num, weekday))
    return days

# Active employees have within the to_date
def get_active_employees(args):
    conditions = ["status = 'Active'"]
    values = []

    if args.get('to_date'):
        conditions.append("date_of_joining <= %s")
        values.append(args['to_date'])

    if args.get('company'):
        conditions.append("company = %s")
        values.append(args['company'])

    condition_str = " AND ".join(conditions)
    employees = frappe.db.sql(f"""
        SELECT name, employee_name, designation, date_of_joining ,status,'relieving_date'
        FROM `tabEmployee`
        WHERE {condition_str}
        ORDER BY employee_name
    """, values, as_dict=True)
    frappe.log_error(
        title="Active Employees Fetch Log",
        message=f"Query Conditions: {condition_str}\nTotal Employees Found: {len(employees)}"
    )

    return employees

# method to get employees relieved within the choosen period
def get_left_employees(args):
    conditions = ["status != 'Active'"]
    values = []
    if args.get('from_date') and args.get('to_date'):
        conditions.append("relieving_date BETWEEN %s AND %s")
        values.extend([args['from_date'], args['to_date']])
    if args.get('company'):
        conditions.append("company = %s")
        values.append(args['company'])
    condition_str = " AND ".join(conditions)
    employees = frappe.db.sql(f"""
        SELECT name, employee_name, designation, date_of_joining ,status,'relieving_date'
        FROM `tabEmployee`
        WHERE {condition_str}
        ORDER BY employee_name
    """, values, as_dict=True)

    return employees

from datetime import datetime

def format_date(date_str):
    if not date_str:
        return None
    if isinstance(date_str, datetime):
        return date_str.date()
    try:
        return datetime.strptime(date_str, "%Y-%m-%d").date()
    except Exception:
        return None

def get_employees(args):
    active_employees = get_active_employees(args)
    left_employees = get_left_employees(args)
    employees_dict = {e['name']: e for e in active_employees}
    for le in left_employees:
        if le['name'] not in employees_dict:
            employees_dict[le['name']] = le
    employees = list(employees_dict.values())
    final_data = []
    dates = get_att_dates(args.from_date, args.to_date)

    from_date = format_date(args.from_date)
    to_date = format_date(args.to_date)

    for e in employees:
        doj = getdate(e.get('date_of_joining'))
        relieving_date_raw = e.get('relieving_date')
        try:
            relieving_date = getdate(relieving_date_raw) if relieving_date_raw else None
        except Exception:
            relieving_date = None
        daily_status_row = []
        daily_ot_row = []
        tot_pre=0
        tot_ab=0
        tot_rd=0
        tot_ot=0
        tot_extra=0
        tot_leave=0
        tot_ml=0
        tot_eml=0

        for date in dates:
            status = ""
            ot_hours = 0
            att_date = getdate(date)

            attendance = frappe.db.get_all(
                'Attendance',
                {'attendance_date': date, 'employee': e['name'], 'docstatus': ['!=', 2]},
                ['status', 'leave_type', 'custom_overtime_hours', 'custom_rest_day'],
                order_by='creation desc',
                limit=1
            )
            hh = check_holiday(date, e['name'])

            if attendance:
                att = attendance[0]  
                ot_hours = att.get('custom_overtime_hours') or 0
                ot_hours = float(ot_hours)
                tot_ot += ot_hours

                if att.get('status') == 'Present':
                    tot_pre += 1
                    if hh:
                        tot_extra += 1
                    status = 'P'
                elif att.get('status') == 'Absent':
                    if att.get('custom_rest_day'):
                        status = 'RD'
                        tot_rd += 1
                    else:
                        tot_ab += 1
                        status = 'A'
                elif att.get('status') == 'Half Day':
                    status = 'HD'
                    tot_pre += 0.5
                    if att.get('leave_type'):
                        if att.get('leave_type')=='Emergency Leave':
                            tot_eml+=0.5
                            tot_leave += 0.5
                        elif att.get('leave_type')=='Sick Leave':
                            tot_ml+=0.5
                            tot_leave += 0.5
                        else:
                            tot_leave += 0.5
                    else:                     
                        tot_ab += 0.5

                elif att.get('status') == 'On Leave':
                    if att.get('leave_type')=='Emergency Leave':
                        tot_eml+=1
                        tot_leave += 1
                    elif att.get('leave_type')=='Sick Leave':
                        tot_ml+=1
                        tot_leave += 1
                    else:
                        tot_leave += 1
                    status = frappe.db.get_value('Leave Type', att.get('leave_type'), 'custom_abbr') or 'L'
                else:
                    status = att.get('status')

            else:
                if doj and att_date and att_date < doj:
                    frappe.log_error(
                            title="NEW DOJ",
                            message=f"Query Conditions:{e['name']}"
                        )
                    status = 'New Joining'
                elif relieving_date and att_date and att_date > relieving_date:
                    status = 'Left'
                else:
                    status = check_holiday(date, e['name']) or 'A'

                if status == 'A':
                    tot_ab += 1

            daily_status_row.append(status)
            daily_ot_row.append(str(ot_hours) if ot_hours else '')
        registers = frappe.db.get_all(
            "Attendance and OT Register",
            filters={"from_date": ["<=", args.to_date], "to_date": [">=", args.from_date], "docstatus": ["!=", 2]},
            fields=["employee", "breakfast", "lunch", "dinner", "mobile_allow"],
        )
        register_map = { r["employee"]: r for r in registers }
        reg = register_map.get(e["name"], {})
        breakfast = reg.get("breakfast", 0)
        lunch = reg.get("lunch", 0)
        dinner = reg.get("dinner", 0)
        mobile_allow = reg.get("mobile_allow", 0)

        final_data.append({
            "name": e['name'],
            "employee_name": e.get('employee_name'),
            "designation": e.get('designation'),
            "daily_status": daily_status_row,
            "daily_ot": daily_ot_row,
            "totals": [
                0,                
                tot_pre,
                tot_extra,
                tot_rd,
                tot_leave,
                tot_ab,
                tot_ml,                
                tot_eml,               
                tot_ot
            ],
            'breakfast':breakfast,
            'lunch':lunch,
            'dinner':dinner,
            'mobile_allow':mobile_allow
        })

    return final_data



@frappe.whitelist()
def check_holiday(date, emp):
    holiday_list = frappe.db.get_value('Employee', emp, 'holiday_list')
    
    holiday = frappe.db.sql("""
        SELECT `tabHoliday`.holiday_date, `tabHoliday`.weekly_off
        FROM `tabHoliday List`
        LEFT JOIN `tabHoliday` ON `tabHoliday`.parent = `tabHoliday List`.name
        WHERE `tabHoliday List`.name = %s AND `tabHoliday`.holiday_date = %s
    """, (holiday_list, date), as_dict=True)
    
    if holiday:
        if not holiday[0].weekly_off:
            return 'H'

