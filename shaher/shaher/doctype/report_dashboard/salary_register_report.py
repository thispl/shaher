# Copyright (c) 2024, TEAMPRO and contributors
# For license information, please see license.txt

# import frappe
from frappe.model.document import Document
import frappe
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO
import io  # Add this import to resolve the NameError

from datetime import datetime
from datetime import datetime, timedelta
from frappe.utils import (getdate, cint, add_months, date_diff, add_days, format_date,
    nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from openpyxl.drawing.image import Image as xlImage
import requests




@frappe.whitelist()
def download():
    filename = 'Salary Register Reports.xlsx'
    build_xlsx_response(filename)

def make_xlsx(sheet_name=None):
    args = frappe.local.form_dict
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name if sheet_name else 'Sheet1'

    data = get_data(args)
    if not data:
        frappe.throw("No data available to generate the report.")

    earnings_count = frappe.db.count("Salary Component", {'type': "Earning",'custom_sequence_id':['!=',0]})
    deductions_count = frappe.db.count("Salary Component",{'type': "Deduction",'custom_sequence_id':['!=',0]})
    total_column = 16 + earnings_count + deductions_count

    for row in data:
        if not isinstance(row, (list, tuple)):
            row = [row]
        ws.append(row)
    for row in range(1,6):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_column)
    for col in range(1,total_column+1):
        ws.merge_cells(start_row=6, start_column=col, end_row=7, end_column=col)
    ws.merge_cells(start_row=ws.max_row, start_column=1, end_row=ws.max_row, end_column=5)
    apply_styles(ws,total_column)
    set_column_widths(ws,total_column)

    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    xlsx_file.seek(0)
    return xlsx_file


def apply_styles(ws, total_column):
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='top', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)

    header_font = Font(bold=True, size=14,color="FFFFFF")
    sub_header_font = Font(bold=True, size=10, color="FFFFFF")
    text_font_data = Font(bold=False, size=9)

    border = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )
    # Green

    header_fill = PatternFill(start_color="FF808000", end_color="FF808000", fill_type="solid")
    header_fill_1 = PatternFill(start_color="FF666699", end_color="FF666699", fill_type="solid")

    for row in ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=total_column):
        for cell in row:
            cell.font = header_font
            cell.alignment = align_center
            # cell.border = border
            cell.fill = header_fill
    for row in ws.iter_rows(min_row=6, max_row=7, min_col=1, max_col=total_column):
        for cell in row:
            cell.font = sub_header_font
            cell.alignment = align_center
            cell.border = border
            cell.fill = header_fill_1

    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            cell.font = text_font_data
            cell.alignment = align_center
            cell.border = border
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=2, max_col=5):
        for cell in row:
            cell.font = text_font_data
            cell.alignment = align_left
            cell.border = border
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=6, max_col=total_column-3):
        for cell in row:
            cell.font = text_font_data
            cell.alignment = align_right
            cell.border = border
            cell.number_format = '#,##0.000' 
    for row in ws.iter_rows(min_row=8, max_row=ws.max_row, min_col=total_column-2, max_col=total_column):
        for cell in row:
            cell.font = text_font_data
            cell.alignment = align_left
            cell.border = border
    for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=total_column):
        for cell in row:
            cell.font = sub_header_font
            cell.alignment = align_center
            cell.border = border
            cell.fill = header_fill_1
            if cell.column not in [1, 2,3,4,5]:
                cell.number_format = '#,##0.000'
                cell.alignment = align_right 

def set_column_widths(ws,total_column):
   # total_column = total number of columns in your report
    column_widths = [10, 15, 25, 30] + [10] * (total_column - 7) + [20, 20, 20, 20,10]

    for i, width in enumerate(column_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
    row_heights = [20, 20, 20,20,20,40] + [20] * (ws.max_row - 1) +[30]*1
    for i, height in enumerate(row_heights, start=1):
        ws.row_dimensions[i].height = height

def build_xlsx_response(filename):
    xlsx_file = make_xlsx(sheet_name=filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'
      
def get_data(args):
    args = frappe.local.form_dict
    start_date = args.get('from_date')
    end_date = args.get('to_date')
    
    company = args.get('company') or 'All Companies'
    formatted_date =''
    if args.get('from_date'):
        formatted_date = frappe.utils.formatdate(start_date, 'dd-mm-yyyy')
    earnings = frappe.get_all("Salary Component", {'type': "Earning",'custom_sequence_id':['!=',0]}, ['name'], order_by='custom_sequence_id ASC')
    deductions = frappe.get_all("Salary Component", {'type': "Deduction",'custom_sequence_id':['!=',0]}, ['name'], order_by='custom_sequence_id ASC')

    earnings_count = len(earnings)
    deductions_count = len(deductions)
    total_column = 17 + earnings_count + deductions_count

    data = [
        [company] + [""] *total_column,
        ["Salary As of " + formatted_date],
        [""] *total_column,
        [""] *total_column,
        [""] *total_column,
        ["S.No", "Employee", "Salary Slip", "Name", "Div", "Days", "Normal OT Hours","Normal Overtime Amount",
        "Night OT Hours","Night Overtime Amount","Holiday OT Hours","Holiday Overtime Amount", "Basic", 
         "Earned Basic"]
        + [e.name for e in earnings if e.name not in  ['Basic',"Normal Overtime Amount","Night Overtime Amount","Holiday Overtime Amount"]]
        + ["Gross Salary"]
        + [d.name for d in deductions]
        + ["Total Deductions", "Net Pay","Bank Account","Bank Swift Code","Bank Account No"],
        [""]
    ]

    row = []
    index = 0
    salary_slips = get_salary_slips(args)
    total = {
        "payment_days": 0.0,
        "gross_pay": 0.0,
        "total_deduction": 0.0,
        "net_pay": 0.0,
    }
    for e in earnings:
        total[e.name] = 0.0
    for d in deductions:
        total[d.name] = 0.0
    total_row =[]
    total_basic =0
    total_ot_hours =0
    total_night_ot_hours =0
    total_holiday_ot_hours =0

    
    for s in salary_slips:
        
        index += 1
        
        basic = frappe.db.get_value('Employee',{'name':s['employee']},['custom_basic']) or 0.0
        total_basic +=basic
        division = frappe.db.get_value('Employee',{'name':s['employee']},['custom_division']) or ''
        bank_ac_no = frappe.db.get_value('Employee',{'name':s['employee']},['bank_ac_no']) or ''
        bank = frappe.db.get_value('Employee',{'name':s['employee']},['bank_name']) or ''
        iban = frappe.db.get_value('Employee',{'name':s['employee']},['iban']) or ''
        earnings_details = {}
        deductions_details = {}
        earning_values =[]
        deduction_values =[]
        joining_date, relieving_date = frappe.get_cached_value(
            "Employee", s['employee'], ["date_of_joining", "relieving_date"]
        )
        ot_hours = s['custom_overtime_hours'] or 0.0
        holiday_ot_hours = s['custom_holiday_ot_hours'] or 0.0
        night_ot_hours = s['custom_night_ot_hours'] or 0.0
        # if relieving_date and (getdate(start_date) <= relieving_date < getdate(end_date)):
        #     ot_hours = get_ot_hours(s['employee'], start_date, relieving_date) or 0.0
        # if joining_date and (getdate(start_date) < joining_date <= getdate(end_date)):
        #     ot_hours = get_ot_hours(s['employee'], joining_date, end_date)  or 0.0
        total_ot_hours +=ot_hours     
        total_holiday_ot_hours +=holiday_ot_hours   
        total_night_ot_hours +=night_ot_hours   
        for e in earnings:
            earnings_details[e.name] = frappe.get_value("Salary Detail", {'parent': s['name'], 'salary_component': e.name}, 'amount') or 0.0
            total[e.name] += earnings_details[e.name]

        for d in deductions:
            deductions_details[d.name] = frappe.get_value("Salary Detail", {'parent': s['name'], 'salary_component': d.name}, 'amount') or 0.0
            total[d.name] += deductions_details[d.name]

        for e in earnings:
            if e.name not in  ['Basic',"Normal Overtime Amount","Night Overtime Amount","Holiday Overtime Amount"]:
                earning = earnings_details.get(e.name, 0.0)
                earning_values.append(earning)
        
        for d in deductions:
            deduction =deductions_details.get(d.name, 0.0)
            deduction_values.append(deduction)

        total['payment_days'] += s['payment_days']
        total['gross_pay'] += s['gross_pay']
        total['total_deduction'] += s['total_deduction']
        total['net_pay'] += s['net_pay']
        row = [
            index, s['employee'], s['name'], s['employee_name'], division, s['payment_days'],
            ot_hours,earnings_details.get('Normal Overtime Amount', 0.0),night_ot_hours,earnings_details.get('Night Overtime Amount', 0.0),holiday_ot_hours,earnings_details.get('Holiday Overtime Amount', 0.0), basic, earnings_details.get('Basic', 0.0)
        ] + earning_values + [s['gross_pay']
        ] + deduction_values + [s['total_deduction'], 
        s['net_pay'],
         bank, iban, bank_ac_no
        ]


        data.append(row)

        total_row = (
        ["TOTAL", "", "", ""] +
        ["", "", total_ot_hours,total.get("Normal Overtime Amount", 0.0),total_night_ot_hours,total.get("Night Overtime Amount", 0.0),
        total_holiday_ot_hours, total.get("Holiday Overtime Amount", 0.0),total_basic,total.get("Basic", 0.0)] +
        [total.get(e.name, 0.0) for e in earnings if e.name not in  ['Basic',"Normal Overtime Amount","Night Overtime Amount","Holiday Overtime Amount"]] +
        [total["gross_pay"]] +
        [total.get(d.name, 0.0) for d in deductions] +
        [total["total_deduction"], total["net_pay"], "", "", ""]
    )
    data.append(total_row)
    return data

def format_currency(value):
    if value is None:
        return "0"
    
    number_str = f"{float(value):,.3f}"
    
    if '.' not in number_str:
        return number_str
    
    integer_part, decimal_part = number_str.split('.')
    
    integer_part = integer_part.replace(",", "")
    
    formatted_integer = f"{int(integer_part):,}"
    
    return f"{formatted_integer}.{decimal_part}"






import re
import frappe

def get_salary_slips(args):
    args = frappe.local.form_dict
    conditions = ["docstatus != 2"]  # Ensures only valid documents are selected
    values = []

    # Date filtering
    if args.get('from_date') and args.get('to_date'):
        conditions.append("start_date >= %s AND end_date <= %s")
        values.extend([args.get('from_date'), args.get('to_date')])
    elif args.get('from_date'):
        conditions.append("start_date >= %s")
        values.append(args.get('from_date'))
    elif args.get('to_date'):
        conditions.append("end_date <= %s")
        values.append(args.get('to_date'))

    # Company filtering
    if args.get('company'):
        conditions.append("company = %s")
        values.append(args.get('company'))

    # Construct SQL condition string
    condition_str = " AND ".join(conditions)

    # Query to fetch salary slips
    salary_slips = frappe.db.sql(f"""
        SELECT * FROM `tabSalary Slip`
        WHERE {condition_str}
        ORDER BY employee
    """, values, as_dict=True)

    # Sorting function for employee names
    def sort_key(emp):
        code = emp['employee']
        match = re.match(r"([A-Za-z]+)(\d*)", code)

        if match:
            prefix = match.group(1)
            number = match.group(2)
            number = int(number) if number.isdigit() else 0
            return (prefix, number)
        else:
            return (code, 0)  # Default to just the name if regex doesn't match

    # Sort the salary slips based on the employee name
    employees = sorted(salary_slips, key=sort_key)

    return employees  # Returning the sorted salary slips

def get_ot_hours(employee, from_date, to_date):

    result = frappe.db.sql("""
        SELECT SUM(custom_overtime_hours) AS ot
        FROM `tabAttendance`
        WHERE employee = %s 
        AND attendance_date BETWEEN %s AND %s 
        AND docstatus != 2
    """, (employee, from_date, to_date), as_dict=True)

    return result[0].get('ot', 0.0) if result else 0.0

