# Copyright (c) 2025, contact@half-ware.com and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import frappe
from frappe.model.document import Document
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from six import BytesIO
from datetime import datetime
from datetime import datetime, timedelta
from frappe.utils import (getdate, cint, add_months, date_diff, add_days, format_date,nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from openpyxl.styles import GradientFill, PatternFill


@frappe.whitelist()
def download():
	filename = 'Std Rated Sales'
	build_xlsx_response(filename)

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(sheet_name=filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def make_xlsx(sheet_name=None):
	args = frappe.local.form_dict
	wb = Workbook()
	ws = wb.active
	ws.title = sheet_name if sheet_name else 'Sheet1'
	# frappe.errprint(args)
	data = get_data(args)
	if not data:  # Check if data is empty or None
		frappe.throw("No data available to generate the report.")
	
	for row in data:
		# if not isinstance(row, (list, tuple)):
		# 	row = [row]  # Convert to a list if it's a single item (likely a string)
		ws.append(row)

	# Merging cells based on the number of rows in the header
	apply_styles(ws)
	set_column_widths(ws)

	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

def apply_styles(ws):
	align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	align_right = Alignment(horizontal='right', vertical='top', wrap_text=True)
	align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
	header_font = Font(bold=True, size=14)
	text_font = Font(bold=False, size=10)
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin')
	)
	bor = Border(
		top=Side(border_style='thin'),
		bottom=Side(border_style='double')
	)
	for rows in ws.iter_rows(min_row=3, max_row=3, min_col=1, max_col=11):
		for cell in rows:
			cell.fill = PatternFill(fgColor="d9d9d9", fill_type = "solid")
			cell.font = text_font
			cell.font = Font(color='96614d',bold=True,size=10)
			cell.alignment = align_center
			cell.border = border
	for rows in ws.iter_rows(min_row=4, max_row=ws.max_row-3, min_col=1, max_col=11):
		for cell in rows:
			cell.font = text_font
			cell.alignment = align_left
			cell.border = border
	for rows in ws.iter_rows(min_row=4, max_row=ws.max_row-3, min_col=5, max_col=5):
		for cell in rows:
			cell.font = text_font
			cell.alignment = align_center
			cell.border = border
	for rows in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row-1, min_col=7, max_col=8):
		for cell in rows:
			cell.font = text_font
			cell.font = Font(bold=True)
			cell.alignment = align_left
			cell.border = bor
	for rows in ws.iter_rows(min_row=ws.max_row-1, max_row=ws.max_row-1, min_col=6, max_col=6):
		for cell in rows:
			cell.font = text_font
			cell.font = Font(bold=True)
			cell.alignment = align_left
	for rows in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=1):
		for cell in rows:
			cell.font = text_font
			cell.font = Font(bold=True)
			# cell.alignment = align_left
	for rows in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=1):
		for cell in rows:
			cell.font = text_font
			cell.font = Font(bold=True)
			# cell.alignment = align_left

def set_column_widths(ws):
	column_widths = [6,15] + [34] * 1 + [15] * 1 + [17, 25]+ [23, 18] + [20, 15] + [80] *1
	for i, width in enumerate(column_widths, start=1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width



def get_data(args):
	datas=frappe.db.sql(f"""SELECT sa.name,sa.posting_date,sa.currency,sa.grand_total,sa.customer,c.tax_id,c.customer_name FROM `tabSales Invoice` sa LEFT JOIN `tabCustomer` c ON sa.customer = c.customer_name Where sa.docstatus=1 AND c.vat_type='Std Vat' AND sa.posting_date BETWEEN '{args.get('from_date')}' AND '{args.get('to_date')}'""",as_dict=1)
	grant=0
	ta=0
	row =[]
	data = [
			["Supplies of goods / services taxed at 5% - Box 1(a)* - If Group Should be Split By Group Member"],
			[""],
			["Serial#","Taxpayer VATIN","Taxpayer Name / Member Company Name (If applicable)","Tax Invoice/Tax Credit Note# ","Tax Invoice/Tax credit note Date - DD/MM/YYYY format only","Reporting period (From DD/MM/YYYY to DD/MM/YYYY format only)"," Tax Invoice/Tax credit note Amount OMR (before VAT)"," VAT Amount OMR","Customer Name","Customer VATIN","Clear description of the supply"],
		]
	index=1
	date1=frappe.utils.getdate(args.get('from_date')).strftime('%d-%m-%Y')
	date2=frappe.utils.getdate(args.get('to_date')).strftime('%d-%m-%Y')
	for rept in datas:
		grant+=rept.grand_total
		posting=frappe.utils.getdate(rept.posting_date).strftime('%d-%m-%Y')
		frappe.errprint('2')
		taxpayervatin='OM1100035159'
		taxpayername ='Shaher United Trading & Cont. Co.'
		tax=((rept.grand_total)*(5/100))
		ta+=tax
		formatted_tax = f"{float(tax):.3f}"
		i=float(formatted_tax) - int(float(formatted_tax))
		s=str(i)
		form_s=s[1:]
		form_a = format_currency(formatted_tax)
		# con_a =f"{form_a}" + "." + f"{form_s}"
		formatted_grand_total = f"{float(rept.grand_total):.3f}"
		i2=float(formatted_grand_total) - int(float(formatted_grand_total))
		s2=str(i2)
		form_s1=s[1:]
		form_b=format_currency(formatted_grand_total)
		# con_b = f"{form_b}" + "." + f"{form_s1}"
		formatted_grant = f"{float(grant):.3f}"
		a = "OMR" + "  " * (14 - len(form_a)) + form_a
		b = "OMR" + "  " * (20 - len(form_b)) + form_b
		form_grant=f"{float(grant):.3f}"
		form_g=format_currency(form_grant)
		form_ta=f"{float(ta):.3f}"
		form_t=format_currency(form_ta)
		g_len = len(formatted_grant)
		row = [index,
				taxpayervatin,
				taxpayername,
				rept.name,
				posting,
				f"{date1} - {date2}",
				b,
				a,
				rept.customer,
				rept.tax_id
			]
		index+=1
		data.append(row)
	row=[""]
	data.append(row)
	row=["","","","","","Total",f"{'OMR':<5}{form_g:>28}",f"{'OMR':<5}{form_t:>18}"]
	data.append(row)
	row=["*Total value of standard rated supplies of goods and services in the Sultanate, including deemed supplies."]
	data.append(row)
	return data



def format_currency(value):
    if value is None:
        return "0.000"
    
    try:
        float_val = float(value)
    except ValueError:
        return "Invalid input" 
    integer_part, dot, decimal_part = f"{float_val:.3f}".partition('.')
    number_str = str(int(integer_part))
    
    if len(number_str) > 3:
        last_three = number_str[-3:]  
        other_digits = number_str[:-3] 
        
        formatted_other = []
        while len(other_digits) > 2:
            formatted_other.append(other_digits[-2:])
            other_digits = other_digits[:-2]
        if other_digits:
            formatted_other.append(other_digits)        
        formatted_other.reverse()
        formatted_number = ','.join(formatted_other) + ',' + last_three
    else:
        formatted_number = number_str

    return f"{formatted_number}.{decimal_part}"
# def format_currency(value):
#     if value is None:
#         return "0"
    
#     number_str = str(int(value))
    
#     if len(number_str) > 3:
#         last_three = number_str[-3:]  
#         other_digits = number_str[:-3] 
        
#         formatted_other = []
#         while len(other_digits) > 2:
#             formatted_other.append(other_digits[-2:])
#             other_digits = other_digits[:-2]
#         if other_digits:
#             formatted_other.append(other_digits)        
#         formatted_other.reverse()
#         formatted_number = ','.join(formatted_other) + ',' + last_three
#     else:
#         formatted_number = number_str
    
#     return formatted_number
