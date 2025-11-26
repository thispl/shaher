# Copyright (c) 2024, TEAMPRO and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO
import io  

from datetime import datetime
from datetime import datetime, timedelta
from frappe.utils import (getdate, cint, add_months, date_diff, add_days, formatdate)
from openpyxl.drawing.image import Image as xlImage
import requests
from frappe.utils import getdate, add_days, date_diff, formatdate
from openpyxl.drawing.image import Image as XLImage


@frappe.whitelist()
def download():
	filename = 'Monthly Attendance Sheet.xlsx'
	build_xlsx_response(filename)

def make_xlsx(sheet_name=None):
	args = frappe.local.form_dict
	wb = Workbook()
	ws = wb.active
	ws.title = sheet_name if sheet_name else 'Sheet1'
	data = get_data(args)
	if not data:
		frappe.throw("No data available to generate the report.")
	start_column_count = 5
	end_column_count =12
	days_count = int(date_diff(args.get('to_date'),args.get('from_date')))+1
	total_column = start_column_count + days_count + end_column_count
	for row in data:
		if not isinstance(row, (list, tuple)):
			row = [row]
		ws.append(row)
	column_merge =start_column_count + days_count 
	
	url = 'https://erp.shaherunited.com/files/shaher_logo.png'
	response = requests.get(url)
	img_bytes = BytesIO(response.content)

	logo = XLImage(img_bytes)
	logo.width = 100 
	logo.height = 40
	# ws.merge_cells('AU1:AV1')
	ws.add_image(logo, "AU2")
	days_count1 = days_count // 2
	days_count2 = days_count1 if days_count % 2 == 0 else days_count1 + 1
	ws.merge_cells(start_row=1, start_column=1, end_row=3, end_column=2)
	ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=days_count1+3)
	ws.merge_cells(start_row=1, start_column=4+days_count1, end_row=1, end_column=5+days_count)
	ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=column_merge)
	ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=column_merge)
	for row in range(1, 4):
		ws.merge_cells(start_row=row, start_column=column_merge+1, end_row=row, end_column=column_merge+2)
		ws.merge_cells(start_row=row, start_column=column_merge+3, end_row=row, end_column=column_merge+10)
	
	
	row = 4
	while row <= ws.max_row:
		for col in range(1, 5):
			ws.merge_cells(start_row=row, start_column=col, end_row=row + 1, end_column=col)
			ws.merge_cells(start_row=row, start_column=column_merge+11, end_row=row + 1, end_column=column_merge+12)
		row += 2  
	for col in range(7+days_count, ws.max_column-1):
		ws.merge_cells(start_row=4, start_column=col, end_row=5, end_column=col)
	ws.merge_cells(start_row=1, start_column=ws.max_column -1, end_row=3, end_column=ws.max_column)
	ws.merge_cells(start_row=1, start_column=column_merge + 11, end_row=3, end_column=column_merge+12) 
	ws.merge_cells(start_row=4, start_column=column_merge + 11, end_row=5, end_column=column_merge+12) 
	align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
	align_right = Alignment(horizontal='right', vertical='top', wrap_text=True)
	align_left = Alignment(horizontal='left', vertical='top', wrap_text=True)
	# color="FFFFFF"
	header_font = Font(bold=True, size=14)
	sub_header_font = Font(bold=True, size=10)
	text_font_data = Font(bold=False, size=9)
	white_color_font = Font(bold=True, size=10, color="FFFFFF")
	white_color_font_data = Font(bold=False, size=9, color="FFFFFF")
	red_color_font = Font(bold=True, size=14, color="ab100d")
	border = Border(
		left=Side(border_style='thin'),
		right=Side(border_style='thin'),
		top=Side(border_style='thin'),
		bottom=Side(border_style='thin')
	)
	
	header_fill = PatternFill(start_color="f8cfb4", end_color="f8cfb4", fill_type="solid")
	purple = PatternFill(start_color="caa7e7", end_color="caa7e7", fill_type="solid")
	light_green = PatternFill(start_color="9dd059", end_color="9dd059", fill_type="solid")
	ash = PatternFill(start_color="cacacc", end_color="cacacc", fill_type="solid")
	dark_green = PatternFill(start_color="399c37", end_color="399c37", fill_type="solid")
	yellow = PatternFill(start_color="f5ff1c", end_color="f5ff1c", fill_type="solid")
	gray = PatternFill(start_color="898997", end_color="898997", fill_type="solid")
	brown = PatternFill(start_color="c0af95", end_color="c0af95", fill_type="solid")
	sky_blue = PatternFill(start_color="c6fbef", end_color="c6fbef", fill_type="solid")
	red = PatternFill(start_color="ab100d", end_color="ab100d", fill_type="solid")
	dark_blue = PatternFill(start_color="01184b", end_color="01184b", fill_type="solid")
	# row 1, 1-2 col
	for row in ws.iter_rows(min_row=1, max_row=1, min_col=3, max_col=(3+days_count1)):
		for cell in row:
			cell.font = header_font
			cell.alignment = align_left
			if cell.column != (3+days_count1):
				cell.border = border
	# row 1, 3-center col
	for row in ws.iter_rows(min_row=1, max_row=1, min_col=(4+days_count1), max_col=(4+days_count)):
		for cell in row:
			cell.font = header_font
			cell.alignment = align_right
			# cell.border = border
	for row in ws.iter_rows(min_row=1, max_row=3, min_col=(5+days_count), max_col=(6+days_count)):
		for cell in row:
			cell.font = header_font
			cell.alignment = align_left
			cell.border = border
	for row in ws.iter_rows(min_row=1, max_row=3, min_col=(7+days_count), max_col=(total_column)):
		for cell in row:
			cell.font = header_font
			cell.alignment = align_center
			cell.border = border
			if cell.row != 1:
				if cell.column == (days_count+8):
					cell.font = red_color_font
	for row in ws.iter_rows(min_row=2, max_row=3, min_col=1, max_col=4+days_count):
		for cell in row:
			cell.font = header_font
			cell.alignment = align_center
			cell.border = border
	# for row in ws.iter_rows(min_row=4, max_row=4, min_col=1, max_col=total_column):
	# 	for cell in row:
	# 		cell.font = sub_header_font
	# 		cell.alignment = align_center
	# 		cell.border = border
	# 		cell.fill = header_fill
	# 		if cell.column == (days_count+7):
	# 			cell.fill = yellow
	# 		elif cell.column == (days_count+8):
	# 			cell.fill = gray
	# 		elif cell.column == (days_count+9):
	# 			cell.fill = brown
	# 			# cell.font =white_color_font
	# 		elif cell.column == (days_count+10):
	# 			cell.fill = purple
	# 		elif cell.column == (days_count+11):
	# 			cell.fill = sky_blue
	# 		elif cell.column == (days_count+12):
	# 			cell.fill = red
	# 		elif cell.column == (days_count+13):
	# 			cell.fill = dark_green
	# 		elif cell.column == (days_count+14):
	# 			cell.fill = dark_blue
	# 			cell.font =white_color_font
	# 		elif cell.column == (days_count+15):
	# 			cell.fill = yellow
	for row in ws.iter_rows(min_row=5, max_row=5, min_col=1, max_col=total_column):
		for cell in row:
			cell.font = sub_header_font
			cell.alignment = align_center
			cell.border = border
			cell.fill = header_fill
			if cell.value in ['SA','F']:
				cell.fill = light_green
			if cell.column == (days_count+7):
				cell.fill = yellow
			elif cell.column == (days_count+8):
				cell.fill = gray
			elif cell.column == (days_count+9):
				cell.fill = brown
				# cell.font =white_color_font
			elif cell.column == (days_count+10):
				cell.fill = purple
			elif cell.column == (days_count+11):
				cell.fill = sky_blue
			elif cell.column == (days_count+12):
				cell.fill = red
			elif cell.column == (days_count+13):
				cell.fill = dark_green
			elif cell.column == (days_count+14):
				cell.fill = dark_blue
				cell.font =white_color_font
			elif cell.column == (days_count+15):
				cell.fill = yellow
	for row in ws.iter_rows(min_row=6, max_row=ws.max_row, min_col=1, max_col=total_column):
		for cell in row:
			cell.font = text_font_data
			cell.alignment = align_center
			cell.border = border
			if cell.value == 'RD':
				cell.fill = purple
			elif cell.value == 'P':
				cell.fill = ash
			elif cell.value == 'A':
				cell.fill = red
			elif cell.value in ['ML']:
				cell.fill = dark_green
				cell.font =white_color_font_data
			elif cell.value in ['EL']:
				cell.fill = dark_blue
				cell.font =white_color_font_data
			elif cell.value != 0:
				for c in range(6,days_count):
					if cell.column == c:
						cell.fill = yellow
			elif cell.value == "L":
				cell.fill = sky_blue
	# for row in ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, min_col=1, max_col=total_column):
	# 	for cell in row:
	# 		cell.font = sub_header_font
	# 		cell.alignment = align_center
	# 		cell.border = border
	column_widths = [10, 15, 15, 15,15] + [5] * (days_count+1) + [9] * (9) + [10]*(ws.max_column)
	for i, width in enumerate(column_widths, start=1):
		ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width
	row_heights =[20,20,20,25,25]+[20] * (ws.max_row - 1) +[30]*1
	for i, height in enumerate(row_heights, start=1):
		ws.row_dimensions[i].height = height
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(sheet_name=filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'


def get_data(args):
	args = frappe.local.form_dict
	start_date = args.get('from_date')
	end_date = args.get('to_date')
	 
	date_label = f"{ordinal(getdate(start_date).day)} {getdate(start_date).strftime('%B')} to {ordinal(getdate(end_date).day)} {getdate(end_date).strftime('%B')} {getdate(end_date).year}"
	month_label = getdate(end_date).strftime("%b %Y").upper()

	company = args.get('company') or 'All Companies'

	formatted_date = ''
	if start_date:
		formatted_date = formatdate(start_date, 'dd-mm-yyyy')

	days_count = int(date_diff(args.get('to_date'),args.get('from_date')))+1
	days_count1 = days_count // 2
	days_count2 = days_count1 if days_count % 2 == 0 else days_count1 + 1
	custom_day_map = {0: "M", 1: "TU", 2: "W", 3: "TH", 4: "F", 5: "SA", 6: "S"}
	date_row = []
	day_row = []

	for i in range(days_count):
		current_date = add_days(getdate(start_date), i)
		date_row.append(current_date.day)
		day_row.append(custom_day_map[current_date.weekday()])

	data = [
		["", "", "Client:OMAN ELECTRICITY TRANSMISSION COMPANY S.A.O.C"]
		+ [""] * days_count1 + ["Contractor:" + company]
		+ [""] * (days_count2+1) + ["Location:", "", "AL DAKHLIYAH (NIZWA)"] + [""] * 7,

		["", "", "Maintenance of all OETC Grid Stations, Overhead Lines and Underground Cables (OETC-58/2023)"]
		+ [""] * (days_count+2) + ["Date:", "", date_label] + [""] * 7,

		["", "", "MONTHLY ATTENDANCE TIME SHEET"]
		+ [""] * (days_count+2) + ["Month:", "", month_label] + [""] * 7,

		["SI", "Emp. Code", "Employee Name", "Designation", ""]
		+ date_row + ["", "OT from OETC", "Total Present Days - P", "Total Extra Work Days",
			"Total Weekend Days - RD", "Total Leave Days", "Total Absent Days",
			"Total M Leave Days", "Total E Leave Days", "Total OT Hours", "Remarks"],

		["", "", "", "", ""] + day_row
	]

	index = 0
	active_employees = get_active_employees(args)
	left_employees = get_left_employees(args)
	employees = active_employees + left_employees
	
	for emp in employees:
		ot_from_oetc_per_employee =''
		present_count_per_employee =0
		extra_days_count_per_employee =0
		rd_count_per_employee =0
		leave_count_per_employee =0
		absent_count_per_employee =0
		m_leave_count_per_employee =0
		e_leave_count_per_employee =0
		ot_hours_per_employee =0
		total_dict ={
			'present_count' :0,
			'ot_from_oetc' :0,
			'extra_days_count' :0,
			'rd_count' :0,
			'leave_count' :0,
			'absent_count' :0,
			'm_leave_count' :0,
			'e_leave_count' :0,
			'ot_hours':0
		}
		
		index += 1
		joining_date, relieving_date = frappe.get_cached_value(
			"Employee", emp['name'], ["date_of_joining", "relieving_date"]
		)
		att_start = start_date
		att_end = end_date
		
		
		row1 =[index,emp['name'],emp['employee_name'],emp['designation'],'Attendance']
		row2 =['','','','','OT Hr']
		
		for i in range(days_count):
			att_date = add_days(getdate(att_start), i)
			
			# attendance = get_att_data(emp['name'], att_date)
			# frappe.log_error(attendance,'attendance')
			if frappe.db.exists('Attendance',{'attendance_date':att_date,'employee':emp['name'],'docstatus':['!=',2]}):
				attendance = frappe.db.get_all('Attendance',{'attendance_date':att_date,'employee':emp['name'],'docstatus':['!=',2]},['status','leave_type','custom_overtime_hours','custom_rest_day'])
				att =attendance[0]
				ot_hours = att['custom_overtime_hours']
				ot_hours_per_employee +=att['custom_overtime_hours']
				if att['status'] =='Present':
					status ='P'
					present_count_per_employee +=1
					if att['custom_rest_day']:
						extra_days_count_per_employee +=1
						# status ='P/RD'
				elif att['status'] =='Absent':
					if att['custom_rest_day']:
						rd_count_per_employee +=1
						status ='RD'
					else:
						if check_holiday(att_date, emp['name']):
							status ='H'
						else:
							status ='A'
							absent_count_per_employee +=1

				elif att['status'] =='Half Day':
					present_count_per_employee +=0.5
					absent_count_per_employee +=0.5
					status ='HD'
					if att['leave_type']:
						status = frappe.db.get_value('Leave Type',{'name':att['leave_type']},'custom_abbr')
						if status =='ML':
							m_leave_count_per_employee +=1
						elif status =='EL':
							e_leave_count_per_employee +=1
				elif att['status'] =='On Leave':
					leave_count_per_employee +=1
					if att['leave_type']:
						status = frappe.db.get_value('Leave Type',{'name':att['leave_type']},'custom_abbr')
						if status =='ML':
							m_leave_count_per_employee +=1
						elif status =='EL':
							e_leave_count_per_employee +=1
			

			else:
				ot_hours =0
				if check_holiday(att_date, emp['name']):
					status ='H'
				else:
					status ='A'
					absent_count_per_employee +=1
			if relieving_date and getdate(att_date) <= relieving_date < getdate(att_date):
				status =''
			if joining_date and getdate(att_date) < joining_date <= getdate(att_date):
				att_start = joining_date
				status = ''
			row1.append(status)
			row2.append(ot_hours)
		row1 +=['',ot_from_oetc_per_employee,present_count_per_employee,extra_days_count_per_employee,rd_count_per_employee,leave_count_per_employee,absent_count_per_employee,m_leave_count_per_employee,e_leave_count_per_employee,round(ot_hours_per_employee)]
		data.append(row1)
		data.append(row2)

	return data


def get_active_employees(args):
	conditions = ["status = 'Active'"]
	values = []

	if args.get('from_date') and args.get('to_date'):
		conditions.append("date_of_joining <= %s and date_of_joining <= %s")
		values.extend([args['from_date'], args['to_date']])

	if args.get('company'):
		conditions.append("company = %s")
		values.append(args['company'])

	condition_str = " AND ".join(conditions)
	employees = frappe.db.sql(f"""
		SELECT name,employee_name,designation FROM `tabEmployee`
		WHERE {condition_str}
		ORDER BY employee_name
	""", values, as_dict=True)

	return employees

def get_left_employees(args):
	conditions = ["status != 'Active'"]
	values = []

	if args.get('from_date') and args.get('to_date'):
		conditions.append("relieving_date >= %s and relieving_date <= %s")
		values.extend([args['from_date'], args['to_date']])

	if args.get('company'):
		conditions.append("company = %s")
		values.append(args['company'])

	condition_str = " AND ".join(conditions)
	employees = frappe.db.sql(f"""
		SELECT name,employee_name,designation FROM `tabEmployee`
		WHERE {condition_str}
		ORDER BY employee_name
	""", values, as_dict=True)

	return employees

def ordinal(n):
	return f"{n}{'th' if 11<=n%100<=13 else {1:'st',2:'nd',3:'rd'}.get(n%10, 'th')}"

@frappe.whitelist()
def check_holiday(date, emp):
    holiday_list = frappe.db.get_value('Employee', emp, 'holiday_list')
    
    holiday = frappe.db.sql("""
        SELECT `tabHoliday`.holiday_date, `tabHoliday`.weekly_off
        FROM `tabHoliday List`
        LEFT JOIN `tabHoliday` ON `tabHoliday`.parent = `tabHoliday List`.name
        WHERE `tabHoliday List`.name = %s AND `tabHoliday`.holiday_date = %s
    """, (holiday_list, date), as_dict=True)
    
    if holiday:
        if not holiday[0].weekly_off:
            return 'H'
