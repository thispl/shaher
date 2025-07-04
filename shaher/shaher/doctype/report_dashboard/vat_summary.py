import frappe
from frappe.model.document import Document
import frappe
from frappe.model.document import Document
from datetime import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from six import BytesIO
from datetime import datetime
from datetime import datetime, timedelta
from frappe.utils import (getdate, cint, add_months, date_diff, add_days, format_date,nowdate, get_datetime_str, cstr, get_datetime, now_datetime, format_datetime)
from openpyxl.styles import GradientFill, PatternFill
from openpyxl.styles import Border, Side

@frappe.whitelist()
def download():
	filename = 'Std Rated Sales'
	build_xlsx_response(filename)

def build_xlsx_response(filename):
	xlsx_file = make_xlsx(sheet_name=filename)
	frappe.response['filename'] = filename + '.xlsx'
	frappe.response['filecontent'] = xlsx_file.getvalue()
	frappe.response['type'] = 'binary'

def make_xlsx(sheet_name=None):
	args = frappe.local.form_dict
	wb = Workbook()
	ws = wb.active
	ws.title = sheet_name
	ws.append(['Shaher United Trading & Cont. Co.', '', ''])
	ws.append(['', '', ''])
	from_date = datetime.strptime(args.get('from_date'), "%Y-%m-%d").strftime("%d/%m/%Y")
	to_date = datetime.strptime(args.get('to_date'), "%Y-%m-%d").strftime("%d/%m/%Y")
	period = f" For the period from {from_date} to {to_date}"
	ws.append([period, '', ''])
	data = get_data(args)
	
	for row in data:
		ws.append(row)
	ws.column_dimensions['A'].width = 24 
	ws.column_dimensions['B'].width = 15
	ws.column_dimensions['C'].width = 15
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
	ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=3)
	thin_border = Border(
	left=Side(style='thin', color='000000'),
	right=Side(style='thin', color='000000'),
	top=Side(style='thin', color='000000'),
	bottom=Side(style='thin', color='000000')
	)
	left_align = Alignment(horizontal='left')
	right_align = Alignment(horizontal='right')
	for row in ws.iter_rows(min_row=4, max_row=15, min_col=1, max_col=3):
		for cell in row:
			cell.border = thin_border
			if cell.column == 1:
				cell.alignment = left_align
			elif cell.row >= 5 and cell.column in (2, 3):
				cell.alignment = right_align
	bold_font = Font(bold=True)

	for row_idx in [1,3, 15]:
		for cell in ws[row_idx]:
			cell.font = bold_font
	xlsx_file = BytesIO()
	wb.save(xlsx_file)
	xlsx_file.seek(0)
	return xlsx_file

def get_data(args):
	data = []
	vat = frappe.db.sql(f"""
		SELECT SUM(sa.total) as total
		FROM `tabSales Invoice` sa
		LEFT JOIN `tabCustomer` c ON sa.customer = c.customer_name
		WHERE sa.docstatus=1 AND c.vat_type='Std Vat' 
		AND sa.posting_date BETWEEN '{args.get('from_date')}' AND '{args.get('to_date')}'
	""", as_dict=1)

	no_vat = frappe.db.sql(f"""
		SELECT SUM(sa.total) as total
		FROM `tabSales Invoice` sa
		LEFT JOIN `tabCustomer` c ON sa.customer = c.customer_name
		WHERE sa.docstatus=1 AND c.vat_type='Zero Vat' 
		AND sa.posting_date BETWEEN '{args.get('from_date')}' AND '{args.get('to_date')}'
	""", as_dict=1)
	purchase_total = frappe.db.sql(f"""
		SELECT SUM(total) as total
		FROM `tabPurchase Invoice`
		WHERE docstatus = 1
		AND posting_date BETWEEN '{args.get('from_date')}' AND '{args.get('to_date')}'
	""", as_dict=1)
	sales_tot=vat[0]['total'] or 0
	pur_tot=purchase_total[0]['total']  or 0
	vat_pur=purchase_total[0]['total'] or 0
	vat_pur_5=vat_pur* 0.05
	five_percent_s = sales_tot * 0.05
	five_percent_p = pur_tot * 0.05
	vat_tot=five_percent_p-five_percent_s
	fp=f'{vat_pur_5}'
	vat_pay=f'{vat_tot}'
	data.append(['Description', 'Amount', 'VAT 5%'])
	data.append(['', '','' ])
	data.append(['VAT OUT - Sales', vat[0]['total'] or 0, five_percent_s])
	data.append(['VAT OUT - Sales-Zero VAT', no_vat[0]['total'] or 0, '-'])
	data.append(['VAT IN - Purchase', purchase_total[0]['total'] or 0,fp ])
	data.append(['VAT IN - Fixed Assets', '-','-' ])
	data.append(['', '','' ])
	data.append(['VAT Payable', '',vat_pay])
	data.append(['', '','' ])
	data.append(['Output Adjustments', '','-' ])
	data.append(['', '','' ])
	data.append(['Net Payable', '',vat_pay])
	return data