import frappe
import json
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from frappe import _
from frappe.utils import cint, flt
from frappe import _dict

from erpnext.accounts.report.financial_statements import (
    compute_growth_view_data,
    get_columns,
    get_data,
    get_filtered_list_for_consolidated_report,
    get_period_list,
)

from erpnext.accounts.report.balance_sheet import balance_sheet

@frappe.whitelist()
def download():
    args = frappe.local.form_dict
    filters = json.loads(args.get("filters"))
    filters = _dict(filters)
    export_balance_sheet(filters)

def export_balance_sheet(filters):
    columns, data, *_ = balance_sheet.execute(filters)
    columns = [col for col in columns if col.get("fieldname") != "currency"]
    short_code = frappe.db.get_value("Company", filters.company, "abbr")
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    bold_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Account - Parent").font = bold_font

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx + 1, value=col.get("label", ""))
        cell.font = bold_font

    row_idx = 2
    for row in data:
        if not any(row.values()):
            continue  # skip completely empty rows

        account_name = row.get("account_name")
        account_name_with_abbr = f"{account_name} - {short_code}" if account_name else None
        parent_account = frappe.db.get_value("Account", account_name_with_abbr, "parent_account")
        is_group = frappe.db.get_value("Account", account_name_with_abbr, "is_group")
        indent_level = row.get("indent", 0)

        parent_cell = ws.cell(row=row_idx, column=1, value=parent_account)
        if not parent_account:
            parent_cell.font = bold_font

        for col_idx, col in enumerate(columns, start=1):
            fieldname = col.get("fieldname")
            value = row.get(fieldname, "")
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)

            if fieldname == "account":
                cell.alignment = Alignment(indent=indent_level)
            if not parent_account:
                cell.font = bold_font
            if col.get('fieldtype') == 'Currency':
                cell.number_format = '#,##0.00'

        row_idx += 1

    for col_idx in range(1, len(columns) + 2):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
        if col_idx in [1, 2]:
            ws.column_dimensions[column_letter].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response['filename'] = 'Balance Sheet.xlsx'
    frappe.response['filecontent'] = output.read()
    frappe.response['type'] = 'binary'
    frappe.response['headers'] = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

@frappe.whitelist()
def remove_outer_quotes(value):
    if isinstance(value, str) and value.startswith("'") and value.endswith("'"):
        return value[1:-1]
    return value

@frappe.whitelist()
def test_check():
    filters = _dict({
        'company': 'SHAHER UNITED TRADING & CONTRACTING CO',
        'filter_based_on': 'Fiscal Year',
        'period_start_date': '2025-01-01',
        'period_end_date': '2025-12-31',
        'from_fiscal_year': '2025',
        'to_fiscal_year': '2025',
        'periodicity': 'Yearly',
        'cost_center': [],
        'project': [],
        'selected_view': 'Report',
        'accumulated_values': 1,
        'include_default_book_entries': 1
    })
    return export_balance_sheet(filters)

@frappe.whitelist()
def test_check1():
    value = frappe.db.get_value("Account", "1100 - Cash In Hand - SUTC", "account_type")
    print(value)


import frappe
import json
import io
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from frappe import _
from frappe.utils import cint, flt
from frappe import _dict

from erpnext.accounts.report.financial_statements import (
    compute_growth_view_data,
    get_columns,
    get_data,
    get_filtered_list_for_consolidated_report,
    get_period_list,
)

from erpnext.accounts.report.balance_sheet import balance_sheet

@frappe.whitelist()
def download():
    args = frappe.local.form_dict
    filters = json.loads(args.get("filters"))
    filters = _dict(filters)
    export_balance_sheet(filters)

def export_balance_sheet(filters):
    columns, data, *_ = balance_sheet.execute(filters)
    columns = [col for col in columns if col.get("fieldname") != "currency"]
    short_code = frappe.db.get_value("Company", filters.company, "abbr")
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"

    bold_font = Font(bold=True)

    ws.cell(row=1, column=1, value="Account - Parent").font = bold_font

    for col_idx, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx + 1, value=col.get("label", ""))
        cell.font = bold_font

    row_idx = 2
    for row in data:
        if not any(row.values()):
            continue  # skip completely empty rows

        account_name = row.get("account_name")
        account_name_with_abbr = f"{account_name} - {short_code}" if account_name else None
        parent_account = frappe.db.get_value("Account", account_name_with_abbr, "parent_account")
        is_group = frappe.db.get_value("Account", account_name_with_abbr, "is_group")
        indent_level = row.get("indent", 0)

        parent_cell = ws.cell(row=row_idx, column=1, value=parent_account)
        if not parent_account:
            parent_cell.font = bold_font

        for col_idx, col in enumerate(columns, start=1):
            fieldname = col.get("fieldname")
            value = row.get(fieldname, "")
            cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)

            if fieldname == "account":
                cell.alignment = Alignment(indent=indent_level)
            if not parent_account:
                cell.font = bold_font
            if col.get('fieldtype') == 'Currency':
                cell.number_format = '#,##0.00'

        row_idx += 1

    for col_idx in range(1, len(columns) + 2):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 15
        if col_idx in [1, 2]:
            ws.column_dimensions[column_letter].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.response['filename'] = 'Balance Sheet.xlsx'
    frappe.response['filecontent'] = output.read()
    frappe.response['type'] = 'binary'
    frappe.response['headers'] = {
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    }

@frappe.whitelist()
def remove_outer_quotes(value):
    if isinstance(value, str) and value.startswith("'") and value.endswith("'"):
        return value[1:-1]
    return value

@frappe.whitelist()
def test_check():
    filters = _dict({
        'company': 'SHAHER UNITED TRADING & CONTRACTING CO',
        'filter_based_on': 'Fiscal Year',
        'period_start_date': '2025-01-01',
        'period_end_date': '2025-12-31',
        'from_fiscal_year': '2025',
        'to_fiscal_year': '2025',
        'periodicity': 'Yearly',
        'cost_center': [],
        'project': [],
        'selected_view': 'Report',
        'accumulated_values': 1,
        'include_default_book_entries': 1
    })
    return export_balance_sheet(filters)

@frappe.whitelist()
def test_check1():
    value = frappe.db.get_value("Account", "1100 - Cash In Hand - SUTC", "account_type")
    print(value)


import frappe
import json
from frappe import _dict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

from erpnext.accounts.report.cash_flow import cash_flow


@frappe.whitelist()
def download_cashflow():
    args = frappe.local.form_dict
    filters = json.loads(args.get("filters"))
    filters = _dict(filters)
    export_cash_flow(filters)

def export_cash_flow(filters):
    columns, data, *_ = cash_flow.execute(filters)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cash Flow"

    # Styles
    bold_font = Font(bold=True)
    align_center = Alignment(horizontal="center")
    align_left = Alignment(horizontal="left")
    align_right = Alignment(horizontal="right")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    CURRENCY_FORMAT = "#,##0.00"

    # ---------------- HEADER ----------------
    ws.cell(row=1, column=1, value="S.No").font = bold_font
    ws.cell(row=1, column=1).alignment = align_center
    ws.cell(row=1, column=1).border = thin_border

    col_index = 1
    for col in columns:
        if col.get("fieldname") == "currency":
            continue

        col_index += 1
        cell = ws.cell(row=1, column=col_index, value=col.get("label"))
        cell.font = bold_font
        cell.alignment = align_center
        cell.border = thin_border

    # ---------------- DATA ----------------
    row_idx = 2
    sno = 1

    for row in data:
        # Skip meaningless empty rows
        if not row.get("account") and not any(
            isinstance(v, (int, float)) and v != 0 for v in row.values()
        ):
            continue

        # S.No
        ws.cell(row=row_idx, column=1, value=sno).alignment = align_right
        ws.cell(row=row_idx, column=1).border = thin_border
        sno += 1

        data_index = 1
        indent = row.get("indent", 0)
        is_group = row.get("is_group")

        for col in columns:
            if col.get("fieldname") == "currency":
                continue

            data_index += 1
            fieldname = col.get("fieldname")
            value = row.get(fieldname, "")

            cell = ws.cell(row=row_idx, column=data_index, value=value)
            cell.border = thin_border

            # -------- Account column --------
            if fieldname == "account":
                cell.alignment = Alignment(
                    horizontal="left", indent=indent
                )
                if is_group:
                    cell.font = bold_font

            # -------- Currency columns --------
            elif col.get("fieldtype") == "Currency" and isinstance(value, (int, float)):
                cell.number_format = CURRENCY_FORMAT
                cell.alignment = align_right
                if is_group:
                    cell.font = bold_font
            else:
                cell.alignment = align_left

        row_idx += 1

    # ---------------- COLUMN WIDTH ----------------
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    # ---------------- RESPONSE ----------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    frappe.local.response.filename = "Cash Flow.xlsx"
    frappe.local.response.filecontent = output.read()
    frappe.local.response.type = "download"
