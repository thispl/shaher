import frappe
import frappe
from frappe import _
from frappe.utils import date_diff, fmt_money, today, add_days, nowdate, unique, cstr
from frappe import _
import math
from frappe.utils.background_jobs import enqueue
from datetime import datetime, time, timedelta
from frappe.utils import cint, flt, formatdate, get_link_to_form, getdate, now,time_diff_in_hours
from frappe.utils import getdate, get_datetime
import calendar
from frappe.model.mapper import get_mapped_doc
from frappe.utils.csvutils import UnicodeWriter,read_csv_content
from frappe.utils.file_manager import get_file
from frappe.utils import nowdate, get_abbr
from datetime import datetime

from frappe.utils import nowdate, add_days, getdate, formatdate
import io
from frappe.utils.pdf import get_pdf
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
import PyPDF2
from PyPDF2 import PdfReader, PdfWriter
from reportlab.lib import colors
import os
import re
from frappe.model.workflow import apply_workflow
from hrms.hr.doctype.leave_application.leave_application import get_leave_details
from frappe.utils import flt, escape_html

import frappe
from frappe.utils import flt, getdate, nowdate
from frappe import _
from html import escape as escape_html

from frappe.utils import flt, getdate, nowdate
from frappe import _
from frappe.utils import strip_html
import json
import openpyxl
from openpyxl.styles import Font, Alignment
from frappe.utils import flt, get_site_path
@frappe.whitelist()
def html_table(employee, date):
    if not employee:
        frappe.throw(_("Employee is required"))

    # ----------- Leave Details ----------
    details = get_leave_details(employee=employee, date=date)
    leave_allocations = details.get("leave_allocation", {})
    rows = []

    for lt, alloc in leave_allocations.items():
        total_allocated = flt(alloc.get("total_leaves", 0))
        expired_leaves = flt(alloc.get("expired_leaves", 0))
        used_leaves = flt(alloc.get("leaves_taken", 0))
        pending = flt(alloc.get("leaves_pending_approval", 0))
        available = flt(alloc.get("remaining_leaves", 0))

        rows.append({
            "leave_type": lt,
            "total": total_allocated,
            "expired": expired_leaves,
            "used": used_leaves,
            "pending": pending,
            "available": available
        })

    if rows:
        leave_html = """
        <div class="leave-table" style="margin-bottom:20px;">
            <h4 style="margin-bottom:10px;">Leave Balance</h4>
            <table style="border:1px solid #000; width:100%; border-collapse: collapse;">
                <thead>
                    <tr style="background-color:#b81a0f; color:white; text-align:center;border:1px solid #000;">
                        <th style="border:1px solid #000;">Leave Type</th>
                        <th style="border:1px solid #000;">Total Allocated Leaves</th>
                        <th style="border:1px solid #000;">Expired Leaves</th>
                        <th style="border:1px solid #000;">Used Leaves</th>
                        <th style="border:1px solid #000;">Leaves Pending Approval</th>
                        <th style="border:1px solid #000;">Available Leaves</th>
                    </tr>
                </thead>
                <tbody>
        """
        for r in rows:
            leave_html += f"""
                    <tr style="text-align:center;border:1px solid #000;">
                        <td style="padding:5px;border:1px solid #000;">{escape_html(r['leave_type'])}</td>
                        <td style="padding:5px;border:1px solid #000;">{r['total']}</td>
                        <td style="padding:5px;border:1px solid #000;">{r['expired']}</td>
                        <td style="padding:5px;border:1px solid #000;">{r['used']}</td>
                        <td style="padding:5px;border:1px solid #000;">{r['pending']}</td>
                        <td style="padding:5px;border:1px solid #000;">{r['available']}</td>
                    </tr>
            """
        leave_html += """
                </tbody>
            </table>
        </div>
        """
    else:
        leave_html = "<p>No leave allocations found.</p>"

    # ----------- Gratuity Card ----------
    doj = frappe.db.get_value("Employee", employee, "date_of_joining")
    if not doj:
        frappe.throw(_("Date of Joining not set for employee"))

    date_join = getdate(doj)
    date_to = getdate(nowdate())
    total_days = (date_to - date_join).days
    custom_nationality =  frappe.db.get_value("Employee", employee, "custom_nationality")
    if custom_nationality and custom_nationality !='Oman':
        gratuity_html = f"""
        <style>
        .small-box {{
            flex: 0 0 150px;
            height: 120px;
            width: 150px;
            border-radius: 0.25rem;
            position: relative;
            overflow: hidden;
            color: #fff;
            box-shadow: 0 1px 3px rgba(0,0,0,.2);
            display: flex;
            flex-direction: column;
            align-items: center; 
            justify-content: center; 
            background-color: #28a745;
            text-align: center;
            padding: 10px;
            box-sizing: border-box;
        }}
        .small-box .icon {{
            position: absolute;
            top: 50%;
            left: 50%;
            transform: translate(-50%, -50%);
            font-size: 2.5rem;
            opacity: 0.15;
        }}
        .small-box .inner h3 {{
            font-size: 2rem;
            font-weight: bold;
            margin: 0 0 5px 0;
            color: white;
        }}
        .small-box .inner p {{
            font-size: 0.9rem;
            margin: 0;
        }}
        </style>

        <div class="small-box">
            <div class="inner">
                <h3>{total_days}</h3>
                <p>Gratuity Days</p>
            </div>
        </div>
        """
    else:
        gratuity_html =""
    # ----------- Full Layout ----------
    return f"""
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <div class="employee-summary" style="
        display: flex;
        gap: 20px;
        align-items: flex-start;
        flex-wrap: wrap;
    ">
        <div class="leave-table-container" style="flex: 1 1 80%;">
            {leave_html}
        </div>
        <div class="gratuity-card-container" style="flex: 1 1 10%;">
            {gratuity_html}
        </div>
    </div>
    """



@frappe.whitelist()
def gratuity_card(employee):
    if not employee:
        frappe.throw("Employee is required")
    doj = frappe.db.get_value("Employee", {'name':employee},'date_of_joining')
    if not doj:
        frappe.throw("Date of Joining not set for employee")

    date_2 = nowdate()
    date_join = getdate(doj)
    date_to = getdate(date_2)

    # diff = relativedelta(date_to, date_join)
    # yos = f"{diff.years} years, {diff.months} months and {diff.days} days"
    total_days = (date_to - date_join).days

    # Constructing the HTML content directly in Python
    html = f"""
    <div class="card" style="max-width: 100px; margin-bottom: 1rem;">
        <div class="card-body">
            <p class="card-text"><strong>Gratuity Days:</strong> {total_days}</p>
        </div>
    </div>
    """
    return html



#Show the below details while open the quotation document if there in profit percentage
@frappe.whitelist()
def margin_html(mar_dis_per,mar_tot_cost,mar_gross_tot,mar_dis_amt,total_selling,prof_per,margin_profit):
    data = ''
    data = "<table style='width:85%'>"
    data += "<tr><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>TOTAL COST</td><td style ='font-weight:bold;text-align:center;border:1px solid black'>%s</td><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>GROSS TOTAL SELLING</td><td style ='font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(mar_tot_cost),2),round(float(mar_gross_tot),2))
    data += "<tr><td colspan  = 2 style = 'border:1px solid black;border-right-color:#dedede;background-color:#dedede;'><td style ='border-left-color:#dedede;background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>DISCOUNT PERCENTAGE</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(mar_dis_per),2))
    data += "<tr><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>PROFIT PERCENTAGE</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>DISCOUNT AMOUNT</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(prof_per),2),round(float(mar_dis_amt),2))
    data += "<tr><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>PROFIT AMOUNT</td><td style ='background-color:#74cee7;font-weight:bold;text-align:center;border:1px solid black'>%s</td><td style ='background-color:#dedede;font-weight:bold;text-align:center;border:1px solid black'>TOTAL SELLING</td><td style ='font-weight:bold;text-align:center;border:1px solid black'>%s</td></tr>"%(round(float(margin_profit),2),round(float(total_selling),2))
    data += "</table>"
    return data

#creates Item Code based on the Item Description
@frappe.whitelist()
def create_item_code(description):
    if not frappe.db.exists("Item",{'description':description,'disabled':0}):
        item=frappe.new_doc("Item")
        item.item_name=description
        item.description=description
        item.item_group="Services"
        item.item_code = ''.join([word[0].upper() for word in description.split() if word.isalpha()])
        item.save(ignore_permissions=True)
        return item.name
    else:
        item_code=frappe.get_doc("Item",{'description':description,'disabled':0})
        return item_code

#create user and user_permission for supplier
@frappe.whitelist()
def create_user_permission(supplier_name,name,user_id):
    if frappe.db.exists("User",{'name':user_id,'enabled':1}):
        user=frappe.get_doc("User",{'name':user_id,'enabled':1})
        user.append("roles",{
            "role":"Supplier"
        })
        user.module_profile='Supplier'
        user.save(ignore_permissions=True)
        if not frappe.db.exists("User Permission",{'user':user_id,'allow':'Supplier','for_value':name}):
            usr_per=frappe.new_doc("User Permission")
            usr_per.user=user_id
            usr_per.allow='Supplier'
            usr_per.for_value=name
            usr_per.save(ignore_permissions=True)
        import string
        import random
        all_characters = string.ascii_letters + string.digits
        length = 6
        password = ''.join(random.choices(all_characters, k=length))
        if password:
            from frappe.utils.password import update_password

            update_password(user=user.name, pwd=password)
    else:
        user=frappe.new_doc("User")
        user.email=user_id
        user.send_welcome_email=0
        user.first_name=supplier_name
        user.append("roles",{
            "role":"Supplier"
        })
        user.module_profile='Supplier'
        user.send_welcome_email=0
        user.save(ignore_permissions=True)
        if not frappe.db.exists("User Permission",{'user':user_id,'allow':'Supplier','for_value':name}):
            usr_per=frappe.new_doc("User Permission")
            usr_per.user=user_id
            usr_per.allow='Supplier'
            usr_per.for_value=name
            usr_per.save(ignore_permissions=True)
        import string
        import random
        all_characters = string.ascii_letters + string.digits
        length = 6
        password = ''.join(random.choices(all_characters, k=length))
        if password:
            from frappe.utils.password import update_password

            update_password(user=user.name, pwd=password)

#Get Supplier Contact details to trigger mail from RFQ
@frappe.whitelist()
def get_supplier_contact(supplier):
    email_id=frappe.db.get_value("Supplier",{'name':supplier},['custom_supplier_login_id'])
    # email_id=frappe.db.get_value("Supplier",{'name':supplier},['email_id'])
    return email_id


@frappe.whitelist()
def mail_for_supplier_quotation(doc, method):
        if doc.items:
            request_quotation = doc.items[0].request_for_quotation

            owner = frappe.db.get_value("Request for Quotation", request_quotation, "owner")

            if owner !='Administrator':
                subject = f"Supplier Quotation - {doc.name}"
                message = f"""
                Dear Sir/Mam,<br><br>
                Supplier Quotation - {doc.name} aganist Request for Quotation - {request_quotation} has been submmitted successfully.
                <br><br>
                <i>This email has been automatically generated. Please do not reply</i>
                """
                frappe.sendmail(
                    recipients = [owner],
                    subject = subject,
                    message = message
                )

@frappe.whitelist()
def update_supplier_status(doc,method):
    supplier_qtn=frappe.db.get_value("Purchase Order Item",{"parent":doc.name},['supplier_quotation'])
    if supplier_qtn:
        frappe.db.set_value("Supplier Quotation",supplier_qtn,'custom_po_status','Converted to PO')

@frappe.whitelist()
def update_supplier_status_on_cancel(doc,method):
    supplier_qtn=frappe.db.get_value("Purchase Order Item",{"parent":doc.name},['supplier_quotation'])
    if supplier_qtn:
        frappe.db.set_value("Supplier Quotation",supplier_qtn,'custom_po_status','PO Cancelled')

@frappe.whitelist()
def update_po_status(doc,method):
    frappe.db.set_value("Supplier Quotation",doc.name,'custom_po_status','Pending for PO')

@frappe.whitelist()
def validate_remarks(doc, method):
    if doc.custom_cancellation_remarks and doc.custom_cancellation_remarks.strip():
        return doc.custom_cancellation_remarks.isalnum()
    else:
        frappe.throw("Field <b>Cancellation Remarks</b> is mandatory")

@frappe.whitelist()
def update_po_remarks(name, remark):
    frappe.db.set_value("Purchase Order", name, "custom_cancellation_remarks", remark)
    frappe.db.set_value("Purchase Order", name, "workflow_state", "Cancelled")
    doc = frappe.get_doc("Purchase Order", name)
    doc.cancel()


@frappe.whitelist()
def update_signature(user, docname, field):
    if frappe.db.exists("Employee", {"user_id": user}):
        signature = frappe.db.get_value("Employee", {"user_id": user}, ["custom_digital_signature"])
        if signature:
            frappe.db.set_value("Purchase Order", docname, field, signature)






@frappe.whitelist()
def update_signature_purchase(user, docname, field):
    if frappe.db.exists("Employee", {"user_id": user}):
        signature = frappe.db.get_value("Employee", {"user_id": user}, ["custom_digital_signature"])
        if signature:
            frappe.db.set_value("Purchase Order", docname, field, signature)


    frappe.db.set_value("Purchase Order", docname, "custom_approved_by", user)


@frappe.whitelist()
def update_signature_project(user, docname, field):
    if frappe.db.exists("Employee", {"user_id": user}):
        signature = frappe.db.get_value("Employee", {"user_id": user}, ["custom_digital_signature"])
        if signature:
            frappe.db.set_value("Purchase Order", docname, field, signature)


    frappe.db.set_value("Purchase Order", docname, "custom_authorized_by", user)





@frappe.whitelist()
def trigger_mail_for_purchase_user(doc, method):
    pu_name = frappe.db.get_value("User", doc.owner, "full_name")
    subject = f"PO {doc.name} has been Approved - Reg"

    message = f"""
                    <div style="font-family: Arial, sans-serif;
                        border: solid 1px silver;
                        border-radius: 10px;
                        padding-top: 50px;
                        padding-bottom: 50px;
                        padding-right: 10px;
                        padding-left: 10px;
                        padding-left: 20px;">

                        <b>Dear {pu_name}
                        </b>
                        <p>Greetings !</p>
                        <p>PO <b>{doc.name}</b> has been approved. Kindly check and share it to the Supplier</p>
                        <a style="background-color: #3488ef;
                        color: white;
                        text-decoration: none;
                        padding: 5px;
                        border-radius: 5px;" href="https://shaher.teamproit.com/app/purchase-order/{doc.name}">Open Purchase Order</a>

                    """
    if doc.custom_project_managers_signature:
        message += f"""
                <p>with regards,</p>
                <img src="{doc.custom_project_managers_signature}" />
                </div>
                """
    else:
        message += "</div>"
    frappe.sendmail(
                    recipients = [doc.owner],
                    subject = subject,
                    message = message,
                )

@frappe.whitelist()
def trigger_mail_for_supplier(name):
    doc = frappe.get_doc("Purchase Order", name)
    # if frappe.db.exists("Supplier Quotation", {"supplier": doc.supplier}):
    # 	supplier_quotation = frappe.db.get_value("Supplier Quotation", {"supplier": doc.supplier}, "quotation_number")
    # else:
    # 	supplier_quotation = ""
    supplier = frappe.db.get_value("Supplier", doc.supplier, "custom_supplier_login_id")
    frappe.log_error
    message = f"""
            <div style="font-family: Arial, sans-serif;
                        border: solid 1px silver;
                        border-radius: 10px;
                        padding-top: 50px;
                        padding-bottom: 50px;
                        padding-right: 10px;
                        padding-left: 10px;
                        padding-left: 20px;">

                        <b>Dear {doc.supplier}</b>
                        <p>Greetings from <b>{doc.company}</b> !</p>
                        <p>Please find the attached Purchase Order</p>
            """
    if supplier:
        subject = "Purchase Order - Reg"
        frappe.sendmail(
                        recipients = [supplier],
                        subject = subject,
                        message = message,
                        attachments = [frappe.attach_print("Purchase Order", name,file_name=name, print_format="Purchase Order")]
                    )
        return "ok"
    else:
        frappe.throw("Couldn't find supplier email")


@frappe.whitelist()
def get_po_qty(doc, method):
    if doc.custom_purchase_order:
        po = frappe.get_doc("Purchase Order", doc.custom_purchase_order, "items")
        if po:
            for po_item in po.items:
                for pi_item in doc.items:
                    if po_item.item_code == pi_item.item_code:
                        if doc.is_new():
                            pi_item.custom_ordered_quantity = po_item.qty
                        pi_item.custom_balance_quantity = po_item.qty - pi_item.qty


#Get the below values and set it to onload of the Vehicle document
@frappe.whitelist()
def record(name):
    maintenance = frappe.db.exists("Vehicle Maintenance Check List",{'register_no':name})
    if maintenance:
        main = frappe.db.get_all("Vehicle Maintenance Check List",{'register_no':name},['complaint','employee_id','vehicle_handover_date','garage_name'])[0]

    accident = frappe.db.exists("Vehicle Accident Report",{'plate_no':name})
    if accident:
        acc = frappe.db.get_all("Vehicle Accident Report",{'plate_no':name},['name','emp_id','date_of_accident','remarks'])[0]

        return main,acc

@frappe.whitelist()
def child_set_value(doc, method):
    wo=''
    if doc.custom_sales_order:
        if frappe.db.exists("Sales Order Item", {"parent": doc.custom_sales_order}):
            wo = frappe.db.get_value("Sales Order Item", {"parent": doc.custom_sales_order}, "custom_work_order_number")
    for item in doc.items:
        item.required_quantity = item.qty
        current_qty = frappe.db.get_value("Bin", {"item_code": item.item_code, "warehouse": doc.set_warehouse}, "actual_qty") or 0
        item.current_quantity = current_qty
        item.balance_qty_to_order = item.required_quantity - item.current_quantity
        additional_qty = float(item.additional_qty) if item.additional_qty is not None else 0
        item.finaly_qty_to_ordered = max(item.balance_qty_to_order - additional_qty, 0)
        # item.qty = item.finaly_qty_to_ordered
        if wo:
            item.custom_work_order_number = wo or ''
    # for item in doc.items:
    #     item.required_quantity = item.qty
    #     current_qty = frappe.db.get_value("Bin", {"item_code": item.item_code, "warehouse": doc.set_warehouse}, "actual_qty") or 0
    #     item.current_quantity = current_qty
    #     item.balance_qty_to_order = item.required_quantity - item.current_quantity
    #     item.finaly_qty_to_ordered = max(item.balance_qty_to_order - item.additional_qty, 0)
    #     item.qty = item.finaly_qty_to_ordered

@frappe.whitelist()
def update_item_table(sup_name,name):
    sq_sows=frappe.get_all("SQ SOW",{'supplier_quotation':sup_name})
    supplier_quotation=frappe.get_doc('Supplier Quotation',sup_name)
    supplier_quotation.set('custom_item_table',[])
    for i in sq_sows:
        sq_sow=frappe.get_doc("SQ SOW",i.name)

        for sow in sq_sow.manpower_table:
            supplier_quotation.append("custom_item_table",{
                'item_code':sow.item_code,
                'item_name':sow.item_name,
                'description':sow.description,
                'qty':sow.qty,
                'rate':sow.rate,
                'amount':sow.amount,
                'sow':sq_sow.sow_name
            })
        for sow in sq_sow.materials_table:
            supplier_quotation.append("custom_item_table",{
                'item_code':sow.item_code,
                'item_name':sow.item_name,
                'description':sow.description,
                'qty':sow.qty,
                'rate':sow.rate,
                'amount':sow.amount,
                'sow':sq_sow.sow_name
            })
        for sq in supplier_quotation.items:
            if sq.item_code==sq_sow.sow_id:
                sq.amount=sq_sow.total_sq_sow_amount
                sq.rate=(sq_sow.total_sq_sow_amount)/float(sq_sow.qty)
    supplier_quotation.save(ignore_permissions=True)

@frappe.whitelist
def update_item_table_without_amount(sup_name,name):
    sq_sows=frappe.get_all("SQ SOW",{'supplier_quotation':sup_name})
    supplier_quotation=frappe.get_doc('Supplier Quotation',sup_name)
    supplier_quotation.set('custom_item_table',[])
    for i in sq_sows:
        sq_sow=frappe.get_doc("SQ SOW",i.name)

        for sow in sq_sow.manpower_table:
            supplier_quotation.append("custom_item_table",{
                'item_code':sow.item_code,
                'item_name':sow.item_name,
                'description':sow.description,
                'qty':sow.qty,
                'rate':sow.rate,
                'amount':sow.amount,
                'sow':sq_sow.sow_name
            })
        for sow in sq_sow.materials_table:
            supplier_quotation.append("custom_item_table",{
                'item_code':sow.item_code,
                'item_name':sow.item_name,
                'description':sow.description,
                'qty':sow.qty,
                'rate':sow.rate,
                'amount':sow.amount,
                'sow':sq_sow.sow_name
            })
    supplier_quotation.save(ignore_permissions=True)

@frappe.whitelist()
def create_user_permission_on_validate(doc,method):
    if frappe.db.exists("User",{'name':doc.custom_supplier_login_id,'enabled':1}):
        user=frappe.get_doc("User",{'name':doc.custom_supplier_login_id,'enabled':1})
        user.append("roles",{
            "role":"Supplier"
        })
        user.module_profile='Supplier'
        user.save(ignore_permissions=True)
        if not frappe.db.exists("User Permission",{'user':doc.custom_supplier_login_id,'allow':'Supplier','for_value':doc.name}):
            usr_per=frappe.new_doc("User Permission")
            usr_per.user=doc.custom_supplier_login_id
            usr_per.allow='Supplier'
            usr_per.for_value=doc.name
            usr_per.save(ignore_permissions=True)
        import string
        import random
        all_characters = string.ascii_letters + string.digits
        length = 6
        password = ''.join(random.choices(all_characters, k=length))
        if password:
            from frappe.utils.password import update_password

            update_password(user=user.name, pwd=password)
    else:
        user=frappe.new_doc("User")
        user.email=doc.custom_supplier_login_id
        user.first_name=doc.supplier_name
        user.append("roles",{
            "role":"Supplier"
        })
        user.module_profile='Supplier'
        user.send_welcome_email=0
        user.save(ignore_permissions=True)
        if not frappe.db.exists("User Permission",{'user':doc.custom_supplier_login_id,'allow':'Supplier','for_value':doc.name}):
            usr_per=frappe.new_doc("User Permission")
            usr_per.user=doc.custom_supplier_login_id
            usr_per.allow='Supplier'
            usr_per.for_value=doc.name
            usr_per.save(ignore_permissions=True)
        import string
        import random
        all_characters = string.ascii_letters + string.digits
        length = 6
        password = ''.join(random.choices(all_characters, k=length))
        if password:
            from frappe.utils.password import update_password

            update_password(user=user.name, pwd=password)

@frappe.whitelist()
def monthly_expiry_doc():
    current_date = nowdate()
    next_date = add_days(current_date, 30)
    docs = frappe.get_all("Legal Compliance Monitor",
                          {'next_due': ('between', (current_date, next_date))},
                          ['name','days_left','next_due'])


    recipients_docs_map = {}  # This dictionary will map recipients to the documents

    for i in docs:
        doc = frappe.get_doc("Legal Compliance Monitor", i.name)
        # For each recipient, append the document information
        for user in doc.alert_mails:
            recipient = user.user
            if recipient not in recipients_docs_map:
                recipients_docs_map[recipient] = []  # Initialize if not present
            recipients_docs_map[recipient].append(i)

    # Now, send an email for each recipient with their corresponding documents
    for recipient, recipient_docs in recipients_docs_map.items():
        # Create the document list table for this recipient
        documents = '''
        <table class="table table-bordered" border=1 style="border-width:2px; border-color:#000;">
            <tr>
                <td colspan="8" style=text-align:center><b>Legal Compliance Monitor List</b></td>
            </tr>
            <tr>
                <td colspan="2"><b>Document Name</b></td>
                <td colspan="2"><b>Expiry Date</b></td>
                <td colspan="2"><b>Days Left</b></td>
            </tr>
        '''

        for doc_info in recipient_docs:
            documents += '''
            <tr>
                <td colspan="2">{name}</td>
                <td colspan="2">{next_due}</td>
                <td colspan="2">{days_left}</td>
            </tr>
            '''.format(name=doc_info.name,
                       next_due=doc_info.next_due,
                       days_left=doc_info.days_left)

        documents += '</table>'

        # Send the email to this recipient
        frappe.sendmail(
            recipients=[recipient],
            subject="Legal Compliance Monitor Expiry Alert",
            message=f"""
                Dear Sir/Mam,<br><br>
                Kindly find the attached Legal Compliance Monitor List for Expiry Soon.<br><br>
                {documents}
                <br><br>
                Regards,<br>
                Your Company
            """
        )

    return True

@frappe.whitelist()
def update_days_left():
    value = frappe.get_all("Legal Compliance Monitor",["next_due","name"])
    for i in value:
        if i.next_due:
            diff = date_diff(i.next_due, today())
            frappe.db.set_value("Legal Compliance Monitor",i.name,"days_left",diff)
            next_due_date = frappe.utils.getdate(i.next_due)
            current_date = frappe.utils.getdate(today())
            next_due_month = next_due_date.month
            next_due_year = next_due_date.year
            current_month = current_date.month
            current_year = current_date.year
            if next_due_month == current_month and next_due_year == current_year and diff>=0:
                frappe.db.set_value("Legal Compliance Monitor", i.name, "status", "Expiring")
                frappe.db.set_value("Legal Compliance Monitor", i.name, "today_status", "Expiring")
            elif diff < 0:
                frappe.db.set_value("Legal Compliance Monitor", i.name, "status", "Expired")
                frappe.db.set_value("Legal Compliance Monitor", i.name, "today_status", "Expired")
            else:
                frappe.db.set_value("Legal Compliance Monitor", i.name, "status", "Active")
                frappe.db.set_value("Legal Compliance Monitor", i.name, "today_status", "Active")

@frappe.whitelist()
def update_days_left_monthly():
    job = frappe.db.exists('Scheduled Job Type', 'send_expiry_notifications_for_visa_on_december_1st')
    if not job:
        emc = frappe.new_doc("Scheduled Job Type")
        emc.update({
            "method": 'shaher.alerts.send_expiry_notifications_for_visa_on_december_1st',
            "frequency": 'Cron',
            "cron_format": '0 0 1 * *'  
        })
        emc.save(ignore_permissions=True)

@frappe.whitelist()
def get_gratuity(employee):
    from datetime import datetime
    from dateutil import relativedelta
    date_2 = datetime.now()
    doj = frappe.db.get_value('Employee',employee,'date_of_joining')
    absent_days = 0
    absent_days += frappe.db.get_value("Employee",{"name":employee},['custom_absent_days'])

    attendance = frappe.get_all(
        "Attendance",
            filters={
                "employee": employee,
                "attendance_date": ["between", (doj, nowdate())],
                "docstatus": ["=", 1],
                "leave_application":''
            },
            fields=["status"]
    )
    for i in attendance:
        if i.status =="Absent":
            absent_days +=1
        if i.status =="Half Day":
            absent_days+=0.5

    diff = relativedelta.relativedelta(date_2, doj)
    yos = cstr(diff.years) + ' years, ' + cstr(diff.months) +' months and ' + cstr(diff.days) + ' days'

    basic_salary = frappe.db.get_value('Employee',employee,'custom_basic')
    current_yoe_days = date_diff(nowdate(),doj) + 1
    current_yoe = round((current_yoe_days / 365),3)

    if(1 <= diff.years <= 3):
        gratuity_per_year = (basic_salary * .5)
        total_gratuity = math.ceil(gratuity_per_year/365*(current_yoe_days-absent_days))
        return total_gratuity
    elif(diff.years > 3):
        gratuity_per_year = basic_salary
        total_gratuity = math.ceil(gratuity_per_year/365*(current_yoe_days-absent_days))
        return total_gratuity

@frappe.whitelist()
def calculate_rejoining_date(to_date):
    return add_days(to_date,1)

@frappe.whitelist()
def validate_next_due_date(doc,method):
    if doc.leave_type=='Annual Leave' and doc.custom_is_extension == 0:
        days,doj = frappe.db.get_value("Employee",{'name':doc.employee},['custom_annual_leave_applicable_after','date_of_joining'])
        if days:
            if frappe.db.exists("Leave Application",{'employee':doc.employee,'leave_type':'Annual Leave','docstatus':('!=',2)}):
                count =frappe.db.count("Leave Application",{'employee':doc.employee,'leave_type':'Annual Leave','docstatus':('!=',2),'name':['!=',doc.name]})
                if count > 0:
                    leave_app = frappe.get_doc("Leave Application",{'employee':doc.employee,'leave_type':'Annual Leave','docstatus':('!=',2)},order_by='creation DESC')
                    if frappe.db.exists("Leave Application",{'employee':doc.employee,'leave_type':'Annual Leave','docstatus':('!=',2),'custom_is_extension':1,'from_date':('>',(add_days(leave_app.to_date,1)))}):
                        leave_app_extn = frappe.get_doc("Leave Application",{'employee':doc.employee,'leave_type':'Annual Leave','docstatus':('!=',2),'custom_is_extension':1,'from_date':('>',(add_days(leave_app.to_date,1)))},order_by='creation DESC')
                        end_date = leave_app_extn.to_date
                    else:
                        end_date = leave_app.to_date
                    next_due_date = add_days(end_date,int(days))
                    current_due_date = date_diff(doc.from_date,end_date) + 1
                    formatted_next_due_date = frappe.utils.formatdate(next_due_date, 'dd-mm-yyyy')
                    if int(days) > current_due_date:
                        frappe.throw(f"You can apply Annual Leave only after Next Due Date: {formatted_next_due_date}")
                else:
                    next_due_date = add_days(doj,int(days))
                    current_due_date = date_diff(today(),doj) + 1
                    formatted_next_due_date = frappe.utils.formatdate(next_due_date, 'dd-mm-yyyy')
                    if int(days) > current_due_date:
                        frappe.throw(f"You can apply Annual Leave after the Due Date: {formatted_next_due_date}")
        else:
            frappe.throw("Annual Leave Applicable after Days is not updated in Employee MIS.Kindly update that and check")


@frappe.whitelist()
def email_salary_slip(from_date):
    ss = frappe.get_all("Salary Slip",{'start_date':from_date,'docstatus':1},['employee','name'])
    for s in ss:
        doc = frappe.get_doc("Salary Slip",s['name'])
        receiver = frappe.db.get_value("Employee", doc.employee, "user_id")
        if not receiver:
            receiver = frappe.db.get_value("Employee", doc.employee, "company_email")
        message = f"Dear {doc.employee_name},<br><br> Kindly check your payslip for the period {doc.start_date.strftime('%d-%m-%Y')} and {doc.end_date.strftime('%d-%m-%Y')}."

        if receiver:
            email_args = {
                "recipients":receiver,
                "message": _(message),
                "subject": 'Salary Slip - from {0} to {1}'.format(doc.start_date, doc.end_date),
                "attachments": [frappe.attach_print("Salary Slip", doc.name,
                    file_name="Pay Slip", print_format="Pay Slip New")],
                "reference_doctype": doc.doctype,
                "reference_name": doc.name
                }
            if not frappe.flags.in_test:
                print('yes')
                enqueue(method=frappe.sendmail, queue='short', timeout=300, is_async=True, **email_args)
            else:
                print('No')
                frappe.sendmail(**email_args)
        else:
            print(_("{0}: Employee email not found, hence email not sent").format(doc.employee_name))

@frappe.whitelist()
def ot_calculation(name):
    doc=frappe.get_doc("Attendance",name)
    basic=frappe.db.get_value("Employee",{'name':doc.employee},['custom_basic'])
    if doc.custom_work_place:
        if doc.custom_work_place=='Office':
            wh=int(9)
        else:
            wh=int(10)
    if not basic:
        basic=0
    att=getdate(doc.attendance_date)
    year = att.year
    month = att.month
    total_days_in_month = calendar.monthrange(year, month)[1]
    hh=check_holiday(att,doc.employee)
    if hh:
        ot=2
    else:
        ot=1.5
    start=frappe.db.get_value("Shift Type",{'name':'G'},['start_time'])
    if isinstance(start, timedelta):
        start = (datetime.min + start).time()
    start_date=datetime.combine(att,start)
    end=frappe.db.get_value("Shift Type",{'name':'G'},['end_time'])
    if isinstance(end, timedelta):
        end = (datetime.min + end).time()
    end_date=datetime.combine(att,end)
    end_time = datetime.combine(att, time(19, 0, 0))
    in_time = get_datetime(doc.in_time) if doc.in_time else None
    out_time = get_datetime(doc.out_time) if doc.out_time else None
    if in_time and out_time:
        if in_time <= start_date and out_time>end_time:
            if basic>0:
                per_hour=basic/total_days_in_month/wh
                ot_amount=per_hour*float(ot)
                if hh:
                    ot_hrs=time_diff_in_hours(out_time,end_date)
                else:
                    ot_hrs=time_diff_in_hours(out_time,in_time)
                tot_ot=ot_amount*ot_hrs
            else:
                ot_hrs=0
                tot_ot=0
        else:
            ot_hrs=0
            tot_ot=0
    else:
        ot_hrs=0
        tot_ot=0
    return ot_hrs,tot_ot


@frappe.whitelist()
def check_holiday(date,emp):
    holiday_list = frappe.db.get_value('Employee',emp,'holiday_list')
    holiday = frappe.db.sql("""select `tabHoliday`.holiday_date,`tabHoliday`.weekly_off from `tabHoliday List`
    left join `tabHoliday` on `tabHoliday`.parent = `tabHoliday List`.name where `tabHoliday List`.name = '%s' and holiday_date = '%s' """%(holiday_list,date),as_dict=True)
    if holiday:
        return 'H'

@frappe.whitelist()
def update_service_entry_number(docname,se_no, se_date):
    frappe.db.set_value("Delivery Note", docname, "service_entry_number", se_no)
    frappe.db.set_value("Delivery Note", docname, "custom_service_entry_date", se_date)
    return "done"

@frappe.whitelist()
def custom_round(value):
    integer_part = int(value)
    decimal_part = value - integer_part
    if decimal_part < 0.50:
        decimal_part = 0.00
    else:
        decimal_part = 0.50
    amount=integer_part + decimal_part
    amount=round(amount,2)
    return integer_part + decimal_part

# To update the DN's date in the SI's items table
@frappe.whitelist()
def udpate_date_of_supply(doc, method):
    for row in doc.items:
        if row.delivery_note:
            delivery_date = frappe.db.get_value("Delivery Note", row.delivery_note, "posting_date")
            service_entry = frappe.db.get_value("Delivery Note", row.delivery_note, "service_entry_number")
            row.custom_date_of_supply = delivery_date
            row.custom_service_entry_number = service_entry

import frappe

import frappe

def validate_purchase_invoice(doc, method):

    purchase_orders = list(set(d.purchase_order for d in doc.items if d.purchase_order))

    if not purchase_orders:
        return


    sales_orders = frappe.get_all("Purchase Order Item",
                                  filters={"parent": ["in", purchase_orders]},
                                  fields=["sales_order"])

    sales_order_names = list(set(so.sales_order for so in sales_orders if so.sales_order))

    if not sales_order_names:
        return


    sales_order_data = frappe.get_all("Sales Order",
                                      filters={"name": ["in", sales_order_names], "docstatus": ["!=", 2]},
                                      fields=["name", "grand_total", "per_delivered"])

    for so in sales_order_data:

        if so.per_delivered < 100:
            frappe.throw(f"Cannot save Purchase Invoice: Sales Order {so.name} is not fully delivered.")


        sales_invoices = frappe.get_all("Sales Invoice",
                                        filters={"sales_order": so.name, "docstatus": ["!=", 2]},
                                        fields=["name", "outstanding_amount", "grand_total"])

        for si in sales_invoices:

            # if si.outstanding_amount > 0:
            #     frappe.throw(f"Cannot save Purchase Invoice: Sales Invoice {si.name} is not fully paid.")


            payment_total = frappe.db.sql("""
                SELECT SUM(per.allocated_amount) FROM `tabPayment Entry Reference` per
                JOIN `tabPayment Entry` pe ON per.parent = pe.name
                WHERE per.reference_doctype = 'Sales Invoice' AND per.reference_name = %s
                AND pe.docstatus = 1
            """, (si.name,))[0][0] or 0


            if payment_total != si.grand_total:
                frappe.throw(f"Cannot save Purchase Invoice: Payment Entry total ({payment_total}) "
                             f"does not match Sales Invoice Grand Total ({si.grand_total}) for {si.name}.")


@frappe.whitelist()
def get_project_items(doctype, txt, searchfield, start, page_len, filters):
    project = filters.get("project") if filters else None
    if project:
        doc = frappe.get_doc("Project", project)
        return [(row.item_code,) for row in doc.custom_items]

@frappe.whitelist()
def get_qty_rate(item_code, project):
    qty = frappe.db.get_value("Project Item Table", {"item_code": item_code, "parent": project}, "qty")
    rate = frappe.db.get_value("Project Item Table", {"item_code": item_code, "parent": project}, "rate")
    return qty, rate

@frappe.whitelist()
def update_employee_certification(doc, method):
    if doc.custom_employee_certification:
        ec = frappe.get_doc("Employee Certification", doc.custom_employee_certification)
        ec.purchase_order = doc.name
        ec.save(ignore_permissions=True)

@frappe.whitelist()
def create_vehicle_maintenance_check(doc, methods):
    if "Vehicle" in doc.custom_department:
        for row in doc.items:
            vmc = frappe.new_doc("Vehicle Maintenance Check List")
            vmc.register_no = row.custom_vehicle
            vmc.supplier = doc.supplier
            vmc.actual_expense = row.base_net_amount
            vmc.vehicle_service = row.item_code
            vmc.work_finished_date = row.schedule_date
            vmc.present_kilometer = row.custom_present_kilometer
            vmc.purchase_order = doc.name
            vmc.save(ignore_permissions=True)
            vmc.submit()
            row.custom_vehicle_maintenance_check = vmc.name
        frappe.msgprint("Vehicle Maintenance Check List created successfully.")

@frappe.whitelist()
def remove_vmc_linked(doc, mthod):
    if "Vehicle" in doc.custom_department:
        for row in doc.items:
            if row.custom_vehicle_maintenance_check:
                row.custom_vehicle_maintenance_check = ""
            if row.custom_vehicle:
                row.custom_vehicle = ""

@frappe.whitelist()
def cancel_vehicle_maintenance_check(doc, methods):
    if "Vehicle" in doc.custom_department:
        if frappe.db.exists("Vehicle Maintenance Check List", {"purchase_order": doc.name, "docstatus": 1}):
            vmcs = frappe.get_all("Vehicle Maintenance Check List", {"purchase_order": doc.name, "docstatus": 1}, "name")
            for i in vmcs:
                vmc = frappe.get_doc("Vehicle Maintenance Check List", i.name)
                if vmc.docstatus == 1:
                    vmc.cancel(ignore_permissions=True)
                if vmc.docstatus == 0:
                    vmc.delete(ignore_permissions=True)

@frappe.whitelist()
def validate_kilometer(doc, methods):
    if "Vehicle" in doc.custom_department:
        for row in doc.items:
            if row.custom_present_kilometer and row.custom_last_kilometer:
                if row.custom_present_kilometer < row.custom_last_kilometer:
                    frappe.throw("<b>Present Kilometer</b> should be greater than the <b>Last Kilometer</b>")

@frappe.whitelist()
def pdo_validation(doc, methods):
    # Check Service Entry
    if doc.customer =="Petroleum Development Oman LLC":
        idx = 0
        for row in doc.items:
            idx+=1
            if not row.custom_service_entry_number:
                frappe.throw(
                    f"<b>Row {idx}</b>: Service Entry Number is not found for the Item {row.item_code}",
                    title="Service Entry Not Found"
                )

@frappe.whitelist()
def validate_grand_total(so,date,total):
    grand_total=frappe.db.get_value("Sales Order",so,'base_grand_total')
    qtn=frappe.db.get_value("Sales Order",so,'quotation')
    profit_per=frappe.db.get_value("Quotation",qtn,'custom_prof_percentage') or 0
    purchase_orders=frappe.get_all("Purchase Order",{'custom_sales_order':so,'docstatus':1,'transaction_date':('<=',date)},['base_grand_total'])
    total_base_grand_total = sum(po['base_grand_total'] for po in purchase_orders) + float(total)
    # if profit_per:
    percentage_difference = ((total_base_grand_total - (grand_total*((100-profit_per)/100))) / (grand_total*((100-profit_per)/100))) * 100
    # else:
    # 	percentage_difference = 0
    if total_base_grand_total > (grand_total*((100-profit_per)/100)):
        return "above",round(percentage_difference,2)
    if total_base_grand_total < (grand_total*((100-profit_per)/100)):
        return "below",round(abs(percentage_difference),2)
    if total_base_grand_total == (grand_total*((100-profit_per)/100)):
        return 'equals',0

@frappe.whitelist()
def validate_grand_total_qtn(supp_qtn,total,profit_per):
    total_base_grand_total=frappe.db.get_value("Supplier Quotation",supp_qtn,'base_grand_total')
    threshold = float(total) * ((100 - float(profit_per)) / 100)

    # Calculate percentage difference
    percentage_difference = ((total_base_grand_total - threshold) / threshold) * 100

    if total_base_grand_total > threshold:
        return "above", round(percentage_difference,2)

@frappe.whitelist()
def update_coc_fields(doc, methods):
    if "pdo" in doc.custom_department.lower() and len(doc.items) > 0:
        first_row = doc.items[0]  # Get the first row
        so = first_row.against_sales_order
        wo = first_row.custom_work_order_number
        # start_date, po_value = frappe.db.get_value("Sales Order", so, ["transaction_date", "grand_total"])
        start_date = frappe.db.get_value("Sales Order", so, ["transaction_date"])
        end_date = doc.custom_approval_date
        # doc.custom_purchase_order_value = po_value if po_value else ''
        # doc.custom_work_start_date = start_date if start_date else ''
        doc.custom_work_finish_date = end_date
        doc.custom_purchase_order_no = doc.po_no
        doc.custom_actual_work_carried_out_ = doc.grand_total
        doc.custom_work_order_no = wo

@frappe.whitelist()
def get_user_details(user):
    employee_sign = frappe.db.get_value("Employee",{'name':user},['custom_digital_signature'])
    if employee_sign:
        return "https://shaher.teamproit.com/"+employee_sign

# @frappe.whitelist()
# def rejection_remark(name, remark):
# 	frappe.db.set_value("Salary Certificate", name, "rejection_remark", remark)
# 	doc = frappe.get_doc("Salary Certificate", name)
# 	doc.save()
# 	frappe.db.set_value("Salary Certificate", name, "workflow_state", "Rejected")

@frappe.whitelist()
def update_approval_date_mr(doc,method):
    if(doc.workflow_state == "Approved"):
        frappe.db.set_value("Material Request",doc.name,"custom_approval_date",today())

@frappe.whitelist()
def update_approval_date_po(doc,method):
    if(doc.workflow_state == "Approved"):
        frappe.db.set_value("Purchase Order",doc.name,"custom_approval_date",today())

@frappe.whitelist()
def update_approval_date_pi(doc,method):
    frappe.db.set_value("Purchase Invoice",doc.name,"custom_approval_date",today())

@frappe.whitelist()
def update_approval_date_pr(doc,method):
    frappe.db.set_value("Purchase Receipt",doc.name,"custom_approval_date",today())

@frappe.whitelist()
def make_quotation(source_name, target_doc=None):
    def postprocess(source, target):
        # Map 'Provisional Budgeting Child Expense' manually since get_mapped_doc doesn't support multiple child tables
        for expense in source.get("expenses"):
            target.append("expenses", {
                "account": expense.account,
                "current_value": expense.current_value,
                "provisional_value": expense.provisional_value
            })

    doclist = get_mapped_doc("Provisional Budgeting", source_name, {
        "Provisional Budgeting": {
            "doctype": "Provisional Budgeting Comparison",
            "field_map": {
                "name": "provisional_budgeting",
            }
        },
        "Provisional Budgeting Child": {
            "doctype": "Provisional Budgeting Comparison Child",
            "field_map": {
                "account": "account",
                "current_value": "current_value",
                "provisional_value": "provisional_value",
            }
        }
    }, target_doc, postprocess)

    return doclist

@frappe.whitelist()
def check_inactive_employee(employee):

    status = frappe.db.get_value("Employee", {"employee": employee}, ["status"])

    if status == "Active":
        return "Active"
    else:
        return "Inactive"

#Enable the hod checkbox in Leave application , Leave salary and final exit request,Rejoining form , additional vacation,HR Request form based on below condition
@frappe.whitelist()
def get_role(employee):
    user_id = frappe.get_value('Employee',{'employee':employee},['user_id'])
    hod = frappe.get_value('User',{'email':user_id},['name'])
    role = "HOD"
    hod = frappe.get_value('Has Role',{'role':role,'parent':hod})
    if hod:
        return  "yes"
    else:
        return "no"

@frappe.whitelist()
def get_next_account_number(parent_account):
    if not parent_account:
        return "Done"

    series = frappe.db.sql(
        """
        SELECT account_number FROM `tabAccount`
        WHERE parent_account=%s AND disabled=0
        ORDER BY account_number DESC
        LIMIT 1
        """,
        (parent_account,),
        as_dict=False
    )
    series_par = frappe.db.get_value("Account",{'name':parent_account,'disabled':0},['account_number'])
    if series and series[0][0]:
        account_number = str(int(series[0][0]) + 1)
        return account_number

    elif series_par:
        account_number = series_par
        return account_number
    else:
        return "Done"

@frappe.whitelist()
def get_leave_periods(date_of_joining, relieving_date, employee):
    emp = frappe.get_doc("Employee", employee)
    basic = emp.custom_basic

    periods = []
    current = date_of_joining
    year = 1
    total_amount=0
    current=getdate(current)
    relieving_date=getdate(relieving_date)
    while current < relieving_date:
        try:
            next_year = current.replace(year=current.year + 1)
        except ValueError:
            next_year = current + timedelta(days=365)

        to_date = min(next_year, relieving_date)
        diff = (to_date - current).days

        if year <= 3:
            amount = (diff / 2) * (basic / 365)
        else:
            amount = (diff) * (basic / 365)
        total_amount+=amount
        periods.append({
            "year_number": year,
            "from_date": current,
            "to_date": to_date,
            "diff_days": diff,
            "amount": round(amount, 3)
        })

        current = to_date
        year += 1

    return {
        "periods": periods,
        "total_amount": round(total_amount, 3)
    }

@frappe.whitelist()
def get_tot_leave_amount(name):
    doc=frappe.get_doc("Full and Final Settlement",name)
    tot=0
    for i in doc.leave_salary:
        tot+=i.amount
    return tot

@frappe.whitelist()
def generate_leave_summary(date_of_joining, relieving_date, employee):
    leave_data = get_leave_periods(date_of_joining, relieving_date, employee)
    periods = leave_data.get("periods", [])
    total_amount = leave_data.get("total_amount", 0)
    html = """
    <table class="table table-bordered">
        <thead>
            <tr>
                <th>From Date</th>
                <th>To Date</th>
                <th>Diff (Days)</th>
                <th>Year</th>
                <th>Amount</th>
            </tr>
        </thead>
        <tbody>
    """
    for row in periods:
        html += f"""
            <tr>
                <td>{row['from_date']}</td>
                <td>{row['to_date']}</td>
                <td>{row['diff_days']}</td>
                <td>{row['year_number']}</td>
                <td>{row['amount']}</td>
            </tr>
        """

    html += f"""
        </tbody>
        <tfoot>
            <tr>
                <td colspan="4"><strong>Total Amount</strong></td>
                <td><strong>{total_amount}</strong></td>
            </tr>
        </tfoot>
    </table>
    """

    return html

@frappe.whitelist()
def validate_project_budget(doc, method):
    if not doc.custom_project_budget and doc.order_type == "Project":
        frappe.throw(
            "Create Project Budget before submitting the Sales Order",
            title = "Not Allowed"
        )


@frappe.whitelist()
def get_template():
    args = frappe.local.form_dict
    w = UnicodeWriter()
    w = add_header(w,args)

    frappe.response['result'] = cstr(w.getvalue())
    frappe.response['type'] = 'csv'
    frappe.response['doctype'] = "Bank Statement"

def add_header(w,args):
    args.company = frappe.db.get_value("Bank Statement Import",args.name,'company')
    frappe.log_error(args.company,args.bank_account)
    if 'Bank Dhofar' in args.bank_account:
        w.writerow(['Transaction Date','Value Date','Transaction Category','Cheque No.','Description','Debit (-)','Credit (+)','Balance','Book Date'])
    elif 'Bank Muscat' in args.bank_account:
        w.writerow(['Transaction Date','Value Date','Transaction Remarks','Dr Amount','Cr Amount','Running Balance','Book Date'])
    elif 'Alizz' in args.bank_account:
        w.writerow(['Transaction Date','Description','Book Date','Value Date','Debit (-)','Credit (+)','Balance'])
    return w

@frappe.whitelist()
def create_bank_transaction(file,company,bank_account):
    filepath = get_file(file)
    pps = read_csv_content(filepath[1])
    count = 0
    for pp in pps:
        if pp[0] != 'Transaction Date':


            doc = frappe.new_doc('Bank Transaction')
            doc.company = company
            doc.bank_account = bank_account

            if 'Bank Dhofar' in bank_account:
                date_obj = datetime.strptime(pp[0], "%d/%m/%Y")  # or "%d/%m/%Y" depending on input
                formatted_date = date_obj.strftime("%Y-%m-%d")
                doc.date = formatted_date
                doc.reference_number = pp[3]
                doc.description = pp[4]
                doc.deposit = pp[5]
                doc.withdrawal = pp[6]
            elif 'Bank Muscat' in bank_account:
                date_obj = datetime.strptime(pp[0], "%d/%m/%Y")  # or "%d/%m/%Y" depending on input
                formatted_date = date_obj.strftime("%Y-%m-%d")
                doc.date = formatted_date
                doc.description = pp[2]
                doc.deposit = pp[3]
                doc.withdrawal = pp[4]
            elif 'Alizz' in bank_account:
                date_obj = datetime.strptime(pp[0], '%d-%b-%y')
                formatted_date = date_obj.strftime('%Y-%m-%d')
                doc.date = formatted_date
                doc.description = pp[1]
                doc.deposit = pp[4]
                doc.withdrawal = pp[5]
            doc.save(ignore_permissions=True)
            doc.submit()
            frappe.db.commit()
            count+=1
    return count

@frappe.whitelist()
def update_dn_status(doc, method):
    if "pdo" in doc.custom_department.lower():
        if doc.workflow_state == "Pending for SE":
            if doc.service_entry_number:
                apply_workflow(doc, "To Bill")

        if doc.workflow_state == "To Bill":
            if not doc.service_entry_number:
                apply_workflow(doc, "Revert")


@frappe.whitelist()
def update_workflow_on_cancel(doc, method):
    doc.workflow_state == "Cancelled"

@frappe.whitelist()
def update_dn_workflow(doc, method):
    delivery_note = frappe.db.get_value("Sales Invoice Item", {"parent": doc.name}, "delivery_note")
    if frappe.db.exists("Delivery Note", {"name": delivery_note, "status": "Completed"}):
        frappe.db.set_value("Delivery Note", {"name": delivery_note, "status": "Completed"}, "workflow_state", "Completed")

@frappe.whitelist()
def update_workflow_on_cancelling_si(doc, method):
    delivery_note = frappe.db.get_value("Sales Invoice Item", {"parent": doc.name}, "delivery_note")
    if frappe.db.exists("Delivery Note", {"name": delivery_note, "workflow_state": "Completed"}):
        frappe.db.set_value("Delivery Note", {"name": delivery_note, "workflow_state": "Completed"}, "workflow_state", "To Bill")



@frappe.whitelist()
def custom_sutc_job_no_validation(doc, method):
    if doc.is_new():
        posting_date =  nowdate()
        if isinstance(posting_date, str):
            posting_date = datetime.strptime(posting_date, "%Y-%m-%d")

        month = posting_date.strftime('%m')
        year = posting_date.strftime('%y')



        loc_abbr = get_abbr(doc.custom_site) if doc.custom_site else ""

        count = frappe.db.count("Delivery Note", {"posting_date": posting_date})
        count_str = str(count + 1).zfill(2)

        if loc_abbr:
            job_no = f"{count_str}-{month}-{loc_abbr}-{year}"
        else:
            job_no = f"{count_str}-{month}-{year}"

        doc.custom_sutc_job_no = job_no


@frappe.whitelist()
def read_service_entry_pdf(file_url):
    # Step 1: Get full path of the uploaded PDF file
    relative_path = file_url.lstrip("/").replace("files/", "", 1)
    pdf_path = frappe.get_site_path("public", "files", relative_path)

    # Step 2: Prepare .txt file path (store PDF as .txt)
    base_filename = os.path.splitext(os.path.basename(pdf_path))[0]
    text_filename = base_filename + ".txt"
    text_path = frappe.get_site_path("public", "files", text_filename)

    # Step 3: Copy raw binary PDF content into .txt file
    with open(pdf_path, "rb") as pdf_file:
        pdf_data = pdf_file.read()

    with open(text_path, "wb") as text_file:
        text_file.write(pdf_data)

    # Step 4: Read binary and decode it (safely)
    with open(text_path, "rb") as f:
        binary_content = f.read()

    try:
        content = binary_content.decode("latin1")
    except UnicodeDecodeError:
        return None

    # Step 5: Extract form field values using regex
    fields = {}

    match_entry = re.search(r"/T\(service_entry\).*?/V\((.*?)\)", content, re.DOTALL)
    if match_entry:
        fields["service_entry"] = match_entry.group(1).strip()

    match_date = re.search(r"/T\(service_entry_date\).*?/V\((.*?)\)", content, re.DOTALL)
    if match_date:
        fields["service_entry_date"] = match_date.group(1).strip()

    if fields:
        return fields
    else:
        return None


@frappe.whitelist()
def editable_coc_print(docname):
    html = frappe.get_print('Delivery Note', docname, print_format='Certificate of Completion')
    base_pdf = get_pdf(html)

    packet = io.BytesIO()
    can = canvas.Canvas(packet, pagesize=A4)

    form = can.acroForm
    form.textfield(
        name='service_entry',
        tooltip='Enter Service Entry Sheet No',
        x=192, y=68,
        width=122, height=20,
        borderColor=colors.black,
        fillColor=colors.white,
        textColor=colors.black,
        fontName="Helvetica",
        fontSize=10,
        forceBorder=True
    )

    form.textfield(
        name='service_entry_date',
        tooltip='Enter Date',
        x=423, y=53,
        width=100, height=17,
        borderColor=colors.black,
        fillColor=colors.white,
        textColor=colors.black,
        fontName="Helvetica",
        fontSize=10,
        forceBorder=True,
        value="dd-mm-yyyy"
    )

    can.showPage()
    can.save()
    packet.seek(0)

    base_pdf_reader = PdfReader(io.BytesIO(base_pdf))
    overlay_reader = PdfReader(packet)
    writer = PdfWriter()

    page = base_pdf_reader.pages[0]
    page.merge_page(overlay_reader.pages[0])
    writer.add_page(page)

    output_filename = f"{docname}-editable.pdf"
    output_path = frappe.get_site_path("public", "files", output_filename)

    with open(output_path, "wb") as f:
        writer.write(f)

    return "/files/" + output_filename


@frappe.whitelist()
def get_pending_payment(pi_name):
    html = """
        <style>
            td {
                padding-right: 8px;
                padding-left: 8px;
                padding-top: 4px;
                padding-bottom: 4px;
            }
        </style>
            <div style="overflow-x: auto; color: white;">
                <p style="overflow: hidden;color:black">This is a Back to Back Order with Pending Payment, please find the below attachment for your reference</p>
                <table class="text-center" style="overflow: hidden; white-space: nowrap; margin-bottom: 50px;width:100%">
                    <tr style="background-color: #e8262e; font-weight: 500; color: white;">
                        <td class="border border-1 border-light" style="min-width: 80px;">S.No.</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Customer PO</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Item Name</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Sales Invoice</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Date of Invoice</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Amount</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Payment Status</td>
                    </tr>
    """
    if pi_name:
        purchase_invoice = pi_name
        pi = frappe.get_doc("Purchase Invoice", purchase_invoice)
        idx = 1
        for row in pi.items:
            purchase_order = row.purchase_order
            purchase_order_item = row.po_detail

            sales_order = frappe.db.get_value("Purchase Order Item", purchase_order_item, "sales_order")
            pdo_po = frappe.db.get_value("Sales Order", sales_order, "po_no")
            sales_order_item = frappe.db.get_value("Purchase Order Item", purchase_order_item, "sales_order_item")

            # service_entry = get_service_entry(sales_order_item, sales_order)
            sales_invoice, sales_invoice_amount, sales_invoice_date, return_sales_invoice = get_sales_invoice_details(sales_order_item, sales_order)
            # purchase_invoice, purchase_invoice_amount, purchase_invoice_date  = get_purchase_invoice_details(row.name, purchase_order)
            # purchase_invoice, purchase_invoice_amount, purchase_invoice_date  = pi.name, pi.grand_total, pi.posting_date
            customer_payment_entry, customer_paid_amount, customer_payment_status = "", 0, "Pending"
            supplier_payment_entry, supplier_paid_amount, supplier_payment_status = "", 0, "Pending"
            if sales_invoice:
                customer_payment_entry, customer_paid_amount, customer_payment_status = get_payment_details(sales_invoice, sales_invoice_amount)
            # if purchase_invoice:
            # 	supplier_payment_entry, supplier_paid_amount, supplier_payment_status = get_payment_details(purchase_invoice, purchase_invoice_amount)

            if not return_sales_invoice and not pi.is_return and customer_payment_status!="Paid" and sales_invoice:
                html += f""""
                        <tr style="color: black;">
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: right;">{idx}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{pdo_po or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{row.item_name or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{sales_invoice or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{sales_invoice_date.strftime('%d-%m-%Y') or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: right;">{fmt_money(sales_invoice_amount, currency = pi.currency) or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{customer_payment_status or ''}</td>
                        </tr>
                """
                idx += 1
            elif not return_sales_invoice and not pi.is_return and not sales_invoice and sales_order:
                html += f""""
                        <tr style="color: black;">
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: right;">{idx}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{pdo_po or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{row.item_name or ''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{'Invoice not yet created'}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: right;">{''}</td>
                            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">{customer_payment_status or ''}</td>
                        </tr>
                """
                idx += 1
        html += """

            </table>
        </div>
        """
    if idx > 1:
        return html

@frappe.whitelist()
def get_sales_invoice_details(sales_order_item, sales_order):
    sales_invoice = frappe.db.sql("""
                SELECT
                    si.name as sales_invoice,
                    sii.amount as amount,
                    si.posting_date as date,
                    si.is_return

                FROM
                    `tabSales Invoice` si
                INNER JOIN
                    `tabSales Invoice Item` sii ON sii.parent = si.name
                WHERE
                    si.docstatus = 1
                    AND sii.so_detail = %s
                    AND sii.sales_order = %s
                LIMIT 1
            """, (sales_order_item, sales_order), as_dict=True)

    name = sales_invoice[0]['sales_invoice'] if sales_invoice else None
    amount = sales_invoice[0]['amount'] if sales_invoice else None
    date = sales_invoice[0]['date'] if sales_invoice else None
    is_return = sales_invoice[0]['is_return'] if sales_invoice else 0
    return name, amount, date, is_return

@frappe.whitelist()
def get_payment_details(invoice, invoice_amount):
    payment_entry = frappe.db.sql("""
                    SELECT
                        pe.name as payment_entry,
                        per.allocated_amount as paid_amount
                    FROM
                        `tabPayment Entry` pe
                    INNER JOIN
                        `tabPayment Entry Reference` per ON per.parent = pe.name
                    WHERE
                        pe.docstatus = 1
                        AND per.reference_name = %s
                    LIMIT 1
                """, (invoice), as_dict=True)

    name = payment_entry[0]['payment_entry'] if payment_entry else None
    amount = payment_entry[0]['paid_amount'] if payment_entry else None
    status = "Pending"
    if payment_entry:
        if amount == invoice_amount or amount > invoice_amount:
            status = "Paid"
        else:
            status = "Partly Paid"
    return name, amount, status

@frappe.whitelist()
def create_journal_entry(doc,method):
    amt = 0

    for item in doc.items:
        if frappe.db.exists('Item', {'name': item.item_code}):
            rate = frappe.db.get_value("Sales Order Item", {'name': item.so_detail}, "base_rate")
            if rate:
                amt += (rate*item.qty) or    0
            else:
                amt += 0
    if amt <= 0:
        return

    je = frappe.new_doc("Journal Entry")
    je.company = doc.company
    je.posting_date = doc.posting_date
    je.voucher_type = "Journal Entry"
    je.custom_cost_center = doc.cost_center
    je.custom_project = doc.project
    debit_acc = frappe.db.get_value(
        "Account",
        {'name': ('like', '%Contract Asset%'), 'company': doc.company}
    )
    # if "pdo" in doc.custom_department.lower():
    #     credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Rev Maintenance Contract PDO%'), 'company': doc.company})
    # elif "oetc" in doc.custom_department.lower():
    #     credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Rev Maintenance Contract OETC%'), 'company': doc.company})
    # else:
    credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Unbilled Revenue%'), 'company': doc.company})
    je.append('accounts', {
        'account': debit_acc,
        'debit_in_account_currency': amt,
        'cost_center':doc.cost_center,
        'project':doc.project
        })
    je.append('accounts', {
        'account': credit_acc,
        'credit_in_account_currency': amt,
        'cost_center':doc.cost_center,
        'project':doc.project
    })
    je.user_remark = doc.name
    je.custom_delivery_note = doc.name
    je.save(ignore_permissions=True)
    je.submit()


import frappe

@frappe.whitelist()
def cancel_journal_entry(doc, method):
    # Get submitted JE linked by user_remark
    je_name = frappe.db.get_value('Journal Entry', {'user_remark': doc.name, 'docstatus': 1})
    if je_name:
        rev_je_name = frappe.db.get_value('Journal Entry', {'reversal_of': je_name, 'docstatus': 1})
        if rev_je_name:
            frappe.get_doc('Journal Entry', {'reversal_of': je_name, 'docstatus': 1})
            if rev_je_name.docstatus == 1:
                rev_je_name.cancel()
        je = frappe.get_doc('Journal Entry', je_name)
        je.cancel()
        frappe.db.commit()

@frappe.whitelist()
def update_dn_field(doc,method):
    doc.custom_delivery_note =None
    rev_je_name = frappe.db.get_value('Journal Entry', {'reversal_of': doc.name, 'docstatus': 1})
    if rev_je_name:
        rev_je = frappe.get_doc('Journal Entry', rev_je_name)
        rev_je.reversal_of = None
        rev_je.save(ignore_permissions=True)
        rev_je.reload()
        rev_je.cancel()

@frappe.whitelist()
def create_reversal_entry(doc,method):
    dn=''
    credit_amt=0
    debit_amt=0
    for i in doc.items:
        if i.delivery_note:
            dn=i.delivery_note
            rate = frappe.db.get_value("Sales Order Item", {'name': i.so_detail}, "base_rate")
            if rate:
                credit_amt += (rate*i.qty) or 0
                debit_amt += (rate*i.qty) or 0
            else:
                credit_amt += 0
                debit_amt += 0
    if credit_amt <= 0:
        return
    if debit_amt <= 0:
        return
    if frappe.db.exists('Journal Entry',{'user_remark':dn,'docstatus':1}):
        je=frappe.get_doc('Journal Entry',{'user_remark':dn,'docstatus':1})
        # for j in je.accounts:
        #     credit_amt+=j.credit_in_account_currency
        #     debit_amt+=j.debit_in_account_currency
        rev_entry = frappe.new_doc("Journal Entry")
        rev_entry.company = doc.company
        rev_entry.posting_date = doc.posting_date
        rev_entry.voucher_type = "Journal Entry"
        debit_acc = frappe.db.get_value(
            "Account",
            {'name': ('like', '%Contract Asset%'), 'company': doc.company}
        )
        # if "pdo" in doc.custom_department.lower():
        #     credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Rev Maintenance Contract PDO%'), 'company': doc.company})
        # elif "oetc" in doc.custom_department.lower():
        #     credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Rev Maintenance Contract OETC%'), 'company': doc.company})
        # else:
        credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Unbilled Revenue%'), 'company': doc.company})
        rev_entry.append('accounts', {
            'account': debit_acc,
            'credit_in_account_currency': credit_amt,
            'reference_type':'Journal Entry',
            'reference_name':je.name,
            'cost_center':doc.cost_center,
            'project':doc.project
        })
        rev_entry.append('accounts', {
            'account': credit_acc,
            'debit_in_account_currency': debit_amt,
            'reference_type':'Journal Entry',
            'reference_name':je.name,
            'cost_center':doc.cost_center,
            'project':doc.project
        })
        rev_entry.user_remark = doc.name
        rev_entry.reversal_of = je.name
        rev_entry.cost_center=doc.cost_center
        rev_entry.project = doc.project
        rev_entry.custom_references=str(je.name)+','+str(doc.name)
        rev_entry.save(ignore_permissions=True)
        rev_entry.submit()

@frappe.whitelist()
def cancel_reverse_entry(doc, method):
    je_name = frappe.db.get_value('Journal Entry', {'user_remark': doc.name, 'docstatus': 1})
    if je_name:
        je = frappe.get_doc('Journal Entry', je_name)
        if je.docstatus == 1:
            je.cancel()
        frappe.db.commit()

@frappe.whitelist()
def create_journal_entry_for_purchase(doc,method):
    amt = 0

    for item in doc.items:
        if item.item_type =='Service Item':
            if frappe.db.exists('Item', {'name': item.item_code}):
                rate = frappe.db.get_value("Purchase Order Item", {'name': item.purchase_order_item}, "base_rate")
                if rate:
                    amt += (rate*item.qty) or    0
                else:
                    amt += 0
    if amt <= 0:
        return

    je = frappe.new_doc("Journal Entry")
    je.company = doc.company
    je.posting_date = doc.posting_date
    je.voucher_type = "Journal Entry"
    je.custom_cost_center = doc.cost_center
    je.custom_project = doc.project
    debit_acc = frappe.db.get_value(
        "Account",
        {'name': ('like', '%Unbilled Revenue%'), 'company': doc.company}
    )
    credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Service Received But Not Billed%'), 'company': doc.company})
    je.append('accounts', {
        'account': debit_acc,
        'debit_in_account_currency': amt,
        'cost_center':doc.cost_center,
        'project':doc.project
        })
    je.append('accounts', {
        'account': credit_acc,
        'credit_in_account_currency': amt,
        'cost_center':doc.cost_center,
        'project':doc.project
    })
    je.user_remark = doc.name
    je.custom_purchase_receipt = doc.name
    je.save(ignore_permissions=True)
    je.submit()


@frappe.whitelist()
def cancel_journal_entry_for_purchase(doc, method):
    je_name = frappe.db.get_value('Journal Entry', {'custom_purchase_receipt': doc.name, 'docstatus': 1})
    if je_name:
        rev_je_name = frappe.db.get_value('Journal Entry', {'reversal_of': je_name, 'docstatus': 1})
        if rev_je_name:
            frappe.get_doc('Journal Entry', {'reversal_of': je_name, 'docstatus': 1})
            if rev_je_name.docstatus == 1:
                rev_je_name.cancel()
        je = frappe.get_doc('Journal Entry', je_name)
        je.cancel()
        frappe.db.commit()

@frappe.whitelist()
def create_reversal_entry_for_purchase(doc,method):
    pr=''
    credit_amt=0
    debit_amt=0
    for i in doc.items:
        if i.purchase_receipt and i.item_type =='Service Item':
            pr=i.purchase_receipt
            rate = frappe.db.get_value("Purchase Order Item", {'name': i.po_detail}, "base_rate")
            if rate:
                credit_amt += (rate*i.qty) or 0
                debit_amt += (rate*i.qty) or 0
            else:
                credit_amt += 0
                debit_amt += 0
    if credit_amt <= 0:
        return
    if debit_amt <= 0:
        return
    if frappe.db.exists('Journal Entry',{'user_remark':pr,'docstatus':1}):
        je=frappe.get_doc('Journal Entry',{'user_remark':pr,'docstatus':1})
        rev_entry = frappe.new_doc("Journal Entry")
        rev_entry.company = doc.company
        rev_entry.posting_date = doc.posting_date
        rev_entry.voucher_type = "Journal Entry"
        debit_acc = frappe.db.get_value(
            "Account",
            {'name': ('like', '%Unbilled Revenue%'), 'company': doc.company}
        )
        credit_acc = frappe.db.get_value("Account", {'name': ('like', '%Service Received But Not Billed%'), 'company': doc.company})
        rev_entry.append('accounts', {
            'account': debit_acc,
            'credit_in_account_currency': credit_amt,
            # 'party_type':'Supplier',
            # 'party':doc.supplier,
            'reference_type':'Journal Entry',
            'reference_name':je.name,
            'cost_center':doc.cost_center,
            'project':doc.project
        })
        rev_entry.append('accounts', {
            'account': credit_acc,
            # 'party_type':'Supplier',
            # 'party':doc.supplier,
            'debit_in_account_currency': debit_amt,
            'reference_type':'Journal Entry',
            'reference_name':je.name,
            'cost_center':doc.cost_center,
            'project':doc.project
        })
        rev_entry.user_remark = doc.name
        rev_entry.reversal_of = je.name
        rev_entry.cost_center=doc.cost_center
        rev_entry.project = doc.project
        rev_entry.custom_references=str(je.name)+','+str(doc.name)
        rev_entry.save(ignore_permissions=True)
        rev_entry.submit()

@frappe.whitelist()
def cancel_reverse_entry_for_purchase(doc, method):
    je_name = frappe.db.get_value('Journal Entry', {'user_remark': doc.name, 'docstatus': 1})
    if je_name:
        je = frappe.get_doc('Journal Entry', je_name)
        if je.docstatus == 1:
            je.cancel()
        frappe.db.commit()

@frappe.whitelist()
def update_dn_doc():
    # doc=frappe.get_doc('GL Entry','ACC-GLE-2025-00371')
    # doc=frappe.get_doc('Payment Ledger Entry','uer9k06oca')
    att=frappe.db.get_all('Salary Certificate',{'reference':'HR-EMP-00007'},['name'])
    for a in att:
        doc=frappe.get_doc('Salary Certificate',a.name)
        doc.delete()

from frappe.model.naming import make_autoname
@frappe.whitelist()
def set_item_code_if_missing(doc, method):
    if not doc.item_code:
        doc.item_code = make_autoname("SKU.####")
        doc.name = doc.item_code

from datetime import date
@frappe.whitelist()
def check_leave_application_creation():
    leave_applications = frappe.get_all("Leave Application",{'workflow_state':'Pending for Reporting Manager','custom_created_before_1_day':0},['name'])
    leave_applications_2 = frappe.get_all("Leave Application",{'workflow_state':'Pending for Reporting Manager','custom_created_before_2_day':0},['name'])
    for la in leave_applications:
        leave_app = frappe.get_doc("Leave Application",la.name)
        if leave_app.creation.date() < date.today():
            leave_app.custom_created_before_1_day = 1
            leave_app.save(ignore_permissions=True)

    for la in leave_applications_2:
        leave_app_2 = frappe.get_doc("Leave Application", la.name)

        if leave_app_2.creation.date() < (date.today() - timedelta(days=1)):
            leave_app_2.custom_created_before_2_day = 1
            leave_app_2.save(ignore_permissions=True)



@frappe.whitelist()
def update_specification(doc, method):
    updated_items = {}

    for it in doc.items:
        if not it.item_code or not it.custom_specifications:
            continue

        if it.item_code in updated_items:
            continue

        item = frappe.get_doc("Item", it.item_code)
        if item.custom_specification != it.custom_specifications:
            item.custom_specification = it.custom_specifications
            item.save(ignore_permissions=True)
            updated_items[it.item_code] = True

@frappe.whitelist()
def get_item_spec_query(doctype, txt, searchfield, start, page_len, filters):
    item_code = filters.get("item_code")

    if not item_code:
        return []

    return frappe.db.sql("""
        SELECT name,specification
        FROM `tabItem Specification`
        WHERE item_code = %s
        LIMIT %s OFFSET %s
    """, (item_code, page_len, start))


# @frappe.whitelist()
# def po_so():
#     print("MuthuSElvan E")
#     purchase_orders = frappe.get_all("Purchase Order", fields=["name"])

#     for po in purchase_orders:
#         # Assuming 'Sales Order' links to PO via a custom link field like 'purchase_order'
#         sales_orders = frappe.get_all(
#             "Sales Order",
#             filters={"purchase_order": po.name},  # adjust field name as needed
#             fields=["name"]
#         )
#         for so in sales_orders:
#             print(so.name)
@frappe.whitelist()
def po_so():

    # Get all Purchase Orders
    purchase_orders = frappe.get_all("Purchase Order", fields=["name"])

    for po in purchase_orders:
        # Get related Purchase Order Items
        po_items = frappe.get_all(
            "Purchase Order Item",
            filters={"parent": po.name},
            fields=["sales_order"]
        )

        for item in po_items:
            if item.sales_order:
                print(item.sales_order)

@frappe.whitelist()
def update_dl_records():
    frappe.delete_doc("Payment Ledger Entry","nav1fbrtkr")
    # frappe.delete_doc("Stock Ledger Entry","MAT-SLE-2025-00026")
    # frappe.delete_doc("Payment Ledger Entry", "")
    # frappe.delete_doc("GL Entry","ACC-GLE-2025-00570")
    # doc = frappe.get_doc("Purchase Order", "PO-PDO-2025-00032")
    # if doc.docstatus == 1:

    #     doc.cancel()
    # frappe.delete_doc("Purchase Order", "PO-PDO-2025-00032")


# import frappe

# @frappe.whitelist()
# def delete_fixed_material_request():
#     mr_name = "MRN-PDO-2025-00017"  # Hardcoded Material Request name
#     print(f"Starting deletion for: {mr_name}")

#     # Step 1: Delete linked Purchase Orders
#     purchase_orders = frappe.get_all(
#         "Purchase Order",
#         filters={"material_request": mr_name},
#         fields=["name"]
#     )

#     for po in purchase_orders:
#         try:
#             frappe.delete_doc("Purchase Order", po.name, force=True)
#             print(f"Deleted Purchase Order: {po.name}")
#         except Exception as e:
#             print(f"Failed to delete PO {po.name}: {str(e)}")

#     # Step 2: Delete the Material Request
#     try:
#         frappe.delete_doc("Material Request", mr_name, force=True)
#         print(f"Deleted Material Request: {mr_name}")
#     except Exception as e:
#         print(f"Failed to delete Material Request {mr_name}: {str(e)}")


# @frappe.whitelist()
# def update_can():
#     frappe.db.sql("""UPDATE `tabAttendance and OT Register` SET docstatus = 0 where name = 'ATT-OTR-00001' """)



@frappe.whitelist()
def merge_account_test():
    # frappe.rename_doc("Account", "1215 - Bank Muscat Smart Card A/c-2 - SUTC", "200104 - Bank Muscat Smart Card A/c-2 - SUTC", force=1)
    # frappe.rename_doc("Account", "1216 - Bank Muscat A/c -297 - SUTC", "200105 - Bank Muscat A/c -297 - SUTC", force=1)
    frappe.rename_doc("Account", "701700 - Direct Salary Cost - SUTC", "701000 - Direct Cost - SUTC", force=1)
    # frappe.rename_doc("Account", "1212 - Guarantee Margins - SUTC", "200202 - Guarantee Margins - SUTC", force=1)
    # frappe.rename_doc("Account", "1210 - Oman Arab Bank - SUTC", "200101 - Oman Arab Bank - SUTC", force=1)

@frappe.whitelist()
def expense_account_table():
    # html = """<p>Please select any one of the following accounts as Expense Account based on the Item type</p>"""
    html = """ """
    html += """
        <style>j
            td {
                padding-right: 8px;
                padding-left: 8px;
                padding-top: 4px;
                padding-bottom: 4px;
            }
        </style>
            <div style="overflow-x: auto; color: white;">
                <table class="text-center" style="overflow: hidden; white-space: nowrap; margin-bottom: 50px;">
                    <tr style="background-color: #e8262e; font-weight: 500; color: white;">
                        <td class="border border-1 border-light" style="min-width: 50px;">S.No.</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Item Type</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Account Number</td>
                        <td class="border border-1 border-light" style="min-width: 100px;">Account Name</td>
                    </tr>
    """
    html += """
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">1</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Material Cost</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701001</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Material, Stores & Spares Cost</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">2</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">All Subcontract Jobs</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701002</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Sub Contract Expenses</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">3</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Medical Bills</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701713</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Staff Welfare Expenses</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">4</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">JCB, Crane, Generator repair and maintenance</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701401</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Equipment Repair & Maintenance</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">5</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Other repairs</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701402</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Other Repair & Maintenance</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">6</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">All fuel related expenses</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701403</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Vehicle Running Expenses</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">7</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Vehicle rent</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701404</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Vehicle Rent Expenses</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">8</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Vehicle repair</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701405</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Vehicle Repair & Maintenance</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">9</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Equipment rental</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701406</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Equipment Rent Expenses</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">10</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Parcel, Freight, Transport related expenses</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701407</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Freight and Transport</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">11</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Stationery Expenses</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701603</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Printing & Stationery Expenses</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">11</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Taxi / Bus fare</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701604</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Travelling & Conveyance Expense</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">12</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Legal fees</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701605</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Legal & Professional fees</td>
        </tr>
        <tr style="color: black;">
            <td class="border border-1 border-danger" style="min-width: 50px; text-align: center;">13</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Training & Course</td>
            <td class="border border-1 border-danger" style="min-width: 80px; text-align: center;">701606</td>
            <td class="border border-1 border-danger" style="min-width: 100px; text-align: left;">Direct Training & Course Expenses</td>
        </tr>

    """
    return html


@frappe.whitelist()
def expense_account_table_new(company=None):
    # fetch expense accounts based on the company
    parent_accounts = frappe.get_all(
    "Account",
    filters={
        "company": company,
        "account_number": ["in", ["701000"]],
        "is_group": 1
    },
    pluck="name"
    )

    # if parent_accounts:
    # 	child_accounts = frappe.get_all(
    # 		"Account",
    # 		filters={
    # 			"company": company,
    # 			"parent_account": ["in", parent_accounts]
    # 		},
    # 		pluck="name"
    # 	)

    # 	accounts = frappe.get_all(
    # 		"Account",
    # 		filters={
    # 			"company": company,
    # 			"is_group": 0,
    # 			"parent_account": ["in", child_accounts]
    # 		},
    # 		fields=["name", "account_number"]
    # 	)
        
    if parent_accounts:
        child_accounts = frappe.get_all(
            "Account",
            filters={
                "company": company,
                "parent_account": ["in", parent_accounts]
            },
            pluck="name"
        )

        # Find child accounts that are NOT parents themselves
        non_parent_children = []
        for acc in child_accounts:
            has_sub = frappe.db.exists("Account", {"company": company, "parent_account": acc})
            if not has_sub:
                non_parent_children.append(acc)

        # Get non-group accounts under those child accounts
        accounts = frappe.get_all(
            "Account",
            filters={
                "company": company,
                "is_group": 0,
                "parent_account": ["in", child_accounts]
            },
            fields=["name", "account_number"]
        )

        # Also include those child accounts which are not parents
        if non_parent_children:
            extra_accounts = frappe.get_all(
                "Account",
                filters={
                    "company": company,
                    "name": ["in", non_parent_children],
                    "is_group": 0
                },
                fields=["name", "account_number"]
            )
            accounts.extend(extra_accounts)



        # Start building HTML table
        html = """
            <style>
                td {
                    padding: 6px 10px;
                }
                table {
                    border-collapse: collapse;
                    width: 100%;
                }
            </style>
            <div style="overflow-x:auto;">
            <table border="1">
                <tr style="background-color:#e8262e; color:white; font-weight:bold;">
                    <td>S.No.</td>
                    <td>Account Number</td>
                    <td>Account Name</td>
                </tr>
        """

        for i, acc in enumerate(accounts, start=1):
            html += f"""
                <tr style="color:black;">
                    <td style='text-align:center'>{i}</td>
                    <td>{acc.get('account_number') or ''}</td>
                    <td>{acc.get('name') or ''}</td>
                </tr>
            """

        html += "</table></div>"

    if not parent_accounts:
        html = f"<p style='color:red;'>No expense accounts found for {company}</p>"

    return html


@frappe.whitelist()
def update_approved_by():
    frappe.db.set_value("Purchase Order",'PO-OETC-2025-00003','custom_approved_by','ho@shaherunited.com')
    frappe.db.set_value("Purchase Order",'PO-OETC-2025-00003','custom_authorized_by','dinesh@shaherunited.com')

@frappe.whitelist()
def create_leave_salary():
    employees = frappe.db.sql("""
        SELECT
            name,
            date_of_joining,
            custom_last_leave_salary_date, custom_basic, custom_nationality_type
        FROM
            `tabEmployee`
        WHERE
            status = 'Active'
            AND custom_annual_leave_applicable_after IS NOT NULL
            AND custom_annual_leave_applicable_after != ''
            AND (
                (
                    custom_last_leave_salary_date IS NOT NULL
                    AND DATE_ADD(custom_last_leave_salary_date, INTERVAL custom_annual_leave_applicable_after DAY) = CURDATE()
                )
                OR (
                    custom_last_leave_salary_date IS NULL
                    AND DATE_ADD(date_of_joining, INTERVAL custom_annual_leave_applicable_after DAY) = CURDATE()
                )
            )
    """, as_dict=True)
    for emp in employees:
        start_date = emp.custom_last_leave_salary_date or emp.date_of_joining

        no_of_days = frappe.db.sql(
            """
            SELECT SUM(no_of_days_worked)
            FROM `tabAttendance and OT Register`
            WHERE (from_date >= %s AND to_date <= %s)
            AND employee = %s
            AND docstatus = 1
            """,
            (start_date, today(), emp.name),
            as_dict=False
        )

        salary_days = no_of_days[0][0] if no_of_days and no_of_days[0][0] else 0
        leave_salary_value = emp.custom_basic * salary_days /31

        leave_sal = frappe.new_doc("Leave Salary")
        leave_sal.purpose = 'Leave Salary'
        leave_sal.employee_number = emp.name
        leave_sal.leave_salary = leave_salary_value
        leave_sal.last_leave_salary_taken_day = emp.custom_last_leave_salary_date or ''
        leave_sal.leave_salary_days_salary = salary_days
        leave_sal.posting_date = today()
        if emp.custom_nationality_type == 'Expat':
            leave_sal.type = 'Expat'
        else:
            leave_sal.type = 'Omani'
        leave_sal.save(ignore_permissions=True)
        frappe.db.set_value("Employee",emp.name,'custom_last_leave_salary_date',today())
        print(leave_salary_value)

@frappe.whitelist()
def create_annual_leave_allocation():
    employees = frappe.db.sql("""
        SELECT
            name,
            date_of_joining,
            custom_rejoining_date,
            custom_annual_leave_applicable_after
        FROM
            `tabEmployee`
        WHERE
            status = 'Active'
            AND custom_annual_leave_applicable_after IS NOT NULL
            AND custom_annual_leave_applicable_after != ''
            AND (
                (
                    custom_rejoining_date IS NOT NULL
                    AND DATE_ADD(custom_rejoining_date, INTERVAL 1 MONTH) <= CURDATE()
                )
                OR (
                    custom_rejoining_date IS NULL
                    AND DATE_ADD(date_of_joining, INTERVAL 1 MONTH) <= CURDATE()
                )
            )
    """, as_dict=True)
    for emp in employees:
        print(emp)
        days_after = int(emp.custom_annual_leave_applicable_after)

        base_date = emp.custom_rejoining_date or emp.date_of_joining
        from_date = add_days(base_date, days_after)

        if days_after == 365:
            allocated_days = 2.5
        elif days_after in [180, 90]:
            allocated_days = 5
        else:
            continue

        # Optional: avoid duplicates
        if frappe.db.exists("Leave Allocation", {
            "employee": emp.name,
            "leave_type": "Annual Leave",
            'docstatus':('!=',2)
        }):
            leave_allocation = frappe.get_doc("Leave Allocation",{
                "employee": emp.name,
                "leave_type": "Annual Leave",
                'docstatus':('!=',2)
            })
            leave_allocation.new_leaves_allocated = allocated_days
            leave_allocation.save()
            leave_allocation.submit()
        else:
            leave_allocation = frappe.new_doc("Leave Allocation")
            leave_allocation.employee = emp.name
            leave_allocation.leave_type = "Annual Leave"
            leave_allocation.from_date = from_date
            leave_allocation.to_date = "31-12-2100"
            leave_allocation.new_leaves_allocated = allocated_days
            leave_allocation.save()
            leave_allocation.submit()







@frappe.whitelist()
def project_site_set_query(doctype, txt, searchfield, start, page_len, filters):
    project = filters.get("project") if filters else None
    if not project:
        return []


    sites = frappe.db.get_all(
        "Project Sites",
        filters={"parent": project},
        pluck="site"
    )


    return [(s,) for s in sites]



@frappe.whitelist()
def send_expiry_alerts():
    send_hse_training_alerts()
    send_ftw_register_alerts()
    send_vehicle_expiry_alerts()


@frappe.whitelist()
def create_scheduled_job_send_expiry_alerts():
    pos = frappe.db.exists('Scheduled Job Type', 'send_expiry_alerts')
    if not pos:
        sjt = frappe.new_doc("Scheduled Job Type")
        sjt.update({
            "method": 'shaher.custom.send_expiry_alerts',
            "frequency": 'Daily'

        })
        sjt.save(ignore_permissions=True)

@frappe.whitelist()
def send_hse_training_alerts():
    today = getdate(nowdate())
    locations = frappe.get_all("Site", filters={'user': ["is","set"]}, fields=["name","user"],order_by="name")
    for site in locations:
        alert_rows_1 = []
        alert_rows_2 = []
        alert_rows_3 = []
        employees = frappe.get_all("Employee", filters={'status': 'Active',"custom_site_location":site.name}, fields=["name", "employee_name"],order_by="name")

        for emp in employees:
            doc = frappe.get_doc("Employee", emp.name)

            for row in doc.get("custom_hse_training_matrix") or []:
                if not row.validity:
                    continue
                # if not row.expiry_date:
                #     continue

                expiry_date = getdate(row.expiry_date)
                days_diff = (expiry_date - today).days

                if days_diff in (30, 15):
                    alert_rows_1.append({
                        "ID": doc.name,
                        "employee": doc.employee_name,
                        "company":doc.company,
                        "course": row.course_name,
                        "course group": row.course_group,
                        "expiry": formatdate(expiry_date),
                        "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                    })
                elif days_diff in (15, 1):
                    alert_rows_2.append({
                        "ID": doc.name,
                        "employee": doc.employee_name,
                        "company":doc.company,
                        "course": row.course_name,
                        "course group": row.course_group,
                        "expiry": formatdate(expiry_date),
                        "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                    })
                elif days_diff < 0:
                    alert_rows_3.append({
                        "ID": doc.name,
                        "employee": doc.employee_name,
                        "company":doc.company,
                        "course": row.course_name,
                        "course group": row.course_group,
                        "expiry": formatdate(expiry_date),
                        "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                    })

        if alert_rows_1:
            html = """
                <p>Dear Sir/Madam,</p>
                <p>
                Please find below the training records that will expire within the next 30 days. Kindly plan for timely renewal.
                </p>
                <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
                    <tr>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee ID</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee Name</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Company</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Course</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Course Group</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Expiry Date</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Status</th>
                    </tr>
                """
            for r in alert_rows_1:
                html += f"""
                <tr>
                    <td style="font-size:12px;">{r.get('ID', '')}</td>
                    <td style="font-size:12px;">{r.get('employee', '')}</td>
                    <td style="font-size:12px;">{r.get('company', '')}</td>
                    <td style="font-size:12px;">{r.get('course', '')}</td>
                    <td style="font-size:12px;">{r.get('course group', '')}</td>
                    <td style="font-size:12px;">{r.get('expiry', '')}</td>
                    <td style="font-size:12px;">{r.get('status', '')}</td>
                </tr>
                """
            html += "</table>"

            frappe.sendmail(
                recipients=[site.user],
                subject="Upcoming HSE Training Renewal – Action Required (30 Days)",
                message=html
            )

        if alert_rows_2:
            html = """
                <p>Dear Sir/Madam,</p>
                <p>
                Please find below the training records that will expire within the next 15 days. Kindly plan for timely renewal.
                </p>
                <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
                    <tr>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee ID</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee Name</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Company</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Course</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Course Group</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Expiry Date</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Status</th>
                    </tr>
                """
            for r in alert_rows_2:
                html += f"""
                <tr>
                    <td style="font-size:12px;">{r.get('ID', '')}</td>
                    <td style="font-size:12px;">{r.get('employee', '')}</td>
                    <td style="font-size:12px;">{r.get('company', '')}</td>
                    <td style="font-size:12px;">{r.get('course', '')}</td>
                    <td style="font-size:12px;">{r.get('course group', '')}</td>
                    <td style="font-size:12px;">{r.get('expiry', '')}</td>
                    <td style="font-size:12px;">{r.get('status', '')}</td>
                </tr>
                """
            html += "</table>"

            frappe.sendmail(
                recipients=[site.user],
                subject="Urgent Reminder: HSE Training Expiring in 15 Days",
                message=html
            )
        if alert_rows_3:
            html = """
                <p>Dear Sir/Madam,</p>
                <p>
                Please find below the training records that have already expired and require urgent renewal to maintain compliance.s
                </p>
                <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
                    <tr>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee ID</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee Name</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Company</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Course</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Course Group</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Expiry Date</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Status</th>
                    </tr>
                """
            for r in alert_rows_3:
                html += f"""
                <tr>
                    <td style="font-size:12px;">{r.get('ID', '')}</td>
                    <td style="font-size:12px;">{r.get('employee', '')}</td>
                    <td style="font-size:12px;">{r.get('company', '')}</td>
                    <td style="font-size:12px;">{r.get('course', '')}</td>
                    <td style="font-size:12px;">{r.get('course group', '')}</td>
                    <td style="font-size:12px;">{r.get('expiry', '')}</td>
                    <td style="font-size:12px;">{r.get('status', '')}</td>
                </tr>
                """
            html += "</table>"

            frappe.sendmail(
                recipients=[site.user],
                subject="Critical Alert: HSE Training Already Expired – Immediate Action Required",
                message=html
            )

@frappe.whitelist()
def send_ftw_register_alerts():
    today = getdate(nowdate())
    locations = frappe.get_all("Site", filters={'user': ["is","set"]}, fields=["name","user"],order_by="name")
    for site in locations:
        alert_rows_1 = []
        alert_rows_2 = []
        alert_rows_3 = []
        employees = frappe.get_all("Employee", filters={'status': 'Active',"custom_site_location":site.name}, fields=["name", "employee_name"],order_by="name")

        for emp in employees:
            doc = frappe.get_doc("Employee", emp.name)

            for row in doc.get("custom_ftw_register") or []:
                if not row.next_due_date:
                    continue

                test_date = getdate(row.medical_test_done_on) if row.medical_test_done_on else ""
                next_due = getdate(row.next_due_date)
                days_diff = (next_due - today).days

                if days_diff in (30, 15) or days_diff < 0:
                    alert_rows_1.append({
                        "ID": doc.name,
                        "employee": doc.employee_name,
                        "company":doc.company,
                        "medical test done on": formatdate(test_date) if test_date else "",
                        "expiry": formatdate(next_due),
                        "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                    })
                elif days_diff in (15, 1) or days_diff < 0:
                    alert_rows_2.append({
                        "ID": doc.name,
                        "employee": doc.employee_name,
                        "company":doc.company,
                        "medical test done on": formatdate(test_date) if test_date else "",
                        "expiry": formatdate(next_due),
                        "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                    })
                elif days_diff < 0:
                    alert_rows_3.append({
                        "ID": doc.name,
                        "employee": doc.employee_name,
                        "company":doc.company,
                        "medical test done on": formatdate(test_date) if test_date else "",
                        "expiry": formatdate(next_due),
                        "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                    })

        if alert_rows_1:
            html = """
                <p>Dear Sir/Madam,</p>
                <p>
                Please find below the FTW register entries that will <b>expire within the next 30 days</b>. Kindly plan the necessary updates to ensure timely compliance.
                </p>
                <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
                    <tr>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee ID</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee Name</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Company</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Medical Test Done On</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Expiry Date</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Status</th>
                    </tr>
                """

            for r in alert_rows_1:
                html += f"""
                <tr>
                    <td style="font-size:12px;">{r.get('ID', '')}</td>
                    <td style="font-size:12px;">{r.get('employee', '')}</td>
                    <td style="font-size:12px;">{r.get('company', '')}</td>
                    <td style="font-size:12px;">{r.get('medical test done on', '')}</td>
                    <td style="font-size:12px;">{r.get('expiry', '')}</td>
                    <td style="font-size:12px;">{r.get('status', '')}</td>
                </tr>
                """
            html += "</table>"

            frappe.sendmail(
                recipients=[site.user],
                subject="Upcoming Employee FTW Register Renewal – Action Required (30 Days)",
                message=html
            )
        if alert_rows_2:
            html = """
                <p>Dear Sir/Madam,</p>
                <p>
                The following FTW register entries will <b>expire within the next 15 days</b>.  Immediate action is recommended to avoid delays or non-compliance.
                </p>
                <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
                    <tr>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee ID</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee Name</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Company</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Medical Test Done On</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Expiry Date</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Status</th>
                    </tr>
                """

            for r in alert_rows_2:
                html += f"""
                <tr>
                    <td style="font-size:12px;">{r.get('ID', '')}</td>
                    <td style="font-size:12px;">{r.get('employee', '')}</td>
                    <td style="font-size:12px;">{r.get('company', '')}</td>
                    <td style="font-size:12px;">{r.get('medical test done on', '')}</td>
                    <td style="font-size:12px;">{r.get('expiry', '')}</td>
                    <td style="font-size:12px;">{r.get('status', '')}</td>
                </tr>
                """
            html += "</table>"

            frappe.sendmail(
                recipients=[site.user],
                subject="Urgent Reminder: Employee FTW Register Renewal Due in 15 Days",
                message=html
            )
        if alert_rows_3:
            html = """
                <p>Dear Sir/Madam,</p>
                <p>
                Please note that the following FTW register entries have <b>already expired</b>.  These require urgent updating to ensure records remain valid and compliant.
                </p>
                <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
                    <tr>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee ID</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Employee Name</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Company</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Medical Test Done On</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Expiry Date</th>
                        <th style="text-align:center; background-color:#b81a0f; color:#ffffff; padding:8px; font-size:12px;">Status</th>
                    </tr>
                """

            for r in alert_rows_3:
                html += f"""
                <tr>
                    <td style="font-size:12px;">{r.get('ID', '')}</td>
                    <td style="font-size:12px;">{r.get('employee', '')}</td>
                    <td style="font-size:12px;">{r.get('company', '')}</td>
                    <td style="font-size:12px;">{r.get('medical test done on', '')}</td>
                    <td style="font-size:12px;">{r.get('expiry', '')}</td>
                    <td style="font-size:12px;">{r.get('status', '')}</td>
                </tr>
                """
            html += "</table>"

            frappe.sendmail(
                recipients=[site.user],
                subject="Critical Alert: Employee FTW Register Already Expired – Immediate Action Required",
                message=html
            )

@frappe.whitelist()
def send_email_hse(rows):
    html = """
    <p>Dear Sir/Madam,</p>
    <p>
    Please find below the training records that require your attention:
    <ul>
        <li><b>Expiring in the next 30 days</b> - Kindly plan for timely renewal to avoid any lapse.</li>
        <li><b>Expiring in the next 15 days</b> - Immediate action is recommended to complete the necessary training.</li>
        <li><b>Already expired</b> - These require urgent renewal to remain compliant.</li>
    </ul>
    </p>
    <h3>Employee HSE Training Expiry Alert</h3>
    <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
        <tr>
            <th>Employee ID</th>
            <th>Employee Name</th>
            <th>Company</th>
            <th>Course</th>
            <th>Course Group</th>
            <th>Expiry Date</th>
            <th>Status</th>
        </tr>
    """
    for r in rows:
        html += f"""
        <tr>
            <td>{r.get('ID', '')}</td>
            <td>{r.get('employee', '')}</td>
            <td>{r.get('company', '')}</td>
            <td>{r.get('course', '')}</td>
            <td>{r.get('course group', '')}</td>
            <td>{r.get('expiry', '')}</td>
            <td>{r.get('status', '')}</td>
        </tr>
        """
    html += "</table>"

    frappe.sendmail(
        recipients=["jothi.m@groupteampro.com","abdulla.pi@groupteampro.com","chitarasu@shaherunited.com"],
        subject="Employee HSE Training Expiry Alert",
        message=html
    )

@frappe.whitelist()
def send_email_ftw(rows):
    html = """
    <p>Dear Sir/Madam,</p>
    <p>
    Please find below the register entries that require your attention:
    <ul>
        <li><b>Expiring in the next 30 days</b> - Kindly initiate the update process to maintain compliance.</li>
        <li><b>Expiring in the next 15 days</b> - Immediate update is recommended to avoid delays.</li>
        <li><b>Already expired</b> - These require urgent updating to ensure records remain valid.</li>
    </ul>
    </p>
    <h3>Employee FTW Register Expiry Alert</h3>
    <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
        <tr>
            <th>Employee ID</th>
            <th>Employee Name</th>
            <th>Company</th>
            <th>Medical Test Done On</th>
            <th>Expiry Date</th>
            <th>Status</th>
        </tr>
    """
    for r in rows:
        html += f"""
        <tr>
            <td>{r.get('ID', '')}</td>
            <td>{r.get('employee', '')}</td>
            <td>{r.get('company', '')}</td>
            <td>{r.get('medical test done on', '')}</td>
            <td>{r.get('expiry', '')}</td>
            <td>{r.get('status', '')}</td>
        </tr>
        """
    html += "</table>"

    frappe.sendmail(
        recipients=["riyaz.a@groupteampro.com"],
        subject="Employee FTW Register Expiry Alert",
        message=html
    )





@frappe.whitelist()
def send_vehicle_expiry_alerts():
    today = getdate(nowdate())

    vehicles = frappe.get_all("Vehicle", fields=["name"])

    expiry_fields = [
        ("custom_rop_due", "ROP Due Date"),
        ("custom_ras_due", "RAS Due Date"),
        ("custom_ivms_certificate_validity", "IVMS Certificate Validity"),
        ("custom_voc_validity", "VOC Validity"),
        ("custom_lifting_tpi_validity", "Lifting TPI Validity"),
        ("custom_speed_limiter__roll_over_cage", "Speed Limiter & Roll Over Cage")
    ]

    for field, label in expiry_fields:
        alert_rows = []

        for v in vehicles:
            doc = frappe.get_doc("Vehicle", v.name)
            due_date = getattr(doc, field, None)
            if not due_date:
                continue

            due_date = getdate(due_date)
            days_diff = (due_date - today).days

            if days_diff in (30, 15) or days_diff < 0:
                alert_rows.append({
                    "vehicle": doc.name,
                    "type": label,
                    "company":doc.custom_company,
                    "expiry": formatdate(due_date),
                    "status": "Expired" if days_diff < 0 else f"Due in {days_diff} days"
                })

        if alert_rows:
            send_vehicle_email(alert_rows, label)


def send_vehicle_email(rows, label):
    html = f"""
    <p>Dear Sir/Madam,</p>
    <p>
    Please find below the list of vehicles with <b>{label}</b> requiring attention:
    <ul>
        <li><b>Expiring in the next 30 days</b> - Plan renewal to avoid non-compliance.</li>
        <li><b>Expiring in the next 15 days</b> - Immediate renewal recommended.</li>
        <li><b>Already expired</b> - Urgent renewal required.</li>
    </ul>
    </p>
    <h3>Vehicle {label} Expiry Alert</h3>
    <table border=1 cellpadding=5 cellspacing=0 style="border:1px solid black; border-collapse:collapse;">
        <tr>
            <th>Vehicle</th>
            <th>Type</th>
            <th>Company</th>
            <th>Expiry Date</th>
            <th>Status</th>
        </tr>
    """

    for r in rows:
        html += f"""
        <tr>
            <td>{r.get('vehicle')}</td>
            <td>{r.get('type')}</td>
            <td>{r.get('company')}</td>
            <td>{r.get('expiry')}</td>
            <td>{r.get('status')}</td>
        </tr>
        """

    html += "</table>"

    frappe.sendmail(
        recipients=["riyaz.a@groupteampro.com"],
        subject=f"Vehicle {label} Expiry Alert",
        message=html
    )


@frappe.whitelist()
def make_delivery_note(source_name, target_doc=None, kwargs=None):
    from erpnext.stock.doctype.packed_item.packed_item import make_packing_list
    from frappe.contacts.doctype.address.address import get_company_address
    from erpnext.setup.doctype.item_group.item_group import get_item_group_defaults
    from erpnext.stock.doctype.item.item import get_item_defaults
    from frappe.model.utils import get_fetch_values
    from erpnext.stock.doctype.stock_reservation_entry.stock_reservation_entry import (
        get_sre_details_for_voucher,
        get_sre_reserved_qty_details_for_voucher,
        get_ssb_bundle_for_voucher,
    )

    if not kwargs:
        kwargs = {
            "for_reserved_stock": frappe.flags.args and frappe.flags.args.for_reserved_stock,
            "skip_item_mapping": frappe.flags.args and frappe.flags.args.skip_item_mapping,
        }

    kwargs = frappe._dict(kwargs)
    so_detail_list = []
    if frappe.flags.args and frappe.flags.args.get("so_detail_list"):
        so_detail_list = frappe.flags.args.get("so_detail_list")
        if isinstance(so_detail_list, str):
            so_detail_list = frappe.parse_json(so_detail_list)
    sre_details = {}
    if kwargs.for_reserved_stock:
        sre_details = get_sre_reserved_qty_details_for_voucher("Sales Order", source_name)

    mapper = {
        "Sales Order": {"doctype": "Delivery Note", "validation": {"docstatus": ["=", 1]}},
        "Sales Taxes and Charges": {"doctype": "Sales Taxes and Charges", "reset_value": True},
        "Sales Team": {"doctype": "Sales Team", "add_if_empty": True},
    }

    def set_missing_values(source, target):
        if kwargs.get("ignore_pricing_rule"):
            # Skip pricing rule when the dn is creating from the pick list
            target.ignore_pricing_rule = 1

        target.run_method("set_missing_values")
        target.run_method("set_po_nos")
        target.run_method("calculate_taxes_and_totals")
        target.run_method("set_use_serial_batch_fields")

        if source.company_address:
            target.update({"company_address": source.company_address})
        else:
            # set company address
            target.update(get_company_address(target.company))

        if target.company_address:
            target.update(get_fetch_values("Delivery Note", "company_address", target.company_address))

        # if invoked in bulk creation, validations are ignored and thus this method is nerver invoked
        if frappe.flags.bulk_transaction:
            # set target items names to ensure proper linking with packed_items
            target.set_new_name()

        make_packing_list(target)

    def condition(doc):
        if doc.name in sre_details:
            del sre_details[doc.name]
            return False

        # Filter based on so_detail_list
        if so_detail_list:
            if doc.name not in so_detail_list:
                return False

        # Only include the clicked line item if so_detail is passed
        if frappe.flags.args and frappe.flags.args.get("so_detail"):
            if doc.name != frappe.flags.args.so_detail:
                return False

        if frappe.flags.args and frappe.flags.args.delivery_dates:
            if cstr(doc.delivery_date) not in frappe.flags.args.delivery_dates:
                return False

        return abs(doc.delivered_qty) < abs(doc.qty) and doc.delivered_by_supplier != 1

    def update_item(source, target, source_parent):
        target.base_amount = (flt(source.qty) - flt(source.delivered_qty)) * flt(source.base_rate)
        target.amount = (flt(source.qty) - flt(source.delivered_qty)) * flt(source.rate)
        target.qty = flt(source.qty) - flt(source.delivered_qty)

        item = get_item_defaults(target.item_code, source_parent.company)
        item_group = get_item_group_defaults(target.item_code, source_parent.company)

        if item:
            target.cost_center = (
                frappe.db.get_value("Project", source_parent.project, "cost_center")
                or item.get("buying_cost_center")
                or item_group.get("buying_cost_center")
            )

    if not kwargs.skip_item_mapping:
        mapper["Sales Order Item"] = {
            "doctype": "Delivery Note Item",
            "field_map": {
                "rate": "rate",
                "name": "so_detail",
                "parent": "against_sales_order",
            },
            "condition": condition,
            "postprocess": update_item,
        }

    so = frappe.get_doc("Sales Order", source_name)
    target_doc = get_mapped_doc("Sales Order", so.name, mapper, target_doc)

    if not kwargs.skip_item_mapping and kwargs.for_reserved_stock:
        sre_list = get_sre_details_for_voucher("Sales Order", source_name)

        if sre_list:

            def update_dn_item(source, target, source_parent):
                update_item(source, target, so)

            so_items = {d.name: d for d in so.items if d.stock_reserved_qty}

            for sre in sre_list:
                if not condition(so_items[sre.voucher_detail_no]):
                    continue

                dn_item = get_mapped_doc(
                    "Sales Order Item",
                    sre.voucher_detail_no,
                    {
                        "Sales Order Item": {
                            "doctype": "Delivery Note Item",
                            "field_map": {
                                "rate": "rate",
                                "name": "so_detail",
                                "parent": "against_sales_order",
                            },
                            "postprocess": update_dn_item,
                        }
                    },
                    ignore_permissions=True,
                )

                dn_item.qty = flt(sre.reserved_qty) / flt(dn_item.get("conversion_factor", 1))

                if sre.reservation_based_on == "Serial and Batch" and (sre.has_serial_no or sre.has_batch_no):
                    dn_item.serial_and_batch_bundle = get_ssb_bundle_for_voucher(sre)

                target_doc.append("items", dn_item)
            else:
                # Correct rows index.
                for idx, item in enumerate(target_doc.items):
                    item.idx = idx + 1

    # Should be called after mapping items.
    set_missing_values(so, target_doc)

    return target_doc



# @frappe.whitelist()
# def dn_coc_Print_purchase_order_value_update():



#    so_item_doc = frappe.get_doc("Sales Order Item", "n00ciiodi0")
#    print(so_item_doc.item_code)
#    print(so_item_doc.qty)
#    print(so_item_doc.rate)
#    print(so_item_doc.amount)
#    print(so_item_doc.net_amount)


@frappe.whitelist()
def get_so_amount(docname):
    doc = frappe.get_doc("Delivery Note", docname)
    so_amount = 0
    for item in doc.items:
        amount = frappe.db.get_value("Sales Order Item", item.so_detail, "base_amount") or 0
        so_amount += amount
    return so_amount



@frappe.whitelist()
def del_att():
    doc = frappe.db.get_all("Attendance",{'attendance_date':('between',('2025-06-21','2025-07-20'))},['name'])
    for i in doc:
        att = frappe.get_doc("Attendance", i.name)
        if att.docstatus==1:
            att.cancel()
        att.delete()
        frappe.db.commit()

@frappe.whitelist()
def warehouse_incharge_mail(doc,method):
    if doc.stock_entry_type == 'Material Transfer' and doc.from_warehouse and doc.to_warehouse:
        item_table = """
        <table style="border-collapse: collapse; width: 100%;">
            <thead>
                <tr>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">S.No</th>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">Source Warehouse</th>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">Target Warehouse</th>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">Item Code</th>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">Item Name</th>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">Qty</th>
                    <th style="border: 1px solid black; padding: 5px; background-color: #003366; color: white;">UOM</th>
                </tr>
            </thead>
            <tbody>
        """
        idx = 1
        for row in doc.items:
            item_table += f"""
                <tr>
                    <td style="border: 1px solid black; padding: 5px;">{idx}</td>
                    <td style="border: 1px solid black; padding: 5px;">{row.s_warehouse or ''}</td>
                    <td style="border: 1px solid black; padding: 5px;">{row.t_warehouse or ''}</td>
                    <td style="border: 1px solid black; padding: 5px;">{row.item_code or ''}</td>
                    <td style="border: 1px solid black; padding: 5px;">{row.item_name or ''}</td>
                    <td style="border: 1px solid black; padding: 5px;">{row.qty or 0}</td>
                    <td style="border: 1px solid black; padding: 5px;">{row.uom or ''}</td>
                </tr>
            """
            idx += 1

        item_table += "</tbody></table>"



        from_wh = frappe.get_doc("Warehouse",doc.from_warehouse)
        email_from = []
        for row in from_wh.custom_warehouse_incharge:
            if row.user_id:
                email_from.append(row.user_id)
        frappe.sendmail(
            recipients=email_from,
            subject="Material Transfer",
            message=f"""
                Dear Sir/Mam,<br><br>
                Kindly find attached the list of items transferred from <b> {doc.from_warehouse} </b> to <b>{doc.to_warehouse}</b>.<br><br>
                {item_table}
                <br><br>
            """
        )



        to_wh= frappe.get_doc("Warehouse", doc.to_warehouse)
        email_to = []
        for row in to_wh.custom_warehouse_incharge:
            if row.user_id:
                email_to.append(row.user_id)
        frappe.sendmail(
            recipients=email_to,
            subject="Material Transfer",
            message=f"""
                Dear Sir/Mam,<br><br>
                Kindly find attached the list of items transferred from <b> {doc.from_warehouse} </b> to <b>{doc.to_warehouse}</b>.<br><br>
                {item_table}
                <br><br>
            """
        )






def set_app(doc, method):
    if not doc.default_app:
        doc.default_app = "erpnext"



# @frappe.whitelist()
# def name_qtn(doc, method):
#     if not doc.amended_from:
#         count = frappe.db.count("Quotation",{'company': doc.company, 'docstatus': ['!=', 2]})
#         count += 1

#         if doc.company != "AMAL PETROLEUM SERVICES CO.":
#             s_code = frappe.db.get_value('Company', {'name': doc.company}, ['abbr'])
#         else:
#             s_code = "APS"

#         year = datetime.today().strftime("%Y")
#         receipt_number = f'QTN-{s_code}-{year}-{str(count).zfill(5)}'
#         doc.custom_quotation_id = receipt_number

import frappe
from datetime import datetime

@frappe.whitelist()
def name_qtn(doc, method):
    if not doc.amended_from:
        count = frappe.db.count("Quotation",{'company': doc.company,'amended_from':["is","not set"]})
        count += 1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])

        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f"QTN-{s_code}-{year}-{str(next_num).zfill(5)}"

            exists = frappe.db.exists("Quotation", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_quotation_id = receipt_number


import frappe
from datetime import datetime
@frappe.whitelist()
def name_si(doc,method):
    if not doc.amended_from:
        count=frappe.db.count("Sales Invoice",{'company':doc.company, 'amended_from':["is","not set"]})
        count+=1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f'SI-{s_code}-{year}-{str(next_num).zfill(5)}'
            exists = frappe.db.exists("Sales Invoice", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_sales_invoice_id=receipt_number



import frappe
from datetime import datetime

@frappe.whitelist()
def name_so(doc,method):
    if not doc.amended_from:
        count=frappe.db.count("Sales Order",{'company':doc.company, 'amended_from':["is","not set"]})
        count+=1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f'SO-{s_code}-{year}-{str(next_num).zfill(5)}'
            exists = frappe.db.exists("Sales Order", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_sales_order_id=receipt_number



import frappe
from datetime import datetime

@frappe.whitelist()
def name_pr(doc,method):
    if not doc.amended_from:
        count=frappe.db.count("Purchase Receipt",{'company':doc.company, 'amended_from':["is","not set"]})
        count+=1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f'PR-{s_code}-{year}-{str(next_num).zfill(5)}'
            exists = frappe.db.exists("Purchase Receipt", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_purchase_receipt_id=receipt_number


import frappe
from datetime import datetime

@frappe.whitelist()
def name_po(doc,method):
    if not doc.amended_from:
        count=frappe.db.count("Purchase Order",{'company':doc.company, 'amended_from':["is","not set"]})
        count+=1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f'PO-{s_code}-{year}-{str(next_num).zfill(5)}'
            exists = frappe.db.exists("Purchase Order", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_purchase_order_id=receipt_number



import frappe
from datetime import datetime

@frappe.whitelist()
def name_pi(doc,method):
    if not doc.amended_from:
        count=frappe.db.count("Purchase Invoice",{'company':doc.company, 'amended_from':["is","not set"]})
        count+=1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f'PI-{s_code}-{year}-{str(next_num).zfill(5)}'
            exists = frappe.db.exists("Purchase Invoice", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_purchase_invoice_id=receipt_number




import frappe
from datetime import datetime

@frappe.whitelist()
def name_mr(doc, method):
    if not doc.amended_from:
        count = frappe.db.count("Material Request", {
            "company": doc.company,
            'amended_from':["is","not set"]

        })
        count += 1


        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")

        next_num = count
        while True:
            receipt_number = f"MR-{s_code}-{year}-{str(next_num).zfill(5)}"
            exists = frappe.db.exists("Material Request", {"name": receipt_number})
            if not exists:
                break
            next_num += 1

        doc.custom_material_request_id = receipt_number


import frappe
from datetime import datetime

@frappe.whitelist()
def name_opp(doc, method):
    if not doc.amended_from:
        count = frappe.db.count("Opportunity", {
            "company": doc.company,
            'amended_from':["is","not set"]
        })
        count += 1

        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])

        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f"OPP-{s_code}-{year}-{str(next_num).zfill(5)}"
            exists = frappe.db.exists("Opportunity", {"name": receipt_number})
            if not exists:
                break
            next_num += 1
        doc.custom_opportunity_id = receipt_number


import frappe
from datetime import datetime

@frappe.whitelist()
def name_dn(doc,method):
    if not doc.amended_from:
        count=frappe.db.count("Delivery Note",{'company':doc.company, 'amended_from':["is","not set"]})
        count+=1
        if doc.company == "AMAL PETROLEUM SERVICES CO.":
            s_code = "APS"
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            s_code = "ALP"
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            s_code = "TOU"
        else:
            s_code=frappe.db.get_value('Company',{'name':doc.company},['abbr'])
        year = datetime.today().strftime("%Y")
        next_num = count
        while True:
            receipt_number = f'DN-{s_code}-{year}-{str(next_num).zfill(5)}'
            exists = frappe.db.exists("Delivery Note", {"name": receipt_number})
            if not exists:
                break

            next_num += 1
        doc.custom_delivery_note_id=receipt_number


# @frappe.whitelist()
# def update_naming_series(doc, method):
#     if not doc.is_new():
#         return
#     if not doc.amended_from:
#         last_doc = frappe.get_all("Purchase Order",{"company": doc.company, "docstatus": ["!=", 2], "name": ["!=", doc.name]},["name"],order_by="name desc", limit_page_length=1)
#         last_seq = 0
#         if last_doc:
#             match = re.search(r"(\d{5})$", last_doc[0].name)
#             if match:
#                 last_seq = int(match.group(1))
#         next_seq = last_seq + 1
#         if doc.company != "AMAL PETROLEUM SERVICES CO.":
#             s_code = frappe.db.get_value("Company", {"name": doc.company}, "abbr")
#         else:
#             s_code = "APS"
#         year = datetime.today().strftime("%Y")
#         doc.name = f"PO-{s_code}-{year}-{str(next_seq).zfill(5)}"
#     else:
#         base = doc.amended_from
#         amend_count = frappe.db.count("Purchase Order", {"amended_from": base}) + 1
#         doc.name = f"{base}-{amend_count}"


@frappe.whitelist()
def update_previous_claim(doc, method):
    if doc.customer == 'Oman Electricity Transmission Company' and doc.custom_invoice_type:
        sales_order=''
        for item in doc.items:
            sales_order=item.sales_order
        if sales_order:
            so=frappe.get_doc('Sales Order',sales_order)
            for s in so.items:
                for si in doc.items:
                    if s.item_code==si.item_code:
                        s.custom_cumulative_amount+=si.amount
            so.save(ignore_permissions=True)

@frappe.whitelist()
def deduct_previous_claim(doc, method):
    if doc.customer == 'Oman Electricity Transmission Company' and doc.custom_invoice_type:
        sales_order=''
        for item in doc.items:
            sales_order=item.sales_order
            
        if sales_order:
            so=frappe.get_doc('Sales Order',sales_order)
            for s in so.items:
                for si in doc.items:
                    if s.item_code==si.item_code:
                        s.custom_cumulative_amount-=si.amount
            so.save(ignore_permissions=True)




# @frappe.whitelist()
# def update_po_and_mr_workflows():

#     docs = frappe.db.get_all(
#         "Leave Application",
#         filters={
#             "workflow_state": "Pending for Level 2 Leave Approver",

#         },
#         fields=["name"],
#         order_by="modified desc"
# )
#     for doc in docs:
#         print(doc)
#         frappe.db.set_value("Leave Application", doc.name, "workflow_state", 'Pending for General Manager')
@frappe.whitelist()
def update_employee_doc_name(doc, method):
    if not doc.employee_number and doc.company and doc.custom_nationality:
        code = ''
        if doc.company == "SHAHER UNITED TRADING & CONTRACTING COMPANY":
            code = 'SUOM' if doc.custom_nationality == 'Oman' else 'SU'
        elif doc.company == "AMAL PETROLEUM SERVICES CO.":
            code = 'APOM' if doc.custom_nationality == 'Oman' else 'AP'
        elif doc.company == "ALPHA ENGINEERING & CONTRACTING LLC":
            code = 'ALPOM' if doc.custom_nationality == 'Oman' else 'ALP'
        elif doc.company == "TAKE OFF UNITED TRADING LLC":
            code = 'TOUOM' if doc.custom_nationality == 'Oman' else 'TOU'
        elif doc.company == "THE PALACE HOTEL":
            code = 'TPHOM' if doc.custom_nationality == 'Oman' else 'TPH'

        if code:
            employees = frappe.db.sql(
                """SELECT name FROM `tabEmployee` WHERE name LIKE %s""",
                (f"{code}%",),
                as_dict=True
            )

            max_number = 0
            if employees:
                for emp in employees:
                    emp_num = emp['name']
                    if emp_num.startswith(code):
                        numeric_part = emp_num[len(code):]
                        if numeric_part.isdigit():
                            num = int(numeric_part)
                            if num > max_number:
                                max_number = num

                next_number = max_number + 1
                doc.employee_number = f"{code}{next_number:03d}"

            else:
                next_number = max_number + 1
                doc.employee_number = f"{code}{next_number}"
@frappe.whitelist()
def get_workflow_states(doctype):
    workflows = frappe.get_all(
        "Workflow",
        filters={"document_type": doctype,"is_active":1},
        pluck="name"
    )
    if not workflows:
        return []

    states = frappe.get_all(
        "Workflow Document State",
        filters={"parent": ["in", workflows],'state':["not in",['Approved','Rejected','Cancelled','Draft']]},
        pluck="state"
    )

    unique_states = sorted(set(states))
    return unique_states

@frappe.whitelist()
def get_workflow_roles(doctype):
    workflows = frappe.get_all(
        "Workflow",
        filters={"document_type": doctype, "is_active": 1},
        pluck="name"
    )
    if not workflows:
        return []

    roles = frappe.get_all(
        "Workflow Document State",
        filters={"parent": ["in", workflows]},
        pluck="allow_edit" 
    )

    unique_roles = sorted(set(roles))
    return unique_roles
    
from frappe import render_template
@frappe.whitelist()
def create_workflow_notification(doctype, name, status,project):
    if project and status:
        filters = {
            "document_type": doctype,
            "workflow_state": status,
            "parent":project,
        }
        notif = frappe.db.get_value(
            "Workflow Notification",
            filters,
            ["subject", "receiver", "message", "receiver_by_role"],
            as_dict=True,
        )
        if notif and (notif.receiver or notif.receiver_by_role):
            recipients = []

            if notif.receiver:
                recipients.extend(
                    [e.strip() for e in re.split(r"[\n,]+", notif.receiver) if e.strip()]
                )

            if notif.receiver_by_role:
                role_users = frappe.get_all(
                    "Has Role",
                    filters={"role": notif.receiver_by_role},
                    fields=["parent"], 
                )
                for ru in role_users:
                    email = frappe.db.get_value("User", {"name": ru.parent, "enabled": 1}, "email")
                    if email:
                        recipients.append(email)

            recipients = list(set(recipients))

            if not recipients:
                return

            try:
                if notif.subject:
                    subject = render_template(notif.subject, {"doc": doc})
                else:
                    subject = f"{doc.doctype} Notification",
                if notif.message:
                    message = render_template(notif.message, {"doc": doc})
                else:
                    message = f"""
                        <p>
                            Dear Sir/Madam,<br>
                            Kindly find the below {doc.doctype}:<br>
                            <a href='/app/{frappe.scrub(doc.doctype)}/{doc.name}'>{doc.name}</a><br><br>
                            Thank you & Regards,<br>
                        </p>
                        """
                frappe.sendmail(
                    recipients=recipients,
                    subject=subject,
                    message=message
                )
            except Exception:
                frappe.log_error(frappe.get_traceback(), "Workflow Notification Error")




@frappe.whitelist()
def update_leaves_in_annual_leave_allocation():
    employees = frappe.db.get_all("Employee", {"status": "Active"},["name"])
    for emp in employees:
        leave_allocation = frappe.get_all("Leave Allocation",{"employee": emp.name,"leave_type": "Annual Leave","docstatus": ("!=", 2) },["name"])
        if leave_allocation:
            print(emp)
            allocation_doc = frappe.get_doc("Leave Allocation", leave_allocation[0].name)
            allocation_doc.new_leaves_allocated += 2.5
            allocation_doc.total_leaves_allocated +=2.5
            allocation_doc.save()
            if allocation_doc.docstatus == 0:
                allocation_doc.submit()

@frappe.whitelist()
def update_employee_doc_course(doc, method):
    if doc.workflow_state=='Approved':
        if doc.custom_course_confirmation:
            for j in doc.custom_course_confirmation:
                if frappe.db.exists('HSE', {"employee": j.employee}):
                    hse_doc = frappe.get_doc('HSE', {"employee": j.employee})
                    hse_doc.append("hse_training_matrix", {
                        "course_name": j.course,
                        "course_group": j.course_group,
                        "expiry_date": j.expiry_date,
                        "validity": j.validity,
                        "po_number":doc.name
                    })
                    hse_doc.save(ignore_permissions=True)
                else:
                    hse_doc =frappe.new_doc('HSE')
                    hse_doc.employee= j.employee
                    hse_doc.append("hse_training_matrix", {
                        "course_name": j.course,
                        "course_group": j.course_group,
                        "expiry_date": j.expiry_date,
                        "validity": j.validity,
                        "po_number":doc.name
                    })
                    hse_doc.save(ignore_permissions=True)

                employee_doc = frappe.get_doc('Employee', j.employee)
                employee_doc.append("custom_hse_training_matrix", {
                    "course_name": j.course,
                    "course_group": j.course_group,
                    "expiry_date": j.expiry_date,
                    "validity": j.validity,
                    "po_number":doc.name
                })
                employee_doc.save(ignore_permissions=True)

@frappe.whitelist()
def update_employee_doc_course_on_cancel(doc, method):
    if doc.name !="PO-SUTC-2025-00121":
        if doc.custom_course_confirmation:
            for j in doc.custom_course_confirmation:
                if frappe.db.exists('HSE', {"employee": j.employee}):
                    hse_doc = frappe.get_doc('HSE', {"employee": j.employee})
                    for row in hse_doc.hse_training_matrix:
                        if row.po_number == doc.name:
                            hse_doc.hse_training_matrix.remove(row)
                            break
                    hse_doc.save(ignore_permissions=True)

                employee_doc = frappe.get_doc('Employee', j.employee)

                for row in employee_doc.custom_hse_training_matrix:
                    if row.po_number == doc.name:
                        employee_doc.custom_hse_training_matrix.remove(row)
                        break
                employee_doc.save(ignore_permissions=True)



@frappe.whitelist()
def create_scheduled_job():
    pos = frappe.db.exists('Scheduled Job Type', 'update_leaves_in_annual_leave_allocation')
    if not pos:
        sjt = frappe.new_doc("Scheduled Job Type")
        sjt.update({
            "method": 'shaher.custom.update_leaves_in_annual_leave_allocation',
            "frequency": 'Cron',
            "cron_format": '0 0 1 * *'
        })
        sjt.save(ignore_permissions=True)



@frappe.whitelist()
def update_gratuity_days_in_employee_doc():
    employees = frappe.db.get_all("Employee",{"status": "Active", "custom_nationality":['not in', "Oman"]},["name", "date_of_joining"])
    for emp in employees:
        if emp.date_of_joining:
            date_join = getdate(emp.date_of_joining)
            date_to = getdate(nowdate())
            gratuity_days = (date_to - date_join).days
        else:
            gratuity_days = 0
        frappe.db.set_value("Employee", emp.name, "custom_gratuity_days", gratuity_days)
    frappe.db.commit()

@frappe.whitelist()
def create_scheduled_job():
    pos = frappe.db.exists('Scheduled Job Type', 'send_alert_notifications')
    if not pos:
        sjt = frappe.new_doc("Scheduled Job Type")
        sjt.update({
            "method": 'shaher.alerts.send_alert_notifications',
            "frequency": 'Cron',
            "cron_format": '0 0 * * *'
        })
        sjt.save(ignore_permissions=True)


@frappe.whitelist()
def set_description_in_long_text_po(doc, method):
    if doc.items:
        for row in doc.items:
            if hasattr(row, "description") and row.description:
                if row.description:
                    clean_desc = row.description.replace('<br>', '\n')\
                                            .replace('<br/>', '\n')\
                                            .replace('<br />', '\n')
                    clean_desc = clean_desc.replace('</p>', '\n').replace('</div>', '\n')
                    clean_desc = strip_html(clean_desc)
                    lines = [line.rstrip() for line in clean_desc.splitlines()]
                    row.custom_long_description = "\n".join(lines)
           

import re
import frappe
from bs4 import BeautifulSoup

@frappe.whitelist()
def update_existing_pos_set_description_in_long_text():
    pos_list = frappe.get_all(
        "Purchase Order",
        {"docstatus": 0},
        ["name"]
    )

    for po in pos_list:
        if po.name == "PO-SUTC-2025-00121":
            doc = frappe.get_doc("Purchase Order", po.name)

            if doc.items:
                for row in doc.items:
                    description = frappe.db.get_value(
                        "Purchase Order Item",
                        {
                            "parent": doc.name,
                            "parenttype": "Purchase Order",
                            "idx": row.idx,
                            "item_code": row.item_code,
                        },
                        "description",
                    ) or ""

                    if description:
                        print(split_html_preserve_tags(description,47))
                        clean_desc = (
                            description.replace("<br>", "\n")
                            .replace("<br/>", "\n")
                            .replace("<br />", "\n")
                            .replace("</p>", "\n")
                            .replace("</div>", "\n")
                        )

                        # Parse with BeautifulSoup
                        soup = BeautifulSoup(clean_desc, "html.parser")

                        # Keep only formatting tags
                        allowed_tags = ["b", "strong", "ul", "ol", "li", "br", "p"]
                        for tag in soup.find_all(True):
                            if tag.name not in allowed_tags:
                                tag.unwrap()

                        # Convert back to HTML (preserves bold, lists, etc.)
                        clean_desc = soup.decode()

                        # Clean unnecessary multiple newlines
                        clean_desc = re.sub(r"\n\s*\n", "\n", clean_desc).strip()

                        # Save into custom long description
                        row.custom_long_description = clean_desc

            doc.save(ignore_permissions=True)
            frappe.db.commit()
            return "Descriptions updated successfully"

from bs4 import BeautifulSoup
import textwrap
import copy

def split_html_preserve_tags(html_content, max_len=47):
    soup = BeautifulSoup(html_content or "", "html.parser")
    lines = []

    def split_text_with_tags(tag):
        """
        Split tag's content into chunks of max_len while preserving HTML tags.
        Returns a list of HTML strings.
        """
        result = []
        buffer = ""
        for child in tag.contents:
            html_str = str(child)
            text_len = len(child.get_text()) if hasattr(child, 'get_text') else len(str(child))
            if len(buffer) + text_len <= max_len:
                buffer += html_str
            else:
                if buffer.strip():  # Skip empty buffers
                    new_tag = soup.new_tag(tag.name)
                    new_tag.append(BeautifulSoup(buffer, "html.parser"))
                    result.append(str(new_tag))
                # If child is bigger than max_len, split recursively
                if hasattr(child, 'contents') and child.contents:
                    result.extend(split_text_with_tags(child))
                else:
                    if html_str.strip():  # Skip empty strings
                        new_tag = soup.new_tag(tag.name)
                        new_tag.append(BeautifulSoup(html_str, "html.parser"))
                        result.append(str(new_tag))
                buffer = ""
        if buffer.strip():  # Final buffer check
            new_tag = soup.new_tag(tag.name)
            new_tag.append(BeautifulSoup(buffer, "html.parser"))
            result.append(str(new_tag))
        return result

    # Process <p>, <td> individually
    for tag in soup.find_all(['p', 'td']):
        lines.extend(split_text_with_tags(tag))

    # Process lists
    for lst in soup.find_all(['ol', 'ul']):
        new_list = soup.new_tag(lst.name)
        for li in lst.find_all('li', recursive=False):
            for chunk in split_text_with_tags(li):
                if BeautifulSoup(chunk, "html.parser").get_text(strip=True):  # Skip empty chunks
                    new_li = soup.new_tag('li')
                    new_li.append(BeautifulSoup(chunk, "html.parser"))
                    new_list.append(new_li)
        if new_list.find_all('li'):  # Only add lists with content
            lines.append(str(new_list))

    # Fallback if empty
    if not lines:
        raw_text = soup.get_text(" ", strip=True)
        for i in range(0, len(raw_text), max_len):
            chunk = raw_text[i:i+max_len].strip()
            if chunk:
                lines.append(f"<p>{chunk}</p>")

    return lines



@frappe.whitelist()
def get_item_rates(doc, month, year):
    if frappe.db.exists(doc, {'month': month, 'year': year, 'docstatus': 1}):
        invoice_doc = frappe.get_doc(doc, {'month': month, 'year': year, 'docstatus': 1})
        item_list = []
        if doc in ['MIS Opex','DHOFAR OPEX']:
            item_list = [
                ["4.1. Permanent Manpower", invoice_doc.manpower_total],
                ["4.3. Permanent Vehicle Rate", invoice_doc.vehicle_total],
                ["4.7. Accommodation, Stationary, Water Supply & Consumables materials", invoice_doc.accommodation_total]
            ]
        elif doc in ['MIS Capex','DHOFAR CAPEX']:
            item_list = [
                ["4.1. Permanent Manpower", invoice_doc.manpower_total],
                ["4.3. Permanent Vehicle Rate", invoice_doc.vehicle_total]
            ]
        elif doc in ['MIS OT','DHOFAR OT']:
            item_list = [
                ["Over Time", invoice_doc.overtime_total]
            ]
        elif doc == 'MIS ADDITIONAL':
            item_list = [
                ["4.2. Need Base Manpower Rate", invoice_doc.manpower_total],
                ["4.4. NEED BASE VEHICLE/ MACHINES COST (The cost should include the driver’s cost.)", invoice_doc.vehicle_total],
                ["4.5. NEED BASE VEHICLE COST PER TRIP", invoice_doc.vehicle_cost_per_trip_total],
                ["4.10. AIRCONDITIONING SYSTEM - SPARE PARTS", invoice_doc.ac_spare_parts_total],
                ["4.11. FIRE FIGHTING SYSTEM SPARE PARTS INCLUDES FIRE ALARM SYSTEM & GENERAL SPARES", invoice_doc.fire_fighting_parts_total],
                ["Over Time", invoice_doc.ot_total],
                ["Additional Manpower", invoice_doc.additional_manpower_total],
                ["Additional works at MIS", invoice_doc.additional_work_total]
            ]
        else:
            item_list = [
                ["4.2. Need Base Manpower Rate", invoice_doc.manpower_total],
                ["4.4. NEED BASE VEHICLE/ MACHINES COST (The cost should include the driver’s cost.)", invoice_doc.vehicle_total],
                ["4.5. NEED BASE VEHICLE COST PER TRIP", invoice_doc.vehicle_cost_per_trip_total],
                ["4.10. AIRCONDITIONING SYSTEM - SPARE PARTS", invoice_doc.ac_spare_parts_total],
                ["4.11. FIRE FIGHTING SYSTEM SPARE PARTS INCLUDES FIRE ALARM SYSTEM & GENERAL SPARES", invoice_doc.fire_fighting_parts_total],
                ["Additional works at MIS", invoice_doc.additional_work_total]
            ]
        return item_list
    else:
        frappe.throw('Document is not created, Kindly create the respective document to raise invoice')
import re
import textwrap
import frappe
from bs4 import BeautifulSoup, NavigableString

@frappe.whitelist()
def update_existing_pos_description_text_editor_to_text_editor():
    pos_list = frappe.get_all(
        "Purchase Order",
        {"docstatus": 0},
        ["name"]
    )

    for po in pos_list:
        if po.name == "PO-SUTC-2025-00138":
            doc = frappe.get_doc("Purchase Order", po.name)

            if doc.items:
                for row in doc.items:
                    description = frappe.db.get_value(
                        "Purchase Order Item",
                        {
                            "parent": doc.name,
                            "parenttype": "Purchase Order",
                            "idx": row.idx,
                            "item_code": row.item_code,
                        },
                        "description",
                    ) or ""

                    if description:
                        wrapped_html = wrap_html_text_preserve_tags(description, width=50)
                        row.custom_long_description = wrapped_html

                doc.save()
                frappe.db.commit()
import textwrap
from bs4 import BeautifulSoup, NavigableString, Tag
import textwrap
import re

@frappe.whitelist()
def update_existing_pos_description_text_editor_to_text_editor():
    pos_list = frappe.get_all(
        "Purchase Order",
        {"docstatus": 0},
        ["name"]
    )

    for po in pos_list:
        if po.name == "PO-SUTC-2025-00138":
            doc = frappe.get_doc("Purchase Order", po.name)

            if doc.items:
                for row in doc.items:
                    description = frappe.db.get_value(
                        "Purchase Order Item",
                        {
                            "parent": doc.name,
                            "parenttype": "Purchase Order",
                            "idx": row.idx,
                            "item_code": row.item_code,
                        },
                        "description",
                    ) or ""

                    if description:
                        wrapped_html = wrap_html_text_preserve_tags(description, width=50)
                        row.custom_long_description = wrapped_html

                doc.save()
                frappe.db.commit()
import textwrap
from bs4 import BeautifulSoup, NavigableString, Tag
import textwrap
import re
from bs4 import BeautifulSoup, NavigableString, Tag
import textwrap
from bs4 import BeautifulSoup, NavigableString, Tag
import re

import textwrap
from bs4 import BeautifulSoup, NavigableString, Tag
import re

def wrap_html_text_preserve_tags(html_content, width=80):
    """
    Wrap long text while preserving all HTML tags.
    Preserves <ul>/<ol>/<li> bullet points and all other HTML tags.
    Inserts <br> only inside plain text nodes for wrapping.
    """

    soup = BeautifulSoup(html_content, "html.parser")

    def wrap_text(text):
        """Wrap plain text without breaking HTML tags."""
        if not text.strip():
            return text
        # Use textwrap to wrap the text
        wrapped = textwrap.fill(
            text,
            width=width,
            break_long_words=False,
            break_on_hyphens=False
        )
        return wrapped.replace("\n", "<br>")

    def process_node(node):
        """Recursively process all children of a node."""
        for child in node.contents:
            if isinstance(child, NavigableString):
                parent = child.parent
                # Skip table, td, tr, ul, ol
                if parent.name not in ["table", "tr", "td", "ul", "ol"]:
                    new_content = wrap_text(str(child))
                    child.replace_with(BeautifulSoup(new_content, "html.parser"))
            elif isinstance(child, Tag):
                # Recursively process child tags
                process_node(child)

    # Start recursive processing
    process_node(soup)

    # Return the final HTML as string
    return str(soup)

from bs4 import BeautifulSoup

def prepare_description_for_template(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    new_lines = []

    for element in soup.contents:
        # If element is a <ul> or <ol>, preserve it as a single chunk
        if element.name in ["ul", "ol"]:
            new_lines.append(str(element))
        else:
            # Split text nodes by <br>
            text = str(element)
            new_lines.extend(text.split("<br>"))

    return new_lines

from bs4 import BeautifulSoup

def get_description_lines(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    lines = []

    for element in soup.contents:
        # Preserve <ul>, <ol>, <p>, <div> as full blocks
        if element.name in ["ul", "ol", "p", "div"]:
            lines.append(str(element))
        elif element.string and element.string.strip():
            # Split plain text by newlines or <br>
            text = str(element)
            lines.extend(text.split("<br>"))
    return lines

@frappe.whitelist()
def get_current_leave_balance(employee,date):
    details = get_leave_details(employee=employee, date=date)
    leave_allocations = details.get("leave_allocation", {})
    rows = []
    available =0
    for lt, alloc in leave_allocations.items():
        available = flt(alloc.get("remaining_leaves", 0))
    return available
import frappe
from frappe.model.document import Document

@frappe.whitelist()
def update_item_type(doc, method):
    if not doc.custom_item_type:
        return
    if doc.custom_item_type == "Stock Item":
        doc.is_stock_item = 1
        doc.is_fixed_asset = 0  
    elif doc.custom_item_type == "Service Item":
        doc.is_stock_item = 0
        doc.is_fixed_asset = 0
    elif doc.custom_item_type == "Asset Item":
        doc.is_stock_item = 0
        doc.is_fixed_asset = 1


@frappe.whitelist()
def download_coa_excel(company, visible_tree=None):
    """Download Chart of Accounts Excel showing visible tree if available, else root accounts only, with summary balances"""

    # Parse visible tree if provided
    visible_accounts = []
    if visible_tree:
        try:
            tree_data = json.loads(visible_tree)
            visible_accounts = [t["name"] for t in tree_data if t.get("name")]
        except Exception:
            visible_accounts = []

    # Fetch all accounts
    accounts = frappe.db.sql("""
        SELECT 
            name,
            account_number,
            account_name,
            is_group,
            parent_account,
            root_type,
            lft
        FROM `tabAccount`
        WHERE company = %s
        ORDER BY lft
    """, (company,), as_dict=True)

    if not accounts:
        frappe.throw(f"No accounts found for company {company}")

    # Balances
    balances = frappe._dict()
    for d in frappe.db.sql("""
        SELECT account, SUM(debit - credit) AS balance
        FROM `tabGL Entry`
        WHERE company = %s
        GROUP BY account
    """, (company,), as_dict=True):
        balances[d.account] = flt(d.balance)

    # Parent → children mapping
    children = {}
    for acc in accounts:
        children.setdefault(acc.parent_account, []).append(acc)

    # Account lookup
    acc_map = {a.name: a for a in accounts}

    # Workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"

    # Title
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{company}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([""])

    # Headers
    ws.append(["Account", "", "", "Balance"])
    ws["A3"].font = Font(bold=True)
    ws["D3"].font = Font(bold=True)
    ws["A3"].alignment = Alignment(horizontal="left")
    ws["D3"].alignment = Alignment(horizontal="right")

    # Recursive balance calculator
    def get_total_balance(acc):
        if not acc.is_group:
            return balances.get(acc.name, 0)
        total = 0
        for child in children.get(acc.name, []):
            total += get_total_balance(child)
        return total

    # Balance formatter
    def format_balance(acc):
        bal = get_total_balance(acc)
        drcr = "Dr" if bal > 0 else "Cr"
        return f"{abs(bal):,.3f} ر.ع.{drcr}"

    # Recursive writer
    def write_account(acc, level=0):
        indent = "    " * level
        icon = "📁" if acc.is_group else "⭕"
        ws.append([f"{indent}{icon} {acc.account_name}", "", "", format_balance(acc)])
        ws[f"D{ws.max_row}"].alignment = Alignment(horizontal="right")
        for child in children.get(acc.name, []):
            write_account(child, level + 1)

    # 🟩 CASE 1: visible tree provided
    if visible_accounts:
        tree_data = json.loads(visible_tree)
        for node_name in visible_accounts:
            acc = acc_map.get(node_name)
            if not acc:
                continue
            # Determine level from input
            level = next((t.get("level", 0) for t in tree_data if t.get("name") == node_name), 0)
            indent = "    " * (level - 1)
            icon = "📁" if acc.is_group else "⭕"
            ws.append([f"{indent}{icon} {acc.name}", "", "", format_balance(acc)])
            ws[f"D{ws.max_row}"].alignment = Alignment(horizontal="right")

    else:
        for root_acc in children.get(None, []):
            write_account(root_acc, level=0)

    # Formatting
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["D"].width = 25

    # Save file
    file_name = f"Chart_of_Accounts_{company.replace(' ', '_')}.xlsx"
    file_path = get_site_path("public", "files", file_name)
    wb.save(file_path)

    return {"file_url": f"/files/{file_name}"}

@frappe.whitelist()
def update_workflow_on_cancel_new():
    frappe.db.set_value('Purchase Order','PO-SUTC-2025-00109','workflow_state', "Cancelled")

@frappe.whitelist()
def check_rates_fetched(doc,method):
    if doc.customer=='Oman Electricity Transmission Company' and doc.custom_department=='OETC - SUTC' and doc.custom_rates_fetched==0:
        frappe.throw("Kindly use <b>Get Rates</b> button at top to fetch rates.")



@frappe.whitelist()
def get_po_item_qty(purchase_order, item_code, idx=None):
    filters = {
        "parent": purchase_order,
        "item_code": item_code,
    }

    if idx is not None:
        filters["idx"] = idx

    return frappe.db.get_value(
        "Purchase Order Item",
        filters,
        "qty"
    )
@frappe.whitelist()
def get_payment_terms_template(parent, parent_type):
    schedule = frappe.db.get_all(
        "Payment Schedule",
        filters={"parent": parent, "parenttype": parent_type},
        fields="*"
    )

    template = frappe.db.get_value(parent_type, parent, "payment_terms_template")

    return {"schedule": schedule, "template": template}




import frappe
@frappe.whitelist()
def send_asset_mails(doc, method):
    asset_company = doc.company
    sponsor_doc = frappe.db.get_value(
        "Visa Sponsor",
        {"name": asset_company},
        "name"
    )

    sponsor = frappe.get_doc("Visa Sponsor", sponsor_doc)
    recipients = [row.user for row in sponsor.recipient_directory if row.user]

    if recipients:
        send_email_to_recipient(recipients, doc, asset_company)


def send_email_to_recipient(email, asset_doc, company):
    subject = f"New Asset Created for {company}"
    message = f"""
        Dear Sir / Mam,<br><br>
        A new Asset <b>{asset_doc.name}</b> has been created for company <b>{company}</b>.
        <table border="1" cellpadding="6" cellspacing="0" style="border-collapse: collapse;">
            <tr style="background-color:#b81a0f;color:white;">
                <th>Asset Code</th>
                <th>Asset Name</th>
                <th>Asset Category</th>
                <th>Company</th>
                <th>Asset Owner</th>
                <th>Asset Owner Name</th>
                <th>Purchase Date</th>
                <th>Purchase Amount</th>
            </tr>
            <tr>
                <td>{asset_doc.item_code or ''}</td>
                <td>{asset_doc.asset_name or ''}</td>
                <td>{asset_doc.asset_category}</td>
                <td>{asset_doc.company or ''}</td>
                <td>{asset_doc.asset_owner or ''}</td>
                <td>{asset_doc.asset_owner_company or asset_doc.supplier or asset_doc.customer or ''}</td>
                <td>{asset_doc.purchase_date or ''}</td>
                <td>{asset_doc.gross_purchase_amount or ''}</td>
            </tr>
        </table>
    """

    frappe.sendmail(
        recipients=email,
        subject=subject,
        message=message
    )
